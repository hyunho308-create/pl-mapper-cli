from __future__ import annotations

import re
from collections import Counter
from collections.abc import Iterable

from openpyxl.utils.cell import range_boundaries

from hotel_pl_normalizer.models.compact import (
    CompactContext,
    CompactHeuristicHints,
    CompactLabel,
    CompactLayout,
    CompactRow,
    CompactRowRole,
    CompactSheetPacket,
    CompactTextCell,
    CompactValue,
    CompactWindow,
    CompactWindowRow,
    PeriodColumnHint,
)
from hotel_pl_normalizer.models.workbook import (
    CellRecord,
    MergedRange,
    WorkbookRecord,
    WorkbookRow,
    WorkbookSheet,
)

_PERIOD_WORDS = ("ytd", "mtd", "actual", "budget", "forecast", "ttm", "year to date", "month to date")
_KPI_WORDS = ("occupancy", "occ", "adr", "revpar", "margin", "percentage", "%")
_CONTROL_LABELS = {
    "account",
    "afterrange",
    "app",
    "apply to",
    "autofitcol",
    "beforerange",
    "bottom",
    "businessunit",
    "category",
    "cellkeyrange",
    "colkeyrange",
    "criteria",
    "comment",
    "datasrc",
    "department",
    "dimension",
    "dumpdatacache",
    "evdre ok",
    "evaluate in",
    "expandin",
    "expandonly",
    "formatrange",
    "format",
    "getonlyrange",
    "groupexpansion",
    "hidecolkeys",
    "hiderowkeys",
    "insert",
    "intco",
    "measures",
    "memberset",
    "no variances",
    "norefresh",
    "nosend",
    "option",
    "optionrange",
    "pagekeyrange",
    "parameter",
    "parameters",
    "pctinput",
    "queryengine",
    "querytype",
    "queryviewname",
    "range",
    "rowkeyrange",
    "rptcurrency",
    "showcomments",
    "shownullaszero",
    "sortcol",
    "sortrange",
    "sqlonly",
    "sumparent",
    "suppress",
    "suppressdatacol",
    "suppressdatarow",
    "suppressnodata",
    "time",
    "time level",
    "top",
    "use",
    "variances only",
}
_BUSINESS_LABEL_TERMS = (
    "allowance",
    "benefit",
    "cost",
    "department",
    "ebitda",
    "expense",
    "expenses",
    "fees",
    "income",
    "labor",
    "payroll",
    "profit",
    "revenue",
    "revenues",
    "sales",
    "salaries",
    "total",
    "wages",
)
_SHORT_BUSINESS_LABELS = {
    "a&g",
    "adr",
    "ebitda",
    "f&b",
    "gop",
    "hvac",
    "it",
    "linen",
    "noi",
    "ood",
    "pom",
    "revpar",
    "sewer",
    "s&m",
    "smerf",
    "water",
}


def build_compact_sheet_packets(
    workbook: WorkbookRecord,
    *,
    sheet_names: set[str] | list[str] | None = None,
    ranges_by_sheet: dict[str, list[tuple[int, int]]] | None = None,
) -> list[CompactSheetPacket]:
    """Build compact row packets only for selected sheets or ranges."""
    included_sheet_names = set(sheet_names) if sheet_names is not None else None
    packets: list[CompactSheetPacket] = []
    for sheet in workbook.sheets:
        if included_sheet_names is not None and sheet.sheet_name not in included_sheet_names:
            continue
        ranges = ranges_by_sheet.get(sheet.sheet_name) if ranges_by_sheet else None
        if not ranges:
            packets.append(build_compact_sheet_packet(workbook, sheet.sheet_name))
            continue
        for start_row, end_row in ranges:
            packets.append(build_compact_sheet_packet(workbook, sheet.sheet_name, start_row=start_row, end_row=end_row))
    return packets


def build_compact_sheet_packet(
    workbook: WorkbookRecord,
    sheet_name: str,
    *,
    start_row: int | None = None,
    end_row: int | None = None,
) -> CompactSheetPacket:
    sheet = _find_sheet(workbook, sheet_name)
    rows = _filter_rows(sheet.rows, start_row=start_row, end_row=end_row)
    compact_rows = _build_sheet_compact_rows(workbook.workbook_id, sheet, rows)
    return CompactSheetPacket(
        packet_id=_packet_id(workbook.workbook_id, sheet_name, start_row, end_row),
        workbook_id=workbook.workbook_id,
        sheet_name=sheet_name,
        start_row=start_row,
        end_row=end_row,
        rows=compact_rows,
    )


def _build_sheet_compact_rows(
    workbook_id: str,
    sheet: WorkbookSheet,
    source_rows: list[WorkbookRow],
) -> list[CompactRow]:
    row_profiles = [_RowProfile(row) for row in source_rows]
    labels = [_label_from_profile(profile) for profile in row_profiles]
    title_guess = _sheet_title_guess(labels)
    column_headers = _column_header_lookup(sheet, [_RowProfile(row) for row in sheet.rows])
    merged_ranges = [_ParsedMergedRange.from_model(merged) for merged in sheet.merged_ranges]

    compact_rows: list[CompactRow] = []
    nearest_heading_above: str | None = None
    nearest_bold_heading_above: str | None = None

    for idx, (row, profile, label) in enumerate(zip(source_rows, row_profiles, labels, strict=True)):
        layout = _layout_from_profile(profile, row_profiles, idx)
        hints = _heuristic_hints(label.raw, profile, layout)
        context = CompactContext(
            nearest_heading_above=nearest_heading_above,
            nearest_bold_heading_above=nearest_bold_heading_above,
            active_merged_header=_active_merged_header(row.row_index, label.source_column, merged_ranges),
            previous_nonblank_label=_previous_nonblank_label(labels, idx),
            next_nonblank_label=_next_nonblank_label(labels, idx),
            sheet_title_guess=title_guess,
        )

        compact_rows.append(
            CompactRow(
                workbook_id=workbook_id,
                sheet_name=sheet.sheet_name,
                row_index=row.row_index,
                row_key=f"{sheet.sheet_name}!{row.row_index}",
                label=label,
                values=[_compact_value(cell, row.row_index, column_headers, merged_ranges) for cell in profile.numeric_cells],
                text_cells=[
                    CompactTextCell(column=cell.column, address=cell.address, text=_clean_text(cell.display_value or ""))
                    for cell in profile.text_cells
                ],
                layout=layout,
                context=context,
                heuristic_hints=hints,
                raw_window=_raw_window(labels, idx),
            )
        )

        if _looks_like_heading(label.raw, profile, layout, hints):
            nearest_heading_above = label.raw
            if layout.bold:
                nearest_bold_heading_above = label.raw

    return compact_rows


class _RowProfile:
    def __init__(self, row: WorkbookRow) -> None:
        self.row = row
        self.text_cells = [cell for cell in row.cells if _is_text_cell(cell)]
        self.numeric_cells = [cell for cell in row.cells if _is_numeric_cell(cell)]
        self.nonblank_cells = [cell for cell in row.cells if _is_nonblank_cell(cell)]


class _ParsedMergedRange:
    def __init__(self, min_row: int, min_col: int, max_row: int, max_col: int, value: str | None) -> None:
        self.min_row = min_row
        self.min_col = min_col
        self.max_row = max_row
        self.max_col = max_col
        self.value = value

    @classmethod
    def from_model(cls, merged: MergedRange) -> "_ParsedMergedRange":
        min_col, min_row, max_col, max_row = range_boundaries(merged.range)
        return cls(
            min_row=min_row,
            min_col=min_col,
            max_row=max_row,
            max_col=max_col,
            value=_clean_text(str(merged.value)) if merged.value is not None else None,
        )

    def contains_column(self, column: int) -> bool:
        return self.min_col <= column <= self.max_col

    def contains_cell(self, row: int, column: int) -> bool:
        return self.min_row <= row <= self.max_row and self.contains_column(column)

    def is_above(self, row: int, column: int) -> bool:
        return self.max_row < row and self.contains_column(column)


def _find_sheet(workbook: WorkbookRecord, sheet_name: str) -> WorkbookSheet:
    for sheet in workbook.sheets:
        if sheet.sheet_name == sheet_name:
            return sheet
    raise ValueError(f"Workbook does not contain sheet: {sheet_name}")


def _filter_rows(rows: list[WorkbookRow], *, start_row: int | None, end_row: int | None) -> list[WorkbookRow]:
    return [
        row
        for row in rows
        if (start_row is None or row.row_index >= start_row) and (end_row is None or row.row_index <= end_row)
    ]


def _packet_id(workbook_id: str, sheet_name: str, start_row: int | None, end_row: int | None) -> str:
    safe_sheet = re.sub(r"[^a-zA-Z0-9]+", "_", sheet_name).strip("_").lower()
    if start_row is None and end_row is None:
        return f"{workbook_id}:{safe_sheet}"
    return f"{workbook_id}:{safe_sheet}:{start_row or ''}-{end_row or ''}"


def _label_from_profile(profile: _RowProfile) -> CompactLabel:
    if not profile.text_cells:
        return CompactLabel()
    cell = label_cell(profile.text_cells)
    if cell is None:
        return CompactLabel(leftmost_text_column=profile.text_cells[0].column)
    raw = _clean_text(cell.display_value or "")
    return CompactLabel(
        raw=raw,
        normalized=_normalize_label(raw),
        source_column=cell.column,
        leftmost_text_column=profile.text_cells[0].column,
    )


def label_cell(
    text_cells: list[CellRecord],
    *,
    preferred_column: int | None = None,
) -> CellRecord | None:
    candidates = [
        (score, cell)
        for cell in text_cells
        if (score := _row_label_score(_clean_text(cell.display_value or ""))) > 0
    ]
    if not candidates:
        return None
    if preferred_column is not None:
        return next(
            (cell for _, cell in candidates if cell.column == preferred_column),
            None,
        )
    return max(candidates, key=lambda item: (item[0], item[1].column))[1]


def dominant_label_column(rows: Iterable[WorkbookRow]) -> int | None:
    """Infer the stable account-label column from rows containing numeric data."""
    counts: Counter[int] = Counter()
    scores: Counter[int] = Counter()
    for row in rows:
        if not any(_is_numeric_cell(cell) for cell in row.cells):
            continue
        for cell in row.cells:
            if not _is_text_cell(cell):
                continue
            score = _row_label_score(_clean_text(cell.display_value or ""))
            if score > 0:
                counts[cell.column] += 1
                scores[cell.column] += score
    if not counts:
        return None
    return max(counts, key=lambda column: (counts[column], scores[column], -column))


def _looks_like_row_label_text(text: str) -> bool:
    if not text:
        return False
    if not re.search(r"[A-Za-z]", text):
        return False
    if re.fullmatch(r"[-+]?\d+(?:\.\d+)?%?", text):
        return False
    return True


def _row_label_score(text: str) -> int:
    if not _looks_like_row_label_text(text):
        return 0
    normalized = _normalize_label(text) or ""
    if not normalized:
        return 0
    if normalized in _CONTROL_LABELS:
        return 0
    if _looks_like_control_or_member_code(text, normalized):
        return 0

    words = re.findall(r"[A-Za-z][A-Za-z&'-]*", text)
    score = 1
    if len(words) >= 2:
        score += 4
    if any(char.islower() for char in text):
        score += 2
    if any(term in normalized for term in _BUSINESS_LABEL_TERMS):
        score += 3
    if text[:1].isupper() and not text.isupper():
        score += 1
    if "_" in text or "|" in text or "!" in text or "$" in text:
        score -= 4
    return max(score, 0)


def _looks_like_control_or_member_code(text: str, normalized: str) -> bool:
    if _looks_like_technical_metadata_text(text):
        return True
    if text.strip().lower() in _SHORT_BUSINESS_LABELS:
        return False
    if "&" in text:
        return False
    compact = re.sub(r"[^A-Za-z0-9]+", "", text)
    if re.fullmatch(r"[A-Z]{1,6}\d{0,4}", compact):
        return True
    if re.fullmatch(r"[A-Z]+(?:_[A-Z0-9]+)+", text):
        return True
    if re.fullmatch(r"[A-Z]\d", text):
        return True
    if normalized.startswith("account style"):
        return True
    if "!" in text and "$" in text:
        return True
    if "dep(" in text.lower():
        return True
    return False


def _looks_like_technical_metadata_text(text: str) -> bool:
    stripped = text.strip()
    if not stripped:
        return False
    normalized = re.sub(r"[^a-z0-9]+", " ", stripped.lower()).strip()
    if "fontbold" in normalized or "indentlevel" in normalized:
        return True
    if re.search(r"\[[^\]]+\]\s*\.\s*\[[^\]]+\]", stripped):
        return True
    if stripped.startswith("<<") and "[" in stripped and "]" in stripped:
        return True
    if "!" in stripped and "$" in stripped:
        return True
    if "dep(" in stripped.lower():
        return True
    return False


def _layout_from_profile(profile: _RowProfile, all_profiles: list[_RowProfile], idx: int) -> CompactLayout:
    style_cell = profile.text_cells[0] if profile.text_cells else (profile.nonblank_cells[0] if profile.nonblank_cells else None)
    row_cells = profile.nonblank_cells
    return CompactLayout(
        indent=style_cell.style.indent if style_cell is not None else None,
        bold=bool(style_cell.style.bold) if style_cell is not None else False,
        italic=bool(style_cell.style.italic) if style_cell is not None else False,
        underline=bool(style_cell.style.underline) if style_cell is not None else False,
        font_size=style_cell.style.font_size if style_cell is not None else None,
        fill_color=style_cell.style.fill_color if style_cell is not None else None,
        has_top_border=False,
        has_bottom_border=any(cell.style.border_bottom is not None for cell in row_cells),
        blank_above=idx == 0 or not all_profiles[idx - 1].nonblank_cells,
        blank_below=idx == len(all_profiles) - 1 or not all_profiles[idx + 1].nonblank_cells,
        numeric_cell_count=len(profile.numeric_cells),
        text_cell_count=len(profile.text_cells),
    )


def _compact_value(
    cell: CellRecord,
    row_index: int,
    column_headers: dict[tuple[int, int], str],
    merged_ranges: list[_ParsedMergedRange],
) -> CompactValue:
    column_header = column_headers.get((row_index, cell.column))
    merged_header = _active_merged_header(row_index, cell.column, merged_ranges)
    period_hint = _period_hint(column_header, merged_header)
    return CompactValue(
        column=cell.column,
        address=cell.address,
        raw_value=cell.raw_value,
        display_value=cell.display_value,
        numeric_value=float(cell.raw_value) if _is_numeric_value(cell.raw_value) else None,
        column_header=column_header,
        merged_header=merged_header,
        period_hint=period_hint,
    )


def _heuristic_hints(label: str | None, profile: _RowProfile, layout: CompactLayout) -> CompactHeuristicHints:
    normalized_label = _normalize_label(label)
    possible_total_line = bool(normalized_label and _is_total_label(normalized_label))
    possible_kpi = bool(normalized_label and any(word in normalized_label for word in _KPI_WORDS))
    possible_section_boundary = bool(
        label
        and profile.numeric_cells == []
        and (layout.bold or layout.blank_above or (label.isupper() and len(label) > 2))
    )

    if not profile.nonblank_cells:
        possible_role = CompactRowRole.BLANK
    elif possible_total_line and profile.numeric_cells:
        possible_role = CompactRowRole.SUBTOTAL
    elif possible_section_boundary:
        possible_role = CompactRowRole.HEADING
    elif label and profile.numeric_cells:
        possible_role = CompactRowRole.DETAIL
    else:
        possible_role = CompactRowRole.UNKNOWN

    return CompactHeuristicHints(
        possible_period_columns=_period_column_hints(profile.text_cells),
        possible_row_role=possible_role,
        possible_section_boundary=possible_section_boundary,
        possible_total_line=possible_total_line,
        possible_kpi=possible_kpi,
    )


def _period_column_hints(text_cells: Iterable[CellRecord]) -> list[PeriodColumnHint]:
    hints: list[PeriodColumnHint] = []
    for cell in text_cells:
        text = _clean_text(cell.display_value or "")
        if not text:
            continue
        normalized = text.lower()
        if any(word in normalized for word in _PERIOD_WORDS) or re.search(r"\b20\d{2}\b", normalized):
            hints.append(
                PeriodColumnHint(
                    column=cell.column,
                    header=text,
                    reason="Header text contains a period keyword or year.",
                )
            )
    return hints


def _column_header_lookup(sheet: WorkbookSheet, profiles: list[_RowProfile]) -> dict[tuple[int, int], str]:
    headers: dict[tuple[int, int], str] = {}
    latest_text_by_column: dict[int, str] = {}
    for row, profile in zip(sheet.rows, profiles, strict=True):
        for cell in profile.numeric_cells:
            if cell.column in latest_text_by_column:
                headers[(row.row_index, cell.column)] = latest_text_by_column[cell.column]
        for cell in profile.text_cells:
            text = _clean_text(cell.display_value or "")
            if text:
                latest_text_by_column[cell.column] = text
    return headers


def _active_merged_header(row: int, column: int | None, merged_ranges: list[_ParsedMergedRange]) -> str | None:
    if column is None:
        return None
    containing = [merged for merged in merged_ranges if merged.value and merged.contains_cell(row, column)]
    if containing:
        return containing[-1].value
    above = [merged for merged in merged_ranges if merged.value and merged.is_above(row, column)]
    if not above:
        return None
    return max(above, key=lambda merged: merged.max_row).value


def _period_hint(*headers: str | None) -> str | None:
    for header in headers:
        if header is None:
            continue
        normalized = header.lower()
        if any(word in normalized for word in _PERIOD_WORDS) or re.search(r"\b20\d{2}\b", normalized):
            return header
    return None


def _looks_like_heading(
    label: str | None,
    profile: _RowProfile,
    layout: CompactLayout,
    hints: CompactHeuristicHints,
) -> bool:
    return bool(label and not profile.numeric_cells and (hints.possible_section_boundary or layout.bold))


def _previous_nonblank_label(labels: list[CompactLabel], idx: int) -> str | None:
    for label in reversed(labels[:idx]):
        if label.raw:
            return label.raw
    return None


def _next_nonblank_label(labels: list[CompactLabel], idx: int) -> str | None:
    for label in labels[idx + 1 :]:
        if label.raw:
            return label.raw
    return None


def _raw_window(labels: list[CompactLabel], idx: int, size: int = 2) -> CompactWindow:
    start = max(0, idx - size)
    end = min(len(labels), idx + size + 1)
    return CompactWindow(
        rows_before=[CompactWindowRow(row_index=row_idx + 1, label=labels[row_idx].raw) for row_idx in range(start, idx)],
        rows_after=[CompactWindowRow(row_index=row_idx + 1, label=labels[row_idx].raw) for row_idx in range(idx + 1, end)],
    )


def _sheet_title_guess(labels: list[CompactLabel]) -> str | None:
    for label in labels[:5]:
        if label.raw:
            return label.raw
    return None


def _is_text_cell(cell: CellRecord) -> bool:
    value = cell.display_value
    return (
        isinstance(cell.raw_value, str)
        and value is not None
        and _clean_text(value) != ""
        and not _looks_like_technical_metadata_text(_clean_text(value))
    )


def _is_numeric_cell(cell: CellRecord) -> bool:
    return _is_numeric_value(cell.raw_value)


def _is_nonblank_cell(cell: CellRecord) -> bool:
    return cell.raw_value is not None and cell.display_value is not None and _clean_text(cell.display_value) != ""


def _is_numeric_value(value) -> bool:
    return isinstance(value, int | float) and not isinstance(value, bool)


def _is_total_label(normalized_label: str) -> bool:
    return (
        normalized_label.startswith("total ")
        or normalized_label.startswith("net ")
        or normalized_label.startswith("gross ")
        or " total " in f" {normalized_label} "
    )


def _clean_text(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def _normalize_label(label: str | None) -> str | None:
    if label is None:
        return None
    cleaned = _clean_text(label).lower()
    if not cleaned:
        return None
    return re.sub(r"[^a-z0-9%]+", " ", cleaned).strip()
