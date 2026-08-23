"""Read bounded structural slices directly from an Excel workbook.

Exploration uses this reader for sheet names, header text, merged ranges, and
coordinates. Values written to the normalized output always come from the
separately ingested `WorkbookRecord`; this reader is never an arithmetic source.

Format differences worth knowing
--------------------------------
`.xlsx`/`.xlsm` stream forward. Reading rows 900-940 rescans from the top, so a
deep read costs about 0.4 s and a shallow one about 0.02 s. Re-reading a sheet
needs a fresh worksheet object from the workbook, which is why `_xlsx_sheet`
never caches one.

`.xls` is the opposite: `on_demand=True` opens in ~0.02 s, loading one sheet
costs 0.05-0.2 s once, and every row read after that is random access and
effectively free.
"""

from __future__ import annotations

import re
import zipfile
from dataclasses import dataclass
from pathlib import Path

# A single read is capped so one tool call cannot pull a whole sheet into the
# prompt. Forty rows is comfortably more than any header block seen in the
# corpus -- the deepest was row 31 -- while leaving room to hunt.
MAX_ROWS_PER_READ = 60
# Long cell text is nearly always a note or a disclaimer, and the useful part is
# the front of it.
MAX_CELL_CHARS = 160


@dataclass(frozen=True)
class SheetSummary:
    """What is known about a sheet before reading any of it."""

    sheet_name: str
    index: int
    visible: bool
    # Declared, not measured. `.xlsx` dimensions are frequently wrong -- PPS
    # declares 889,098 cells and stores 132,726 -- so these orient a reader
    # rather than bound it.
    approx_rows: int | None
    approx_columns: int | None


@dataclass(frozen=True)
class Cell:
    coordinate: str
    value: str


def _clean(value) -> str | None:
    """Render a cell as short display text, or None when it holds nothing."""
    if value is None:
        return None
    if isinstance(value, str):
        text = " ".join(value.split())
        return text[:MAX_CELL_CHARS] or None
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)[:MAX_CELL_CHARS]


def _looks_numeric(text: str) -> bool:
    """Is this a figure, a date or a bare code rather than words?"""
    return not any(character.isalpha() for character in text)


def _column_letter(index: int) -> str:
    """1 -> A. Local so this module needs nothing from openpyxl for .xls."""
    letters = ""
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


class LazyWorkbook:
    """A workbook opened for structural questions only.

    Use as a context manager; both backends hold file handles.
    """

    def __init__(self, path: str | Path) -> None:
        self.path = Path(path)
        suffix = self.path.suffix.lower()
        if suffix in {".xlsx", ".xlsm"}:
            self._kind = "xlsx"
            import openpyxl

            self._book = openpyxl.load_workbook(
                self.path, read_only=True, data_only=True
            )
        elif suffix == ".xls":
            self._kind = "xls"
            import xlrd

            self._book = xlrd.open_workbook(str(self.path), on_demand=True)
        else:
            raise ValueError(
                f"LazyWorkbook supports .xlsx, .xlsm and .xls: {self.path.name}"
            )

    def __enter__(self) -> "LazyWorkbook":
        return self

    def __exit__(self, *_) -> None:
        self.close()

    def close(self) -> None:
        try:
            if self._kind == "xlsx":
                self._book.close()
            else:
                self._book.release_resources()
        except Exception:  # noqa: BLE001 - closing must not mask a real error
            pass

    # -- what is in here --------------------------------------------------

    def sheets(self) -> list[SheetSummary]:
        """Every sheet, in workbook order. Cheap on both formats."""
        if self._kind == "xlsx":
            summaries = []
            for index, name in enumerate(self._book.sheetnames):
                worksheet = self._book[name]
                summaries.append(
                    SheetSummary(
                        sheet_name=name,
                        index=index,
                        visible=getattr(worksheet, "sheet_state", "visible") == "visible",
                        approx_rows=worksheet.max_row,
                        approx_columns=worksheet.max_column,
                    )
                )
            return summaries

        visibility = getattr(self._book, "sheet_visibility", [])
        summaries = []
        for index, name in enumerate(self._book.sheet_names()):
            summaries.append(
                SheetSummary(
                    sheet_name=name,
                    index=index,
                    visible=index >= len(visibility) or visibility[index] == 0,
                    # Loading a sheet to measure it would defeat the point.
                    approx_rows=None,
                    approx_columns=None,
                )
            )
        return summaries

    # -- reading ----------------------------------------------------------

    def read_rows(
        self, sheet_name: str, start_row: int = 1, end_row: int | None = None
    ) -> list[list[Cell]]:
        """Populated cells in a row range, one list per row.

        Rows are 1-based and inclusive, matching what Excel shows and what every
        artifact in this project records. Empty cells are omitted rather than
        returned as blanks, because a header block is mostly empty and the gaps
        carry no information a coordinate does not already give.
        """
        if start_row < 1:
            start_row = 1
        if end_row is None:
            end_row = start_row + MAX_ROWS_PER_READ - 1
        end_row = min(end_row, start_row + MAX_ROWS_PER_READ - 1)

        if self._kind == "xlsx":
            return self._read_rows_xlsx(sheet_name, start_row, end_row)
        return self._read_rows_xls(sheet_name, start_row, end_row)

    def _read_rows_xlsx(self, sheet_name, start_row, end_row) -> list[list[Cell]]:
        # A fresh worksheet each time: read-only worksheets are forward-only, so
        # a cached one is exhausted after the first pass.
        worksheet = self._book[sheet_name]
        rows = []
        for offset, values in enumerate(
            worksheet.iter_rows(min_row=start_row, max_row=end_row, values_only=True)
        ):
            row_index = start_row + offset
            cells = [
                Cell(coordinate=f"{_column_letter(column)}{row_index}", value=text)
                for column, raw in enumerate(values, start=1)
                if (text := _clean(raw)) is not None
            ]
            rows.append(cells)
        return rows

    def _read_rows_xls(self, sheet_name, start_row, end_row) -> list[list[Cell]]:
        sheet = self._book.sheet_by_name(sheet_name)
        rows = []
        for row_index in range(start_row, min(end_row, sheet.nrows) + 1):
            values = sheet.row_values(row_index - 1)
            cells = [
                Cell(coordinate=f"{_column_letter(column)}{row_index}", value=text)
                for column, raw in enumerate(values, start=1)
                if (text := _clean(raw)) is not None
            ]
            rows.append(cells)
        return rows

    def find_text(
        self, query: str, sheet_name: str | None = None, *, limit: int = 40
    ) -> list[tuple[str, Cell]]:
        """Where a string appears, case-insensitively, as (sheet name, cell).

        The sheet name travels with the hit rather than being implied by the
        argument, because the argument is optional: a search across the whole
        workbook returns a bare `C13` that means nothing on its own.

        Bounded by `limit` and by `MAX_ROWS_PER_READ * 8` rows per sheet: this is
        for locating a header block that a fixed window missed, not for scanning
        a workbook end to end.
        """
        needle = query.strip().lower()
        if not needle:
            return []
        names = [sheet_name] if sheet_name else [s.sheet_name for s in self.sheets()]
        row_budget = MAX_ROWS_PER_READ * 8
        hits: list[tuple[str, Cell]] = []
        for name in names:
            for row in self.read_rows_scan(name, row_budget):
                for cell in row:
                    if needle in cell.value.lower():
                        hits.append((name, cell))
                        if len(hits) >= limit:
                            return hits
        return hits

    def peek(self, sheet_name: str, *, max_rows: int = 12, limit: int = 24) -> list[str]:
        """Candidate identifying text from the top of a sheet.

        Coded tab names (`0400`, `F89_RC_ROOMS`) carry no meaning on their own,
        and operators put the real title inside the sheet: `0232 - Group
        Banquet`. Without this, a session deciding from names alone skips real
        departmental schedules, which is exactly what it did on Ritz-Carlton --
        23 coded tabs marked `skip`, unread.

        Returns more candidates than a caller wants to show, because the top of
        these sheets is mostly run metadata and export boilerplate. Narrowing
        them to what actually identifies the sheet needs all the sheets at once,
        so it happens in the caller rather than here.

        Numbers are excluded: a schedule's opening rows are mostly figures, and
        a figure says nothing about what the sheet is.
        """
        found: list[str] = []
        seen: set[str] = set()
        for row in self.read_rows(sheet_name, 1, max_rows):
            for cell in row:
                text = cell.value.strip()
                if len(text) <= 2 or _looks_numeric(text) or text in seen:
                    # Repeats within a sheet are dropped here rather than by the
                    # caller: a header row of `%,C` across a dozen columns would
                    # otherwise consume the whole budget before row 2.
                    continue
                seen.add(text)
                found.append(text)
                if len(found) >= limit:
                    return found
        return found

    def read_rows_scan(self, sheet_name: str, max_rows: int):
        """Yield rows from the top of a sheet, for searching. Internal to find."""
        if self._kind == "xlsx":
            worksheet = self._book[sheet_name]
            for offset, values in enumerate(
                worksheet.iter_rows(min_row=1, max_row=max_rows, values_only=True)
            ):
                row_index = offset + 1
                yield [
                    Cell(coordinate=f"{_column_letter(column)}{row_index}", value=text)
                    for column, raw in enumerate(values, start=1)
                    if (text := _clean(raw)) is not None
                ]
            return
        sheet = self._book.sheet_by_name(sheet_name)
        for row_index in range(1, min(max_rows, sheet.nrows) + 1):
            values = sheet.row_values(row_index - 1)
            yield [
                Cell(coordinate=f"{_column_letter(column)}{row_index}", value=text)
                for column, raw in enumerate(values, start=1)
                if (text := _clean(raw)) is not None
            ]

    def merged_ranges(self, sheet_name: str) -> list[str]:
        """Merged cell refs, read from sheet XML without loading any cells.

        Period headers are very often merged across the columns they label, so a
        reader that cannot see merges misreads a band of months as one column.
        `.xls` merge info needs the sheet loaded, which is cheap there.
        """
        if self._kind == "xls":
            sheet = self._book.sheet_by_name(sheet_name)
            return [
                f"{_column_letter(c1 + 1)}{r1 + 1}:{_column_letter(c2)}{r2}"
                for r1, r2, c1, c2 in getattr(sheet, "merged_cells", [])
            ]
        index = self._book.sheetnames.index(sheet_name)
        try:
            with zipfile.ZipFile(self.path) as archive:
                xml = archive.read(f"xl/worksheets/sheet{index + 1}.xml")
        except (KeyError, OSError, zipfile.BadZipFile):
            return []
        return re.findall(r'<mergeCell ref="([^"]+)"', xml.decode("utf-8", "replace"))
