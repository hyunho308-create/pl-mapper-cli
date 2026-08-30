"""Narrow OpenPyXL fallback for invalid, nonfinancial print metadata."""

from __future__ import annotations

import io
import os
import tempfile
import xml.etree.ElementTree as ET
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl

_WORKBOOK_XML = "xl/workbook.xml"
_MAIN_NAMESPACE = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_ALLOWED_NAMES = {"_xlnm.Print_Titles", "_xlnm.Print_Area"}
_INVALID_PRINT_VALUES = {
    "#DIV/0!",
    "#N/A",
    "#NAME?",
    "#NULL!",
    "#NUM!",
    "#REF!",
    "#VALUE!",
}
_EXPECTED_ERROR_FRAGMENTS = (
    "not a valid print titles definition",
    "not a valid print area definition",
)


@dataclass(frozen=True, slots=True)
class PrintMetadataRepair:
    name: str
    value: str
    local_sheet_id: str | None

    def describe(self) -> str:
        scope = (
            f" (localSheetId={self.local_sheet_id})"
            if self.local_sheet_id is not None
            else ""
        )
        return f"{self.name}={self.value}{scope}"


@dataclass(slots=True)
class CompatibleOpenpyxlWorkbook:
    workbook: Any
    read_path: Path
    repairs: tuple[PrintMetadataRepair, ...] = ()
    _temporary_path: Path | None = None

    def close(self) -> None:
        try:
            self.workbook.close()
        finally:
            if self._temporary_path is not None:
                try:
                    self._temporary_path.unlink(missing_ok=True)
                finally:
                    self._temporary_path = None


def load_openpyxl_workbook(path: str | Path, **kwargs) -> CompatibleOpenpyxlWorkbook:
    """Load normally, repairing only invalid built-in print metadata on failure."""
    source_path = Path(path)
    try:
        workbook = openpyxl.load_workbook(source_path, **kwargs)
        return CompatibleOpenpyxlWorkbook(workbook=workbook, read_path=source_path)
    except ValueError as original_error:
        if not _is_expected_print_metadata_error(original_error):
            raise

        temporary_path, repairs = _sanitized_print_metadata_copy(source_path)
        if not repairs:
            temporary_path.unlink(missing_ok=True)
            raise
        try:
            workbook = openpyxl.load_workbook(temporary_path, **kwargs)
        except Exception:
            temporary_path.unlink(missing_ok=True)
            raise
        return CompatibleOpenpyxlWorkbook(
            workbook=workbook,
            read_path=temporary_path,
            repairs=tuple(repairs),
            _temporary_path=temporary_path,
        )


def _is_expected_print_metadata_error(error: BaseException) -> bool:
    """OpenPyXL wraps the useful print parser error in a generic ValueError."""
    seen: set[int] = set()
    current: BaseException | None = error
    while current is not None and id(current) not in seen:
        seen.add(id(current))
        message = str(current).lower()
        if any(fragment in message for fragment in _EXPECTED_ERROR_FRAGMENTS):
            return True
        current = current.__cause__ or current.__context__
    return False


def repair_warning(repairs: tuple[PrintMetadataRepair, ...]) -> str:
    details = ", ".join(item.describe() for item in repairs)
    return (
        "Workbook was ingested from a temporary compatibility copy after removing "
        f"invalid nonfinancial print metadata: {details}. The source file was not changed."
    )


def _sanitized_print_metadata_copy(
    source_path: Path,
) -> tuple[Path, list[PrintMetadataRepair]]:
    suffix = source_path.suffix if source_path.suffix.lower() in {".xlsx", ".xlsm"} else ".xlsx"
    handle, temporary_name = tempfile.mkstemp(
        prefix="hotel_pl_print_metadata_", suffix=suffix
    )
    os.close(handle)
    temporary_path = Path(temporary_name)
    repairs: list[PrintMetadataRepair] = []
    try:
        with zipfile.ZipFile(source_path, "r") as source, zipfile.ZipFile(
            temporary_path, "w"
        ) as target:
            for entry in source.infolist():
                payload = source.read(entry.filename)
                if entry.filename == _WORKBOOK_XML:
                    payload, repairs = _remove_invalid_print_names(payload)
                target.writestr(entry, payload)
        return temporary_path, repairs
    except Exception:
        temporary_path.unlink(missing_ok=True)
        raise


def _remove_invalid_print_names(
    workbook_xml: bytes,
) -> tuple[bytes, list[PrintMetadataRepair]]:
    for _, namespace in ET.iterparse(io.BytesIO(workbook_xml), events=("start-ns",)):
        prefix, uri = namespace
        try:
            ET.register_namespace(prefix, uri)
        except ValueError:
            # Reserved generated prefixes are harmless; ElementTree will choose one.
            pass
    parser = ET.XMLParser(target=ET.TreeBuilder(insert_comments=True))
    root = ET.fromstring(workbook_xml, parser=parser)
    defined_names = root.find(f"{{{_MAIN_NAMESPACE}}}definedNames")
    repairs: list[PrintMetadataRepair] = []
    if defined_names is None:
        return workbook_xml, repairs

    for item in list(defined_names):
        name = item.attrib.get("name")
        value = (item.text or "").strip()
        if name not in _ALLOWED_NAMES or value not in _INVALID_PRINT_VALUES:
            continue
        repairs.append(
            PrintMetadataRepair(
                name=name,
                value=value,
                local_sheet_id=item.attrib.get("localSheetId"),
            )
        )
        defined_names.remove(item)

    if not repairs:
        return workbook_xml, repairs
    return ET.tostring(root, encoding="utf-8", xml_declaration=True), repairs
