from __future__ import annotations

import hashlib
import importlib
import json
import xml.etree.ElementTree as ET
import zipfile
from datetime import date, datetime, timezone
from pathlib import Path

from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter, range_boundaries

from hotel_pl_normalizer.models.common import Severity
from hotel_pl_normalizer.models.workbook import (
    CellRecord,
    CellStyle,
    FileType,
    IngestionWarning,
    MergedRange,
    WorkbookMetadata,
    WorkbookRecord,
    WorkbookRow,
    WorkbookSheet,
    WorkbookSource,
)

from .openpyxl_compat import load_openpyxl_workbook, repair_warning


def read_excel_workbook(
    path: str | Path,
    *,
    source_id: str | None = None,
) -> WorkbookRecord:
    """Read an Excel workbook into the hotel P&L WorkbookRecord contract.

    `.xlsx` and `.xlsm` files are read with OpenPyXL. Legacy `.xls` files are
    read natively with xlrd, avoiding Excel/COM conversion dependencies.
    """
    source_path = Path(path)
    suffix = source_path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return _read_openpyxl_workbook(
            read_path=source_path,
            source_path=source_path,
            file_type=FileType.XLSM if suffix == ".xlsm" else FileType.XLSX,
            ingestion_warnings=[],
            source_id=source_id,
        )
    if suffix == ".xls":
        return _read_xlrd_workbook(
            source_path=source_path,
            source_id=source_id,
        )
    raise ValueError(f"read_excel_workbook only supports .xlsx, .xlsm, and .xls files: {source_path}")


def _read_xlrd_workbook(
    *,
    source_path: Path,
    source_id: str | None = None,
) -> WorkbookRecord:
    xlrd = _load_xlrd()
    warnings: list[IngestionWarning] = []
    try:
        workbook = xlrd.open_workbook(str(source_path), formatting_info=True)
    except NotImplementedError:
        workbook = xlrd.open_workbook(str(source_path), formatting_info=False)
        warnings.append(
            IngestionWarning(
                severity=Severity.WARNING,
                message="Legacy .xls workbook was ingested without formatting metadata.",
            )
        )

    sheets: list[WorkbookSheet] = []
    sheet_visibility = getattr(workbook, "sheet_visibility", [])
    # Shared across sheets but never across workbooks -- see `_xlrd_cell_style`.
    style_cache: dict = {}
    for idx in range(workbook.nsheets):
        worksheet = workbook.sheet_by_index(idx)
        visible = idx >= len(sheet_visibility) or sheet_visibility[idx] == 0
        sheets.append(
            _read_xlrd_sheet(
                xlrd, workbook, worksheet, idx + 1, visible, style_cache=style_cache
            )
        )
        if not visible:
            warnings.append(
                IngestionWarning(
                    severity=Severity.WARNING,
                    message="Sheet is hidden but was still ingested.",
                    sheet_name=worksheet.name,
                )
            )

    source = _source_record(
        source_path=source_path,
        file_type=FileType.XLS,
        source_id=source_id,
    )
    return WorkbookRecord(
        workbook_id=_workbook_id(_parsed_content_hash(sheets)),
        source=source,
        workbook_metadata=WorkbookMetadata(sheet_count=len(sheets)),
        sheets=sheets,
        ingestion_warnings=warnings,
    )


def _read_openpyxl_workbook(
    *,
    read_path: Path,
    source_path: Path,
    file_type: FileType,
    ingestion_warnings: list[IngestionWarning],
    source_id: str | None = None,
) -> WorkbookRecord:
    warnings = list(ingestion_warnings)
    # read_only streams one row at a time instead of materialising every parsed
    # cell up front. Measured on GRY that is the difference between a 373 MB peak
    # and roughly what the rows themselves cost, because openpyxl's own objects
    # are released as we go. Styles survive the mode -- checked against the normal
    # reader, identical bold and indent counts -- which is what makes it usable
    # here, since bold and indent are the two style signals the mapper reads.
    loaded = load_openpyxl_workbook(read_path, data_only=True, read_only=True)
    if loaded.repairs:
        warnings.append(
            IngestionWarning(
                severity=Severity.WARNING,
                message=repair_warning(loaded.repairs),
            )
        )
    sheets: list[WorkbookSheet] = []
    try:
        # Shared across sheets but never across workbooks -- see `_cell_style`.
        style_cache: dict = {}
        for idx, worksheet in enumerate(loaded.workbook.worksheets, start=1):
            merge_refs = _streaming_merge_refs(
                loaded.read_path, getattr(worksheet, "_worksheet_path", "")
            )
            sheets.append(
                _read_sheet_streaming(worksheet, idx, merge_refs, style_cache=style_cache)
            )
            if getattr(worksheet, "sheet_state", "visible") != "visible":
                warnings.append(
                    IngestionWarning(
                        severity=Severity.WARNING,
                        message="Sheet is hidden but was still ingested.",
                        sheet_name=worksheet.title,
                    )
                )
    finally:
        loaded.close()
    source = _source_record(
        source_path=source_path,
        file_type=file_type,
        source_id=source_id,
    )
    return WorkbookRecord(
        workbook_id=_workbook_id(_parsed_content_hash(sheets)),
        source=source,
        workbook_metadata=WorkbookMetadata(sheet_count=len(sheets)),
        sheets=sheets,
        ingestion_warnings=warnings,
    )


def _read_sheet_streaming(
    worksheet, sheet_index: int, merge_refs: list[str], *, style_cache: dict | None = None
) -> WorkbookSheet:
    """Build one sheet from a single streaming pass.

    The random-access reader this replaces walked every cell twice: once to find
    the used bounds and again to record it. Streaming allows one pass only, so the
    bounds are discovered while building and the result trimmed afterwards -- the
    rows are retained either way, and they are the cheap part.

    Row height and hidden flags are not available in read_only mode. Nothing reads
    them, but they are part of the hashed payload, so workbook ids shift once with
    this change.
    """
    merge_specs = [(*range_boundaries(ref), ref) for ref in merge_refs]
    merged_parent_by_cell: dict[tuple[int, int], str] = {}
    top_left_positions: set[tuple[int, int]] = set()
    for first_col, first_row, last_col, last_row, _ in merge_specs:
        parent = f"{get_column_letter(first_col)}{first_row}"
        top_left_positions.add((first_row, first_col))
        for row in range(first_row, last_row + 1):
            for column in range(first_col, last_col + 1):
                merged_parent_by_cell[(row, column)] = parent

    rows: list[WorkbookRow] = []
    top_left_values: dict[tuple[int, int], str | int | float | bool | None] = {}
    max_row = 0
    max_col = 0

    for row_number, row_cells in enumerate(worksheet.iter_rows(), start=1):
        if not row_cells:
            continue
        # Gaps come back as EmptyCell, which carries no coordinates at all, so
        # position is taken from the iteration order and only confirmed against a
        # real cell when the row has one.
        row_index = next(
            (cell.row for cell in row_cells if hasattr(cell, "row")),
            row_number,
        )
        records: list[CellRecord] = []
        for offset, cell in enumerate(row_cells, start=1):
            column = getattr(cell, "column", None) or offset
            value = getattr(cell, "value", None)
            merged_parent = merged_parent_by_cell.get((row_index, column))
            if value is None:
                records.append(
                    _empty_cell_record(
                        row_index, column, merged_parent=merged_parent
                    )
                )
                continue
            record = _cell_record_streaming(
                cell,
                row_index,
                column,
                merged_parent=merged_parent,
                style_cache=style_cache,
            )
            records.append(record)
            if (row_index, column) in top_left_positions:
                top_left_values[(row_index, column)] = record.raw_value
            max_row = max(max_row, row_index)
            max_col = max(max_col, column)
        rows.append(WorkbookRow(row_index=row_index, cells=records))

    merged_ranges: list[MergedRange] = []
    for first_col, first_row, last_col, last_row, ref in merge_specs:
        value = top_left_values.get((first_row, first_col))
        merged_ranges.append(
            MergedRange(
                range=ref,
                top_left_row=first_row,
                top_left_column=first_col,
                value=value,
            )
        )
        if value is not None:
            max_row = max(max_row, last_row)
            max_col = max(max_col, last_col)

    # Trim to the used range, then pad short rows so every row is the same width,
    # which is what the random-access reader produced.
    trimmed: list[WorkbookRow] = []
    for row in rows:
        if row.row_index > max_row:
            continue
        cells = [cell for cell in row.cells if cell.column <= max_col]
        present = {cell.column for cell in cells}
        for column in range(1, max_col + 1):
            if column not in present:
                cells.append(
                    _empty_cell_record(
                        row.row_index,
                        column,
                        merged_parent=merged_parent_by_cell.get(
                            (row.row_index, column)
                        ),
                    )
                )
        cells.sort(key=lambda cell: cell.column)
        row.cells = cells
        trimmed.append(row)

    return WorkbookSheet(
        sheet_id=f"sheet_{sheet_index}",
        sheet_name=worksheet.title,
        visible=getattr(worksheet, "sheet_state", "visible") == "visible",
        max_row=max_row,
        max_column=max_col,
        merged_ranges=merged_ranges,
        rows=trimmed,
    )


def _cell_record_streaming(
    cell,
    row: int,
    column: int,
    *,
    merged_parent: str | None = None,
    style_cache: dict | None = None,
) -> CellRecord:
    """One populated cell from a read_only worksheet.

    Merge references come from the sheet XML because read-only cells do not expose
    them directly.
    """
    return CellRecord(
        row=row,
        column=column,
        address=f"{get_column_letter(column)}{row}",
        raw_value=_serializable_value(cell.value),
        display_value=_display_value(cell.value),
        number_format=getattr(cell, "number_format", None),
        is_merged=merged_parent is not None,
        merged_parent=merged_parent,
        style=_cell_style(cell, style_cache),
    )


def _empty_cell_record(
    row: int, column: int, *, merged_parent: str | None = None
) -> CellRecord:
    return CellRecord(
        row=row,
        column=column,
        address=f"{get_column_letter(column)}{row}",
        is_merged=merged_parent is not None,
        merged_parent=merged_parent,
    )


def _streaming_merge_refs(path: Path, worksheet_path: str) -> list[str]:
    """Read merge references directly from sheet XML without loading its cells."""
    if not worksheet_path:
        return []
    archive_path = worksheet_path.lstrip("/")
    try:
        with zipfile.ZipFile(path) as archive, archive.open(archive_path) as stream:
            refs: list[str] = []
            for _, element in ET.iterparse(stream, events=("end",)):
                if element.tag.rsplit("}", 1)[-1] == "mergeCell":
                    ref = element.attrib.get("ref")
                    if ref:
                        refs.append(ref)
                element.clear()
            return refs
    except (KeyError, OSError, zipfile.BadZipFile, ET.ParseError):
        return []


def _source_record(
    *,
    source_path: Path,
    file_type: FileType,
    source_id: str | None = None,
) -> WorkbookSource:
    return WorkbookSource(
        source_id=source_id or f"primary:{source_path.stem}",
        original_filename=source_path.name,
        file_type=file_type,
        local_path=str(source_path),
        file_hash=_sha256(source_path),
        ingested_at=datetime.now(timezone.utc),
    )


def _read_xlrd_sheet(
    xlrd,
    workbook,
    worksheet,
    sheet_index: int,
    visible: bool,
    *,
    style_cache: dict | None = None,
) -> WorkbookSheet:
    max_row, max_col = _xlrd_used_bounds(xlrd, workbook, worksheet)
    merged_ranges = [_xlrd_merged_range_record(xlrd, workbook, worksheet, merged) for merged in worksheet.merged_cells]
    merged_parent_by_cell = _xlrd_merged_parent_lookup(worksheet)
    rows: list[WorkbookRow] = []
    rowinfo_map = getattr(worksheet, "rowinfo_map", {})

    for row_idx in range(1, max_row + 1):
        row_info = rowinfo_map.get(row_idx - 1)
        height = float(row_info.height) if row_info is not None and getattr(row_info, "height", None) else None
        hidden = bool(getattr(row_info, "hidden", False)) if row_info is not None else False
        cells = [
            _xlrd_cell_record(
                xlrd,
                workbook,
                worksheet,
                row_idx - 1,
                col_idx - 1,
                merged_parent_by_cell,
                style_cache=style_cache,
            )
            for col_idx in range(1, max_col + 1)
        ]
        rows.append(WorkbookRow(row_index=row_idx, height=height, hidden=hidden, cells=cells))

    return WorkbookSheet(
        sheet_id=f"sheet_{sheet_index}",
        sheet_name=worksheet.name,
        visible=visible,
        max_row=max_row,
        max_column=max_col,
        merged_ranges=merged_ranges,
        rows=rows,
    )


def _xlrd_used_bounds(xlrd, workbook, worksheet) -> tuple[int, int]:
    max_row = 0
    max_col = 0
    for rowx in range(worksheet.nrows):
        for colx in range(worksheet.ncols):
            value = _xlrd_cell_value(xlrd, workbook, worksheet.cell(rowx, colx))
            if value is None:
                continue
            max_row = max(max_row, rowx + 1)
            max_col = max(max_col, colx + 1)

    for row_start, row_end, col_start, col_end in worksheet.merged_cells:
        value = _xlrd_cell_value(xlrd, workbook, worksheet.cell(row_start, col_start))
        if value is None:
            continue
        max_row = max(max_row, row_end)
        max_col = max(max_col, col_end)

    return max_row, max_col


def _xlrd_merged_range_record(xlrd, workbook, worksheet, merged) -> MergedRange:
    row_start, row_end, col_start, col_end = merged
    top_left = worksheet.cell(row_start, col_start)
    return MergedRange(
        range=f"{_xlrd_address(row_start, col_start)}:{_xlrd_address(row_end - 1, col_end - 1)}",
        top_left_row=row_start + 1,
        top_left_column=col_start + 1,
        value=_serializable_value(_xlrd_cell_value(xlrd, workbook, top_left)),
    )


def _xlrd_cell_record(
    xlrd,
    workbook,
    worksheet,
    rowx: int,
    colx: int,
    merged_parent_by_cell: dict[str, str],
    *,
    style_cache: dict | None = None,
) -> CellRecord:
    cell = worksheet.cell(rowx, colx)
    value = _xlrd_cell_value(xlrd, workbook, cell)
    address = _xlrd_address(rowx, colx)
    merged_parent = merged_parent_by_cell.get(address)
    return CellRecord(
        row=rowx + 1,
        column=colx + 1,
        address=address,
        raw_value=_serializable_value(value),
        display_value=_display_value(value),
        number_format=_xlrd_number_format(workbook, cell),
        is_merged=merged_parent is not None,
        merged_parent=merged_parent,
        style=_xlrd_cell_style(workbook, cell, style_cache),
    )


def _cell_style(cell: Cell, style_cache: dict | None = None) -> CellStyle:
    """One cell's style, reusing the instance for cells that share a style.

    `cell.font` and friends are properties that resolve against the workbook's
    style tables on every access, so the uncached path costs six table lookups per
    cell. A workbook has a few hundred distinct styles and hundreds of thousands of
    cells -- GRY, measured, is 301 styles across 10,949 cells -- so keying on the
    style index collapses almost all of that work. `CellStyle` is frozen, which is
    what makes handing the same instance to many cells safe.

    The cache is per `load_workbook` call and must stay that way: `style_array`
    holds indices into *this* workbook's tables, and the same indices mean
    different styles in a different file.
    """
    if style_cache is None:
        return _build_cell_style(cell)
    style_array = getattr(cell, "style_array", None)
    if style_array is None:
        return _build_cell_style(cell)
    key = tuple(style_array)
    cached = style_cache.get(key)
    if cached is None:
        cached = style_cache[key] = _build_cell_style(cell)
    return cached


def _build_cell_style(cell: Cell) -> CellStyle:
    font = cell.font
    alignment = cell.alignment
    return CellStyle(
        bold=bool(font.bold),
        italic=bool(font.italic),
        underline=bool(font.underline),
        font_size=float(font.sz) if font.sz is not None else None,
        indent=float(alignment.indent) if alignment.indent is not None else None,
        horizontal_alignment=alignment.horizontal,
        fill_color=_color(cell.fill.fgColor),
        font_color=_color(font.color),
        border_bottom=cell.border.bottom.style,
    )


def _xlrd_cell_style(workbook, cell, style_cache: dict | None = None) -> CellStyle:
    """The .xls equivalent of `_cell_style`, keyed on the extended format index.

    `xf_index` points into this workbook's `xf_list`, so the cache is per workbook
    for the same reason.
    """
    if style_cache is not None:
        key = getattr(cell, "xf_index", None)
        if key is not None:
            cached = style_cache.get(key)
            if cached is None:
                cached = style_cache[key] = _build_xlrd_cell_style(workbook, cell)
            return cached
    return _build_xlrd_cell_style(workbook, cell)


def _build_xlrd_cell_style(workbook, cell) -> CellStyle:
    xf = _xlrd_xf(workbook, cell)
    if xf is None:
        return CellStyle()

    font = None
    font_list = getattr(workbook, "font_list", [])
    if getattr(xf, "font_index", None) is not None and xf.font_index < len(font_list):
        font = font_list[xf.font_index]

    alignment = getattr(xf, "alignment", None)
    background = getattr(xf, "background", None)
    border = getattr(xf, "border", None)
    return CellStyle(
        bold=bool(getattr(font, "bold", False)) if font is not None else False,
        italic=bool(getattr(font, "italic", False)) if font is not None else False,
        underline=bool(getattr(font, "underlined", False)) if font is not None else False,
        font_size=(float(font.height) / 20.0) if font is not None and getattr(font, "height", None) else None,
        indent=float(alignment.indent_level) if alignment is not None and getattr(alignment, "indent_level", None) else None,
        horizontal_alignment=(
            f"xlrd:{alignment.hor_align}" if alignment is not None and getattr(alignment, "hor_align", None) is not None else None
        ),
        fill_color=(
            f"indexed:{background.pattern_colour_index}"
            if background is not None and getattr(background, "pattern_colour_index", None)
            else None
        ),
        font_color=(
            f"indexed:{font.colour_index}" if font is not None and getattr(font, "colour_index", None) is not None else None
        ),
        border_bottom=(
            f"xlrd:{border.bottom_line_style}" if border is not None and getattr(border, "bottom_line_style", None) else None
        ),
    )


def _xlrd_merged_parent_lookup(worksheet) -> dict[str, str]:
    lookup: dict[str, str] = {}
    for row_start, row_end, col_start, col_end in worksheet.merged_cells:
        parent = _xlrd_address(row_start, col_start)
        for rowx in range(row_start, row_end):
            for colx in range(col_start, col_end):
                lookup[_xlrd_address(rowx, colx)] = parent
    return lookup


def _xlrd_cell_value(xlrd, workbook, cell):
    if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
        return None
    if cell.ctype == xlrd.XL_CELL_DATE:
        return xlrd.xldate_as_datetime(cell.value, workbook.datemode)
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return bool(cell.value)
    if cell.ctype == xlrd.XL_CELL_ERROR:
        error_lookup = getattr(xlrd, "error_text_from_code", {})
        error_text = error_lookup.get(cell.value, cell.value) if isinstance(error_lookup, dict) else cell.value
        return f"#ERROR:{error_text}"
    if cell.ctype == xlrd.XL_CELL_NUMBER and isinstance(cell.value, float) and cell.value.is_integer():
        return int(cell.value)
    return cell.value


def _xlrd_number_format(workbook, cell) -> str | None:
    xf = _xlrd_xf(workbook, cell)
    if xf is None:
        return None
    format_key = getattr(xf, "format_key", None)
    format_map = getattr(workbook, "format_map", {})
    if format_key in format_map:
        return getattr(format_map[format_key], "format_str", None)
    return None


def _xlrd_xf(workbook, cell):
    xf_index = getattr(cell, "xf_index", None)
    xf_list = getattr(workbook, "xf_list", [])
    if xf_index is None or xf_index >= len(xf_list):
        return None
    return xf_list[xf_index]


def _xlrd_address(rowx: int, colx: int) -> str:
    return f"{get_column_letter(colx + 1)}{rowx + 1}"


def _load_xlrd():
    try:
        return importlib.import_module("xlrd")
    except ImportError as exc:
        raise RuntimeError("Reading legacy .xls files requires xlrd. Install project dependencies and retry.") from exc


def _serializable_value(value):
    if value is None or isinstance(value, str | int | float | bool):
        return value
    if isinstance(value, datetime | date):
        return value.isoformat()
    return str(value)


def _display_value(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime | date):
        return value.isoformat()
    return str(value)


def _color(color) -> str | None:
    if color is None:
        return None
    if color.type == "rgb" and color.rgb:
        return str(color.rgb)
    if color.type == "indexed":
        return f"indexed:{color.indexed}"
    if color.type == "theme":
        return f"theme:{color.theme}"
    return None


def _workbook_id(file_hash: str) -> str:
    """Derive a stable workbook id from the parsed workbook content hash.

    The workbook id is embedded in every ``packet_id`` and therefore in every
    prompt, so a random id makes each run's prompts unique and silently defeats
    the response cache. Deriving it from content means re-ingesting the same
    workbook reproduces byte-identical prompts, which keeps repeat runs cheap
    and makes two runs comparable when only the code under test changed.
    """
    return f"wb_{file_hash[:32]}"


def _parsed_content_hash(sheets: list[WorkbookSheet]) -> str:
    """Hash every parsed field that can affect downstream interpretation.

    Excel rewrites archive metadata when a workbook is merely opened and saved.
    Hashing the parsed sheets keeps that metadata churn out of packet ids while
    still invalidating the cache for cell, layout, style, visibility, or merged
    range changes that can affect routing and mapping. Feed one small record at a
    time so hashing does not duplicate the complete workbook as dictionaries and
    then duplicate it again as one JSON string.
    """
    digest = hashlib.sha256()
    _hash_piece(digest, ("workbook_hash_version", 2, len(sheets)))
    for sheet in sheets:
        _hash_piece(
            digest,
            (
                "sheet",
                sheet.sheet_id,
                sheet.sheet_name,
                sheet.max_row,
                sheet.max_column,
                sheet.visible,
                len(sheet.rows),
                len(sheet.merged_ranges),
            ),
        )
        for merged in sheet.merged_ranges:
            _hash_piece(
                digest,
                (
                    "merge",
                    merged.range,
                    merged.top_left_row,
                    merged.top_left_column,
                    merged.value,
                ),
            )
        for row in sheet.rows:
            _hash_piece(
                digest,
                ("row", row.row_index, row.height, row.hidden, len(row.cells)),
            )
            for cell in row.cells:
                style = cell.style
                _hash_piece(
                    digest,
                    (
                        "cell",
                        cell.row,
                        cell.column,
                        cell.address,
                        cell.raw_value,
                        cell.display_value,
                        cell.number_format,
                        cell.is_merged,
                        cell.merged_parent,
                        style.bold,
                        style.italic,
                        style.underline,
                        style.font_size,
                        style.indent,
                        style.horizontal_alignment,
                        style.fill_color,
                        style.font_color,
                        style.border_bottom,
                    ),
                )
    return digest.hexdigest()


def _hash_piece(digest, value) -> None:
    encoded = json.dumps(
        value,
        ensure_ascii=False,
        separators=(",", ":"),
    ).encode("utf-8")
    digest.update(len(encoded).to_bytes(8, "big"))
    digest.update(encoded)


def _sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()
