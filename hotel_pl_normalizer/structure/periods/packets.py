"""Build the packets a period stage is given to reason about.

One packet per department location for column selection, and one workbook-wide
packet for discovery. Both reduce a parsed workbook to the header context and
candidate columns a model can hold at once, which is the whole reason the stages
do not simply read the workbook themselves.
"""

from __future__ import annotations

import re
from collections import defaultdict

from openpyxl.utils import get_column_letter, range_boundaries

from hotel_pl_normalizer.models.department_location import (
    BoundaryConfidence,
    DepartmentLocation,
    DepartmentLocationKind,
    DepartmentLocationMap,
    DepartmentSectionRole,
)
from hotel_pl_normalizer.models.period_selection import (
    PeriodColumnCandidate,
    PeriodColumnPacket,
    PeriodColumnSample,
    PeriodHeaderCell,
    PeriodMergedHeader,
    PeriodSheetPacket,
)
from hotel_pl_normalizer.models.workbook import (
    WorkbookRecord,
    WorkbookRow,
    WorkbookSheet,
)

from .signals import (
    _looks_like_period_header_text,
    _looks_like_rejected_metric_column,
    _looks_like_supporting_schedule_sheet,
)

# A merged cell spanning the target column is treated as a header for it, which
# is right for the merges period headers actually use: `K5:S5` over "Current
# Period" covers the nine columns it labels.
#
# It is wrong for a banner. SFI titles its sheets with `F1:Y1` merged across
# twenty columns reading "Forecast December Monitoring", the name of the report.
# Every data column fell inside that span, so every column inherited the word
# "Forecast", and the scenario check refused an Actual period on a column whose
# own headers read "Year-to-Date" and "2025 Actual". The run was lost to a title.
#
# Width alone does not separate the two -- a real header ran to nine columns and
# a banner to twenty, and any cut-off between them is a guess about the next
# workbook. A banner is also the full width of the sheet, so both tests have to
# hold: wider than any plausible period group *and* covering most of the sheet.
MAX_PERIOD_HEADER_MERGE_COLUMNS = 12
BANNER_SHEET_WIDTH_SHARE = 0.5


def build_period_column_packet(
    workbook: WorkbookRecord,
    location_map: DepartmentLocationMap,
    *,
    requested_period: str = "YTD Actual",
    max_candidates_per_sheet: int = 80,
) -> PeriodColumnPacket:
    sheets = []
    for location in _primary_locations(location_map):
        sheet = _find_sheet(workbook, location.sheet_name)
        candidate_columns = _candidate_columns_for_location(
            sheet,
            location,
            max_candidates=max_candidates_per_sheet,
        )
        sheets.append(
            PeriodSheetPacket(
                location_id=location.location_id,
                sheet_name=sheet.sheet_name,
                department=location.department,
                start_row=location.start_row,
                end_row=location.end_row,
                header_cells=_header_snapshot(sheet),
                merged_headers=_merged_header_snapshot(sheet),
                candidate_columns=candidate_columns,
            )
        )
    return PeriodColumnPacket(
        packet_id=f"{workbook.workbook_id}:period_columns",
        workbook_id=workbook.workbook_id,
        source_filename=workbook.source.original_filename,
        requested_period=requested_period,
        sheets=sheets,
        notes=[
            "Infer the dominant workbook convention, then allow sheet-specific overrides when a sheet has a different layout."
        ],
    )


def build_workbook_period_discovery_packet(
    workbook: WorkbookRecord,
    *,
    requested_period: str = "YTD Actual",
    max_candidates_per_sheet: int = 80,
    max_sheets: int | None = 5,
    include_candidates: bool = False,
) -> PeriodColumnPacket:
    """Build period evidence before sheet routing or department identification.

    One synthetic whole-sheet location is enough for discovery. Department-level
    ranges are added only after the human has selected the periods to map.
    """
    ranked_sheets = [
        (index, sheet)
        for index, sheet in enumerate(workbook.sheets)
        if sheet.visible and _discovery_sheet_has_period_evidence(sheet)
    ]
    if not ranked_sheets:
        ranked_sheets = [
            (index, sheet)
            for index, sheet in enumerate(workbook.sheets)
            if sheet.visible
        ]
    ranked_sheets.sort(
        key=lambda item: _discovery_workbook_sheet_score(item[1], item[0]),
        reverse=True,
    )
    if max_sheets is not None:
        ranked_sheets = ranked_sheets[:max_sheets]

    sheets = []
    for index, sheet in ranked_sheets:
        location = DepartmentLocation(
            location_id=f"sheet:{index}",
            department="workbook",
            section_role=DepartmentSectionRole.PRIMARY,
            location_kind=DepartmentLocationKind.SHEET,
            sheet_name=sheet.sheet_name,
            start_row=1,
            end_row=sheet.max_row,
            boundary_confidence=BoundaryConfidence.EXACT,
        )
        candidates = (
            _candidate_columns_for_location(
                sheet, location, max_candidates=max_candidates_per_sheet
            )
            if include_candidates
            else []
        )
        spatial_header = _spatial_header_snapshot(
            sheet,
            max_candidate_column=max(1, min(sheet.max_column, 60)),
        )
        sheets.append(
            PeriodSheetPacket(
                location_id=location.location_id,
                sheet_name=sheet.sheet_name,
                start_row=1,
                end_row=sheet.max_row,
                header_cells=list(spatial_header["spatial_header_cells"]),
                merged_headers=list(spatial_header["spatial_merged_headers"]),
                **spatial_header,
                candidate_columns=candidates,
            )
        )
    return PeriodColumnPacket(
        packet_id=f"{workbook.workbook_id}:period_discovery",
        workbook_id=workbook.workbook_id,
        source_filename=workbook.source.original_filename,
        requested_period=requested_period,
        sheets=sheets,
        notes=[
            "Period discovery includes coordinate-preserving header windows for up to five cheaply selected visible sheets; exact numeric candidates are built only after period selection."
        ],
    )


def _discovery_sheet_has_period_evidence(sheet: WorkbookSheet) -> bool:
    return any(
        _row_has_period_header_evidence(row)
        for row in sheet.rows
        if row.row_index <= 60
    )


def _discovery_workbook_sheet_score(
    sheet: WorkbookSheet, workbook_index: int
) -> tuple[int, int, int, int, int, int]:
    """Rank likely core P&L tabs without inspecting numeric body columns."""
    name = re.sub(r"[^a-z0-9]+", " ", sheet.sheet_name.lower()).strip()
    preferred = int(
        any(
            term in name
            for term in (
                "summary",
                "total hotel",
                "profit loss",
                "p l",
                "operating statement",
            )
        )
    )
    supporting = int(_looks_like_supporting_schedule_sheet(sheet.sheet_name))
    explicit_supporting_title = int(
        any(
            isinstance(cell.raw_value, str)
            and re.search(
                r"\b(?:statistical summary|statistics|kpi|critique)\b",
                cell.display_value or "",
                re.IGNORECASE,
            )
            for row in sheet.rows
            if row.row_index <= 60
            for cell in row.cells
        )
    )
    header_evidence = sum(
        1
        for row in sheet.rows
        if row.row_index <= 60
        for cell in row.cells
        if isinstance(cell.raw_value, str)
        and _looks_like_broad_header_text(cell.display_value or "")
    )
    # Prefer a named statement, then ordinary P&L tabs over obvious Stats/KPI
    # schedules. Header evidence and width break ties without reading body values.
    return (
        1 - explicit_supporting_title,
        1 - supporting,
        preferred,
        min(header_evidence, 50),
        min(sheet.max_column, 60),
        -workbook_index,
    )


def _primary_locations(location_map: DepartmentLocationMap) -> list[DepartmentLocation]:
    return [
        location
        for location in location_map.locations
        if location.section_role.value in {"primary", "summary"} and location.location_kind.value in {"sheet", "range"}
    ]


def _candidate_columns_for_location(
    sheet: WorkbookSheet,
    location: DepartmentLocation,
    *,
    max_candidates: int,
) -> list[PeriodColumnCandidate]:
    first_data_row = _first_business_row(sheet, location)
    rows = [
        row
        for row in _rows_in_location(sheet, location)
        if row.row_index >= first_data_row
    ]
    labelled_values_by_row: dict[int, tuple[str, dict[int, float]]] = {}
    numeric_count: dict[int, int] = defaultdict(int)
    nonzero_count: dict[int, int] = defaultdict(int)
    percentage_format_count: dict[int, int] = defaultdict(int)
    subunit_nonzero_count: dict[int, int] = defaultdict(int)
    material_value_count: dict[int, int] = defaultdict(int)
    percentage_scale_count: dict[int, int] = defaultdict(int)
    large_amount_count: dict[int, int] = defaultdict(int)

    for row in rows:
        label = _row_label(row)
        if not label:
            continue
        row_values: dict[int, float] = {}
        for cell in row.cells:
            if not _is_number(cell.raw_value):
                continue
            value = float(cell.raw_value)
            row_values[cell.column] = value
            numeric_count[cell.column] += 1
            if _is_percentage_number_format(cell.number_format):
                percentage_format_count[cell.column] += 1
            if value != 0:
                nonzero_count[cell.column] += 1
                if abs(value) <= 1.5:
                    subunit_nonzero_count[cell.column] += 1
                if abs(value) >= 10:
                    material_value_count[cell.column] += 1
                if abs(value) <= 200:
                    percentage_scale_count[cell.column] += 1
                if abs(value) >= 1000:
                    large_amount_count[cell.column] += 1
        if row_values:
            labelled_values_by_row[row.row_index] = (label, row_values)

    candidate_columns = sorted(
        column for column, count in numeric_count.items() if count >= 2
    )[:max_candidates]
    candidate_column_set = set(candidate_columns)
    sample_rows = sorted(
        sorted(
            labelled_values_by_row,
            key=lambda row_index: (
                -len(
                    candidate_column_set.intersection(
                        labelled_values_by_row[row_index][1]
                    )
                ),
                row_index,
            ),
        )[:3]
    )

    candidates = [
        PeriodColumnCandidate(
            column=column,
            excel_column=get_column_letter(column),
            header_context=_header_context(sheet, column, first_data_row),
            numeric_count=numeric_count[column],
            nonzero_count=nonzero_count[column],
            percentage_format_count=percentage_format_count[column],
            subunit_nonzero_count=subunit_nonzero_count[column],
            material_value_count=material_value_count[column],
            percentage_scale_count=percentage_scale_count[column],
            large_amount_count=large_amount_count[column],
            sample_values=[
                PeriodColumnSample(
                    row_index=row_index,
                    label=labelled_values_by_row[row_index][0],
                    value=labelled_values_by_row[row_index][1].get(column),
                )
                for row_index in sample_rows
            ],
        )
        for column in candidate_columns
    ]
    return candidates


def _rows_in_location(sheet: WorkbookSheet, location: DepartmentLocation) -> list[WorkbookRow]:
    start = location.start_row or 1
    end = location.end_row or sheet.max_row
    return [row for row in sheet.rows if start <= row.row_index <= end]


def _first_business_row(sheet: WorkbookSheet, location: DepartmentLocation) -> int:
    numeric_rows = [
        row
        for row in _rows_in_location(sheet, location)
        if _row_label(row)
        and any(_is_number(cell.raw_value) for cell in row.cells)
    ]
    for row in numeric_rows:
        label = _row_label(row)
        if label and _looks_like_business_label(label):
            # Short sections such as Utilities may contain only generic account
            # labels before their Total row. Starting at that Total leaves fewer
            # than the two numeric samples required for any period candidate, so
            # retain the section's earlier labelled numeric rows in that case.
            remaining = sum(
                1 for item in numeric_rows if item.row_index >= row.row_index
            )
            if remaining >= 2:
                return row.row_index
            return numeric_rows[0].row_index
    if numeric_rows:
        return numeric_rows[0].row_index
    return location.start_row or 1



def _looks_like_banner(first_column: int, last_column: int, sheet: WorkbookSheet) -> bool:
    """Is this merge a sheet-wide title rather than a header over some columns?"""
    span = last_column - first_column + 1
    if span <= MAX_PERIOD_HEADER_MERGE_COLUMNS:
        return False
    width = max(sheet.max_column or 0, last_column)
    return span >= width * BANNER_SHEET_WIDTH_SHARE


def _header_context(sheet: WorkbookSheet, column: int, first_data_row: int) -> list[str]:
    entries: list[tuple[int, str]] = []
    header_rows = _header_search_rows(sheet, first_data_row)
    for row in header_rows:
        header_label = _header_row_label(row, column)
        has_direct_text = False
        has_direct_period_header = False
        has_merged_period_header = any(
            first_row == row.row_index
            and first_column <= column <= last_column
            and _looks_like_period_header_text(str(merged.value or ""))
            for merged in sheet.merged_ranges
            for first_column, first_row, last_column, _ in [
                range_boundaries(merged.range)
            ]
        )
        for cell in row.cells:
            display_value = (cell.display_value or "").strip()
            if (
                cell.column == column
                and isinstance(cell.raw_value, str)
                and display_value
                and (
                    row.row_index < first_data_row
                    or _looks_like_broad_header_text(display_value)
                )
            ):
                has_direct_text = True
                has_direct_period_header = _looks_like_period_header_text(display_value)
                if header_label:
                    entries.append(
                        (row.row_index, f"{cell.address} {header_label}: {display_value}")
                    )
                else:
                    entries.append((row.row_index, f"{cell.address}: {display_value}"))
        if (
            not has_direct_text
            and not has_direct_period_header
            and not has_merged_period_header
        ):
            nearby = [
                cell
                for cell in row.cells
                if abs(first_data_row - row.row_index) <= 6
                and isinstance(cell.raw_value, str)
                and (cell.display_value or "").strip()
                and abs(cell.column - column) <= 6
                and _looks_like_period_header_text(cell.display_value or "")
            ]
            if nearby:
                nearest_distance = min(abs(cell.column - column) for cell in nearby)
                nearest = [
                    cell
                    for cell in nearby
                    if abs(cell.column - column) == nearest_distance
                ]
                nearest_values = {
                    re.sub(r"\s+", " ", (cell.display_value or "")).strip().lower()
                    for cell in nearest
                }
                if len(nearest_values) == 1:
                    cell = nearest[0]
                    entries.append(
                        (
                            row.row_index,
                            f"{cell.address} nearby: {(cell.display_value or '').strip()}",
                        )
                    )
    for merged in sheet.merged_ranges:
        first_column, first_row, last_column, _ = range_boundaries(merged.range)
        value = str(merged.value or "").strip()
        if (
            value
            and any(row.row_index == first_row for row in header_rows)
            and (
                first_row < first_data_row
                or _looks_like_broad_header_text(value)
            )
            and first_column <= column <= last_column
            and column != first_column
            and not _looks_like_banner(first_column, last_column, sheet)
        ):
            entries.append((first_row, f"{merged.range} merged: {value}"))
    context = []
    for _, text in sorted(entries, key=lambda item: item[0]):
        if text not in context:
            context.append(text)
    return context[-10:]


def _header_search_rows(
    sheet: WorkbookSheet, first_data_row: int
) -> list[WorkbookRow]:
    """Keep ordinary leading headers plus explicit/repeated period header rows."""
    leading_end = max(first_data_row - 1, min(sheet.max_row, 40))
    return [
        row
        for row in sheet.rows
        if row.row_index <= leading_end or _row_has_period_header_evidence(row)
    ]


def _row_has_period_header_evidence(row: WorkbookRow) -> bool:
    return any(
        isinstance(cell.raw_value, str)
        and _looks_like_broad_header_text(cell.display_value or "")
        for cell in row.cells
    )


def _looks_like_broad_header_text(text: str) -> bool:
    lowered = re.sub(r"\s+", " ", str(text or "")).strip().lower()
    if len(lowered) > 80:
        return False
    return bool(
        _looks_like_period_header_text(lowered)
        or lowered == "total"
        or re.search(r"\b20\d{2}(?:[./-][a-z]{3,9}|[./-]\d{1,2})?\b", lowered)
        or re.search(
            r"\b(actual|budget|forecast|prior year|last year|current period|"
            r"year to date|year-to-date|periodic|month|timeid|time|"
            r"category|measure|actualflag)\b",
            lowered,
        )
    )


def _spatial_header_snapshot(
    sheet: WorkbookSheet,
    *,
    max_candidate_column: int,
    max_rows: int = 12,
    max_columns: int = 60,
) -> dict:
    """Capture one contiguous header window without flattening its row hierarchy."""
    evidence = [
        row for row in sheet.rows if _row_has_period_header_evidence(row)
    ]
    best_group: list[int] = []
    if evidence:
        groups: list[list[WorkbookRow]] = []
        for row in evidence:
            if groups and row.row_index - groups[-1][-1].row_index <= 3:
                groups[-1].append(row)
            else:
                groups.append([row])

        def group_score(group: list[WorkbookRow]) -> tuple[int, int, int]:
            strength = sum(
                2
                for row in group
                for cell in row.cells
                if isinstance(cell.raw_value, str)
                and re.search(
                    r"\b(?:actual|budget|forecast|prior year|last year|ytd|"
                    r"year.to.date|current period|period.to.date|mtd|ptd|total)\b",
                    cell.display_value or "",
                    re.IGNORECASE,
                )
            )
            return strength, len(group), group[-1].row_index

        strongest = max(groups, key=group_score)
        best_group = [row.row_index for row in strongest]
        start_row = best_group[0]
        row_by_index = {row.row_index: row for row in sheet.rows}
        for row_index in range(start_row - 1, max(0, start_row - 4), -1):
            row = row_by_index.get(row_index)
            if row is None:
                continue
            if any(
                isinstance(cell.raw_value, (int, float))
                and not isinstance(cell.raw_value, bool)
                for cell in row.cells
            ):
                break
            start_row = row_index
        end_row = best_group[-1]
        if end_row - start_row + 1 > max_rows:
            start_row = end_row - max_rows + 1
    else:
        start_row = 1
        end_row = min(sheet.max_row, max_rows)

    provisional_end_column = min(max_columns, max_candidate_column)

    nonempty_rows = [
        row.row_index
        for row in sheet.rows
        if start_row <= row.row_index <= end_row
        and any(
            cell.column <= provisional_end_column
            and str(cell.display_value or "").strip()
            for cell in row.cells
        )
    ]
    if nonempty_rows:
        end_row = max(nonempty_rows)

    merged_end_columns = []
    spatial_merged_headers = []
    for merged in sheet.merged_ranges:
        first_column, first_row, last_column, last_row = range_boundaries(merged.range)
        if last_row < start_row or first_row > end_row:
            continue
        merged_end_columns.append(last_column)
        value = str(merged.value or "").strip()
        if value:
            spatial_merged_headers.append(
                PeriodMergedHeader(range=merged.range, value=value)
            )

    end_column = min(
        max_columns,
        max([max_candidate_column, *merged_end_columns], default=max_candidate_column),
    )
    cells = []
    for row in sheet.rows:
        if not start_row <= row.row_index <= end_row:
            continue
        for cell in row.cells:
            if cell.column > end_column:
                continue
            value = re.sub(r"\s+", " ", str(cell.display_value or "")).strip()
            if value:
                cells.append(PeriodHeaderCell(coordinate=cell.address, value=value))
    return {
        "spatial_header_start_row": start_row,
        "spatial_header_end_row": end_row,
        "spatial_header_end_column": end_column,
        "spatial_header_cells": cells,
        "spatial_merged_headers": spatial_merged_headers,
    }


def _header_snapshot(
    sheet: WorkbookSheet, *, max_cells: int = 120
) -> list[PeriodHeaderCell]:
    """Return compact, coordinate-preserving header rows for model discovery."""
    rows = [row for row in sheet.rows if _row_has_period_header_evidence(row)]
    cells: list[PeriodHeaderCell] = []
    for row in rows:
        for cell in row.cells:
            value = (cell.display_value or "").strip()
            if not value or not isinstance(cell.raw_value, str):
                continue
            if not (
                _looks_like_broad_header_text(value)
                or _looks_like_rejected_metric_column(value.lower())
            ):
                continue
            cells.append(PeriodHeaderCell(coordinate=cell.address, value=value))
            if len(cells) >= max_cells:
                return cells
    return cells


def _merged_header_snapshot(
    sheet: WorkbookSheet, *, max_ranges: int = 40
) -> list[PeriodMergedHeader]:
    headers = []
    for merged in sheet.merged_ranges:
        value = str(merged.value or "").strip()
        if value and _looks_like_broad_header_text(value):
            headers.append(PeriodMergedHeader(range=merged.range, value=value))
            if len(headers) >= max_ranges:
                break
    return headers


def _header_row_label(row: WorkbookRow, column: int) -> str | None:
    text_cells = [
        cell
        for cell in row.cells
        if cell.column < column and isinstance(cell.raw_value, str) and (cell.display_value or "").strip()
    ]
    for cell in text_cells:
        normalized = re.sub(r"[^a-z0-9]+", " ", (cell.display_value or "").lower()).strip()
        if normalized in {
            "actualflag",
            "actual flag",
            "alt scenario cat",
            "category",
            "measure",
            "month",
            "period",
            "time",
            "timeid",
        }:
            return (cell.display_value or "").strip()
    return None


def _row_label(row: WorkbookRow) -> str | None:
    text_cells = [
        cell
        for cell in row.cells
        if isinstance(cell.raw_value, str) and cell.display_value and re.search(r"[A-Za-z]", cell.display_value)
    ]
    if not text_cells:
        return None
    candidates = [cell for cell in text_cells if not _looks_like_control_text(cell.display_value or "")]
    if not candidates:
        return None
    return max(candidates, key=lambda cell: (cell.column, len(cell.display_value or ""))).display_value


def _looks_like_control_text(text: str) -> bool:
    stripped = text.strip().lower()
    if stripped.startswith("%,") or "fchartfield" in stripped or "faccount" in stripped:
        return True
    normalized = re.sub(r"[^a-z0-9]+", " ", text.lower()).strip()
    return normalized in {
        "app",
        "account",
        "afterrange",
        "beforerange",
        "businessunit",
        "category",
        "cellkeyrange",
        "colkeyrange",
        "criteria",
        "datasrc",
        "department",
        "dimension",
        "actualflag",
        "alt scenario cat",
        "altscenariocat",
        "forecastyear",
        "measures",
        "measure",
        "memberset",
        "month",
        "parameter",
        "period",
        "rowkeyrange",
        "time",
        "timeid",
    }


def _is_number(value) -> bool:
    return isinstance(value, int | float) and not isinstance(value, bool)


def _looks_like_business_label(label: str) -> bool:
    normalized = re.sub(r"[^a-z0-9]+", " ", label.lower()).strip()
    return any(
        term in normalized
        for term in (
            "adr",
            "available rooms",
            "cost",
            "ebitda",
            "expense",
            "expenses",
            "fees",
            "income",
            "labor",
            "occupancy",
            "occupied rooms",
            "payroll",
            "profit",
            "rent",
            "rev par",
            "revenue",
            "revenues",
            "rooms sold",
            "rooms",
            "salaries",
            "tax",
            "total",
            "wages",
        )
    )


def _is_percentage_number_format(number_format: str | None) -> bool:
    return bool(number_format and "%" in number_format)


def _find_sheet(workbook: WorkbookRecord, sheet_name: str) -> WorkbookSheet:
    for sheet in workbook.sheets:
        if sheet.sheet_name == sheet_name:
            return sheet
    raise ValueError(f"Workbook does not contain sheet: {sheet_name}")
