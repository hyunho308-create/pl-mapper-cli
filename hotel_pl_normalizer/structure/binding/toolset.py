"""Reader and submission tools for one period-column binding session."""

from __future__ import annotations

from typing import Any

from openpyxl.utils import get_column_letter

from hotel_pl_normalizer.models.binding import (
    PeriodBinding,
    UnavailablePeriod,
    WorkbookBindings,
)
from hotel_pl_normalizer.models.workbook import WorkbookRecord
from hotel_pl_normalizer.structure.period_headers import (
    column_forbidden_markers,
    column_scenario_markers,
)
from hotel_pl_normalizer.structure.period_headers import (
    header_markers as _header_markers,
)
from hotel_pl_normalizer.structure.representation import (
    infer_label_layout,
    select_row_label,
)

from .checks import check_bindings
from .column_stats import column_stats

MAX_ROWS_PER_READ = 60
MAX_FIND_HITS = 40
MAX_PREVIEW_LINES = 6

# A row cap alone does not bound a read: hotel P&Ls get wide, and 57 rows of an
# 85-column sheet came back as 125 KB -- about 31,000 tokens. That lands in the
# conversation and is re-sent on every later turn, so one such call cost WAPC
# most of its 1.47 M input tokens. Cells are what the payload is actually made
# of, so cells are what is capped; the reader is told it was cut and can ask for
# a narrower range.
MAX_CELLS_PER_READ = 1_200

# `read_headers` reads the top of many sheets at once. Bounded on both axes
# because the payload is the product of the two: twenty sheets by twenty rows is
# already a large tool result, and the point is to make full coverage affordable
# rather than to move the whole workbook into the prompt.
MAX_SHEETS_PER_HEADER_READ = 20
MAX_HEADER_ROWS = 30
DEFAULT_HEADER_ROWS = 15
LAYOUT_HEADER_ROWS = 30
MAX_LAYOUT_SUBMISSIONS = 3


def _layout_binding_submission_schema() -> dict[str, Any]:
    layout_binding = {
        "type": "object",
        "additionalProperties": False,
        "properties": {
            "layout_id": {"type": "string"},
            "period_id": {"type": "string"},
            "excel_column": {"type": "string"},
            "evidence": {"type": "array", "items": {"type": "string"}},
        },
        "required": ["layout_id", "period_id", "excel_column"],
    }
    layout_unavailable = {
        "type": "object",
        "additionalProperties": False,
        "properties": {
            "layout_id": {"type": "string"},
            "period_id": {"type": "string"},
            "reason": {"type": "string"},
        },
        "required": ["layout_id", "period_id", "reason"],
    }
    sheet_binding = {
        "type": "object",
        "additionalProperties": False,
        "properties": {
            "sheet_name": {"type": "string"},
            "period_id": {"type": "string"},
            "excel_column": {"type": "string"},
            "evidence": {"type": "array", "items": {"type": "string"}},
        },
        "required": ["sheet_name", "period_id", "excel_column"],
    }
    sheet_unavailable = {
        "type": "object",
        "additionalProperties": False,
        "properties": {
            "sheet_name": {"type": "string"},
            "period_id": {"type": "string"},
            "reason": {"type": "string"},
        },
        "required": ["sheet_name", "period_id", "reason"],
    }
    return {
        "type": "object",
        "additionalProperties": False,
        "properties": {
            "layout_bindings": {"type": "array", "items": layout_binding},
            "layout_unavailable": {"type": "array", "items": layout_unavailable},
            "sheet_bindings": {"type": "array", "items": sheet_binding},
            "sheet_unavailable": {"type": "array", "items": sheet_unavailable},
            "notes": {"type": "array", "items": {"type": "string"}},
        },
        "required": ["layout_bindings", "layout_unavailable"],
    }


class PeriodBindingToolset:
    """Reader tools plus one period-binding submission."""

    # Tool results are reads of one specific record, and the prompt names the
    # workbook, so a cached session would be safe -- but the periods vary per
    # run and the prompt would have to carry them all. Not worth the subtlety.
    cacheable = False

    # How many times a submission may be refused before its next well-formed
    # submission is taken anyway. A check that keeps refusing is not teaching the
    # model anything, and a session that argues until it runs out of turns
    # returns *nothing* -- which is strictly worse than the imperfect answer it
    # was refusing. Hotel Kabuki burned all forty turns this way.
    MAX_REJECTIONS_PER_PHASE = 2

    # The coverage gate gets its own, larger allowance. Every other rejection
    # asks the model to reconsider a judgement, which it may reasonably keep;
    # this one asks it to go and read, which is unambiguous and always available.
    # Two attempts is the wrong budget for an instruction that cannot be misread.
    MAX_COVERAGE_REFUSALS = 5

    def __init__(
        self,
        workbook: WorkbookRecord,
        *,
        period_ids: list[str],
        financial_sheets: list[str] | None = None,
        controlling_summary_sheet: str | None = None,
        max_reads: int = 120,
    ) -> None:
        self.workbook = workbook
        self.period_ids = list(period_ids)
        self.sheets = {sheet.sheet_name: sheet for sheet in workbook.sheets}
        # Routing defines the binding scope. Other sheets remain readable for
        # context, but they cannot contribute a binding outcome.
        self.financial_sheets = [
            name for name in (financial_sheets or list(self.sheets)) if name in self.sheets
        ]
        self.controlling_summary_sheet = (
            controlling_summary_sheet
            if controlling_summary_sheet in self.financial_sheets
            else None
        )
        self.max_reads = max_reads
        self.reads = 0
        # Only `read_rows` and `column_stats` count as opening a sheet. A
        # `find_rows` hit does not: a workbook-wide search returns cells from
        # twenty tabs without the model having seen any of their header blocks,
        # and "which column is December" is a question only the header answers.
        self.opened_sheets: set[str] = set()
        self.submission: WorkbookBindings | None = None
        self.rejections: list[str] = []
        self.observations: list[str] = []
        self._refusals: dict[str, int] = {}
        self.layout_groups: list[dict[str, Any]] | None = None
        self.layout_submission_count = 0
        self.header_end_rows = {
            name: _header_end_row(self.sheets[name]) for name in self.financial_sheets
        }
        self.value_column_counts = {
            name: _labelled_column_counts(
                self.sheets[name], self.header_end_rows[name]
            )
            for name in self.financial_sheets
        }

    def signature(self) -> str:
        return f"period_binding:{self.workbook.workbook_id}"

    # -- declarations -----------------------------------------------------

    def declarations(self) -> list[dict[str, Any]]:
        sheet_name = {
            "type": "string",
            "description": "Exact sheet name, as returned by list_sheets.",
            "enum": list(self.sheets),
        }
        return [
            {
                "name": "list_sheet_layouts",
                "description": (
                    "Group routed financial sheets only when their deterministic "
                    "period-header markers occupy the same Excel columns. Returns "
                    "representative headers and candidate value-column profiles. "
                    "Start here; sheets without a safe match remain singleton layouts."
                ),
                "parameters": {"type": "object", "properties": {}},
            },
            {
                "name": "list_sheets",
                "description": (
                    "Every sheet in the workbook, whether routing selected it, "
                    "its extent, and a preview of the first words of text on it. "
                    "Start here."
                ),
                "parameters": {"type": "object", "properties": {}},
            },
            {
                "name": "read_rows",
                "description": (
                    "Populated cells in a row range of one sheet, with their "
                    f"coordinates. At most {MAX_ROWS_PER_READ} rows per call."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "sheet_name": sheet_name,
                        "start_row": {"type": "integer", "minimum": 1},
                        "end_row": {"type": "integer", "minimum": 1},
                    },
                    "required": ["sheet_name", "start_row"],
                },
            },
            {
                "name": "read_headers",
                "description": (
                    "The top rows of several sheets in ONE call, with "
                    "coordinates. This is the cheap way to answer for every "
                    f"sheet: name up to {MAX_SHEETS_PER_HEADER_READ} of them and "
                    "compare their header blocks side by side. Counts as having "
                    "opened each one."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "sheet_names": {
                            "type": "array",
                            "items": {"type": "string"},
                            "description": (
                                "Exact sheet names, as returned by list_sheets."
                            ),
                        },
                        "rows": {
                            "type": "integer",
                            "minimum": 1,
                            "maximum": MAX_HEADER_ROWS,
                            "description": (
                                f"How many rows from the top. Default "
                                f"{DEFAULT_HEADER_ROWS}; raise it when a header "
                                "block sits deeper."
                            ),
                        },
                    },
                    "required": ["sheet_names"],
                },
            },
            {
                "name": "find_rows",
                "description": (
                    "Where a string appears, case-insensitive, with coordinates "
                    "and the sheet each hit is on. Use it to locate a period "
                    "label without guessing a row number."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "query": {"type": "string"},
                        "sheet_name": sheet_name,
                    },
                    "required": ["query"],
                },
            },
            {
                "name": "column_stats",
                "description": (
                    "What every numeric column holds over a row range you name: "
                    "how many values, how many non-zero, how many are "
                    "percentage-formatted, their magnitudes, and real (label, "
                    "value) samples. Use it to check a column carries money "
                    "before binding it."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "sheet_name": sheet_name,
                        "start_row": {"type": "integer", "minimum": 1},
                        "end_row": {"type": "integer", "minimum": 1},
                    },
                    "required": ["sheet_name", "start_row", "end_row"],
                },
            },
            {
                "name": "submit_layout_bindings",
                "description": (
                    "Bind every selected period once per deterministic sheet layout. "
                    "Python expands those choices to sheets, with optional sheet-level "
                    "overrides, and runs the existing exact per-sheet verifier."
                ),
                "parameters": _layout_binding_submission_schema(),
            },
        ]

    # -- dispatch ---------------------------------------------------------

    def dispatch(self, name: str, arguments: dict[str, Any]) -> dict[str, Any]:
        handlers = {
            "list_sheet_layouts": lambda: self._list_sheet_layouts(),
            "list_sheets": lambda: self._list_sheets(),
            "read_rows": lambda: self._read_rows(arguments),
            "read_headers": lambda: self._read_headers(arguments),
            "find_rows": lambda: self._find_rows(arguments),
            "column_stats": lambda: self._column_stats(arguments),
            "submit_layout_bindings": lambda: self._submit_layout_bindings(arguments),
            # Hidden compatibility path for focused tests and saved replays. Live
            # sessions see only the compact layout submission tool.
            "submit_bindings": lambda: self._submit_bindings(arguments),
        }
        handler = handlers.get(name)
        if handler is None:
            return {
                "ok": False,
                "error": f"Unknown tool {name!r}.",
                "instruction": f"Call one of: {', '.join(handlers)}.",
            }
        return handler()

    def terminal_result(self, name: str, result: dict[str, Any]):
        """End the session once the binding submission is accepted."""
        if name == "submit_layout_bindings" and result.get("accepted"):
            return result.get("structure")
        return None

    @staticmethod
    def final_response_error(_result: WorkbookBindings) -> str:
        """A bare final object cannot bypass the tool's coverage verifier."""
        return (
            "Do not end with bare JSON. Call submit_layout_bindings with the complete "
            "layout outcomes and any sheet overrides; only accepted=true completes "
            "this stage."
        )

    # -- readers ----------------------------------------------------------

    def _list_sheet_layouts(self) -> dict[str, Any]:
        """Return conservative header-equivalent groups for routed sheets."""
        if self.layout_groups is None:
            self.layout_groups = self._build_sheet_layouts()
        # Building the groups inspects every routed header deterministically.
        # The model is making one claim about that exact shared header, not
        # guessing from another unexamined sheet.
        self.opened_sheets.update(self.financial_sheets)
        return {
            "ok": True,
            "layout_groups": self.layout_groups,
            "selected_period_ids": self.period_ids,
            "instruction": (
                "Return one layout outcome for every layout_id and selected period. "
                "Use a sheet override only when a named member is a real exception."
            ),
        }

    def _build_sheet_layouts(self) -> list[dict[str, Any]]:
        grouped: dict[tuple[Any, ...], list[str]] = {}
        singleton = 0
        for name in self.financial_sheets:
            sheet = self.sheets[name]
            signature = _sheet_header_signature(sheet)
            if not signature:
                # No semantic header evidence means no safe equivalence claim.
                singleton += 1
                key: tuple[Any, ...] = ("singleton", singleton, name)
            else:
                # Used-range tails vary with blank formatting and technical
                # export cells. The semantic marker/column signature, not
                # max_column, defines whether a period choice can be shared.
                key = ("header", signature)
            grouped.setdefault(key, []).append(name)

        layouts: list[dict[str, Any]] = []
        for index, names in enumerate(grouped.values(), start=1):
            sheets = [self.sheets[name] for name in names]
            representative = sheets[0]
            columns = sorted(
                {
                    column
                    for sheet in sheets
                    for column in self.value_column_counts[sheet.sheet_name]
                }
            )
            candidates = []
            for column in columns:
                numeric = [
                    sheet.sheet_name
                    for sheet in sheets
                    if self.value_column_counts[sheet.sheet_name]
                    .get(column, {})
                    .get("numeric", 0)
                ]
                nonzero = [
                    sheet.sheet_name
                    for sheet in sheets
                    if self.value_column_counts[sheet.sheet_name]
                    .get(column, {})
                    .get("nonzero", 0)
                ]
                candidates.append(
                    {
                        "excel_column": get_column_letter(column),
                        "header_text": _column_header_text(representative, column),
                        "scenario_hints": _column_scenario_hints(
                            representative, column
                        ),
                        "forbidden_hints": sorted(
                            column_forbidden_markers(representative, column)
                        ),
                        "sheets_with_numbers": len(numeric),
                        "sheets_with_nonzero": len(nonzero),
                    }
                )
            layouts.append(
                {
                    "layout_id": f"layout_{index}",
                    "representative_sheet": representative.sheet_name,
                    "sheet_count": len(names),
                    "sheet_names": names,
                    "header_cells": _semantic_header_cells(representative),
                    "candidate_columns": candidates,
                }
            )
        return layouts

    def _list_sheets(self) -> dict[str, Any]:
        routed = set(self.financial_sheets)
        return {
            "ok": True,
            "filename": self.workbook.source.original_filename,
            "sheets": [
                {
                    "sheet_name": sheet.sheet_name,
                    "routed_as_financial": sheet.sheet_name in routed,
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column,
                    "preview": self._preview(sheet),
                }
                for sheet in self.workbook.sheets
            ],
        }

    def _preview(self, sheet) -> list[str]:
        """A few identifying words from the top of a sheet.

        A coded tab name carries no signal on its own -- the real title sits
        inside the sheet, like `0232 - Group Banquet`.
        """
        found: list[str] = []
        for row in sheet.rows[:12]:
            for cell in row.cells:
                text = str(cell.display_value or cell.raw_value or "").strip()
                if len(text) > 2 and any(c.isalpha() for c in text) and text not in found:
                    found.append(text)
                    if len(found) >= MAX_PREVIEW_LINES:
                        return found
        return found

    def _budget_spent(self) -> dict[str, Any] | None:
        if self.reads < self.max_reads:
            return None
        return {
            "ok": False,
            "error": f"Read budget of {self.max_reads} calls is spent.",
            "instruction": "Submit what you have. A partial answer is useful.",
        }

    def _sheet_or_error(self, sheet_name: str):
        sheet = self.sheets.get(sheet_name)
        if sheet is None:
            return None, {
                "ok": False,
                "error": f"No sheet named {sheet_name!r}.",
                "instruction": "Use a name from list_sheets.",
            }
        return sheet, None

    def _read_rows(self, arguments: dict[str, Any]) -> dict[str, Any]:
        if (spent := self._budget_spent()) is not None:
            return spent
        sheet, error = self._sheet_or_error(str(arguments.get("sheet_name") or ""))
        if error is not None:
            return error
        start = max(1, int(arguments.get("start_row") or 1))
        end = arguments.get("end_row")
        end = int(end) if end is not None else start + MAX_ROWS_PER_READ - 1
        end = min(end, start + MAX_ROWS_PER_READ - 1)
        self.reads += 1
        self.opened_sheets.add(sheet.sheet_name)

        rows: list[dict[str, Any]] = []
        cells_returned = 0
        truncated_at: int | None = None
        for row in sheet.rows:
            if not start <= row.row_index <= end:
                continue
            cells = [
                {"at": cell.address, "value": _text(cell)}
                for cell in row.cells
                if _text(cell) is not None
            ]
            if not cells:
                continue
            if cells_returned + len(cells) > MAX_CELLS_PER_READ and rows:
                truncated_at = row.row_index
                break
            rows.append({"row": row.row_index, "cells": cells})
            cells_returned += len(cells)

        result: dict[str, Any] = {
            "ok": True,
            "sheet_name": sheet.sheet_name,
            "merged_ranges": [merged.range for merged in sheet.merged_ranges[:40]],
            "rows": rows,
        }
        if truncated_at is not None:
            # Said plainly so a silently short read cannot hide a header row.
            result["truncated_before_row"] = truncated_at
            result["instruction"] = (
                f"This sheet is {sheet.max_column} columns wide, so the read was "
                f"cut at {MAX_CELLS_PER_READ} cells. Rows {truncated_at}-{end} "
                "were not returned. Ask for a narrower row range to see them."
            )
        return result

    def _read_headers(self, arguments: dict[str, Any]) -> dict[str, Any]:
        """The top of many sheets at once.

        Full coverage was costing one tool call per sheet, and a session with
        twenty-five routed tabs would rather submit a short answer than spend
        twenty-five turns earning a complete one -- which is exactly what one run
        did, growing its answer two bindings at a time until it was accepted.
        Making the reads cheap is a better fix than insisting harder.

        Folded into one call for the same reason `list_sheets` folds a preview of
        every sheet into itself: the model should not have to remember to do
        something the workbook is already open for.
        """
        if (spent := self._budget_spent()) is not None:
            return spent
        requested = arguments.get("sheet_names") or []
        if isinstance(requested, str):
            requested = [requested]
        known = [str(name) for name in requested if str(name) in self.sheets]
        unknown = [str(name) for name in requested if str(name) not in self.sheets]
        if not known:
            return {
                "ok": False,
                "error": f"None of those sheets are in this workbook: {', '.join(unknown[:8])}.",
                "instruction": "Use names from list_sheets.",
            }
        rows = int(arguments.get("rows") or DEFAULT_HEADER_ROWS)
        rows = max(1, min(rows, MAX_HEADER_ROWS))
        taken = known[:MAX_SHEETS_PER_HEADER_READ]
        self.reads += 1
        self.opened_sheets.update(taken)

        result: dict[str, Any] = {
            "ok": True,
            "rows_per_sheet": rows,
            "sheets": [
                {
                    "sheet_name": name,
                    "merged_ranges": [
                        merged.range for merged in self.sheets[name].merged_ranges[:12]
                    ],
                    "rows": [
                        {
                            "row": row.row_index,
                            "cells": [
                                {"at": cell.address, "value": _text(cell)}
                                for cell in row.cells
                                if _text(cell) is not None
                            ],
                        }
                        for row in self.sheets[name].rows
                        if row.row_index <= rows
                        and any(_text(cell) is not None for cell in row.cells)
                    ],
                }
                for name in taken
            ],
        }
        if unknown:
            result["ignored_unknown_sheets"] = unknown[:8]
        if len(known) > MAX_SHEETS_PER_HEADER_READ:
            result["not_read"] = known[MAX_SHEETS_PER_HEADER_READ:]
            result["instruction"] = (
                f"Only the first {MAX_SHEETS_PER_HEADER_READ} were read. Call "
                "read_headers again for the rest."
            )
        return result

    def _find_rows(self, arguments: dict[str, Any]) -> dict[str, Any]:
        if (spent := self._budget_spent()) is not None:
            return spent
        query = str(arguments.get("query") or "").strip().lower()
        if not query:
            return {"ok": False, "error": "find_rows needs a query."}
        requested = arguments.get("sheet_name")
        if requested and requested not in self.sheets:
            return {
                "ok": False,
                "error": f"No sheet named {requested!r}.",
                "instruction": "Use a name from list_sheets, or omit it to search all.",
            }
        names = [requested] if requested else list(self.sheets)
        self.reads += 1
        hits = []
        for name in names:
            for row in self.sheets[name].rows:
                for cell in row.cells:
                    text = _text(cell)
                    if text and query in text.lower():
                        hits.append(
                            {"sheet_name": name, "at": cell.address, "value": text}
                        )
                        if len(hits) >= MAX_FIND_HITS:
                            return {"ok": True, "query": query, "hits": hits}
        return {"ok": True, "query": query, "hits": hits}

    def _column_stats(self, arguments: dict[str, Any]) -> dict[str, Any]:
        if (spent := self._budget_spent()) is not None:
            return spent
        sheet, error = self._sheet_or_error(str(arguments.get("sheet_name") or ""))
        if error is not None:
            return error
        start = max(1, int(arguments.get("start_row") or 1))
        end = int(arguments.get("end_row") or start)
        self.reads += 1
        self.opened_sheets.add(sheet.sheet_name)
        return column_stats(sheet, start, end).as_dict()

    # -- submission -------------------------------------------------------

    def _submit_layout_bindings(self, arguments: dict[str, Any]) -> dict[str, Any]:
        """Expand compact layout choices into the existing sheet-level contract."""
        self.layout_submission_count += 1
        if self.layout_submission_count > MAX_LAYOUT_SUBMISSIONS:
            raise RuntimeError(
                "Excel layout binding exceeded one initial submission and two repairs."
            )
        if self.layout_groups is None:
            self.layout_groups = self._build_sheet_layouts()
        layouts = {item["layout_id"]: item for item in self.layout_groups}
        layout_for_sheet = {
            name: layout_id
            for layout_id, layout in layouts.items()
            for name in layout["sheet_names"]
        }
        layout_bindings = arguments.get("layout_bindings") or []
        layout_unavailable = arguments.get("layout_unavailable") or []
        sheet_bindings = arguments.get("sheet_bindings") or []
        sheet_unavailable = arguments.get("sheet_unavailable") or []
        arrays = {
            "layout_bindings": layout_bindings,
            "layout_unavailable": layout_unavailable,
            "sheet_bindings": sheet_bindings,
            "sheet_unavailable": sheet_unavailable,
        }
        if any(not isinstance(value, list) for value in arrays.values()):
            return self._reject(
                "All compact binding collections must be arrays.",
                "submit_layout_bindings",
            )

        outcomes: dict[tuple[str, str], tuple[str, dict[str, Any]]] = {}
        for kind, items in (
            ("binding", layout_bindings),
            ("unavailable", layout_unavailable),
        ):
            for item in items:
                if not isinstance(item, dict):
                    return self._reject(
                        f"Each layout {kind} must be an object.",
                        "submit_layout_bindings",
                    )
                layout_id = str(item.get("layout_id") or "")
                period_id = str(item.get("period_id") or "")
                if layout_id not in layouts:
                    return self._reject(
                        f"Unknown layout_id {layout_id!r}.",
                        "submit_layout_bindings",
                    )
                if period_id not in self.period_ids:
                    return self._reject(
                        f"Unknown selected period_id {period_id!r}.",
                        "submit_layout_bindings",
                    )
                pair = (layout_id, period_id)
                if pair in outcomes:
                    return self._reject(
                        f"{layout_id} and {period_id} have more than one outcome.",
                        "submit_layout_bindings",
                    )
                outcomes[pair] = (kind, item)

        required = {
            (layout_id, period_id)
            for layout_id in layouts
            for period_id in self.period_ids
        }
        missing = sorted(required - set(outcomes))
        if missing:
            preview = ", ".join(f"{layout}/{period}" for layout, period in missing[:12])
            return self._reject(
                "Every layout and selected period requires one compact outcome. "
                f"Missing: {preview}.",
                "submit_layout_bindings",
            )

        overrides: dict[tuple[str, str], tuple[str, dict[str, Any]]] = {}
        for kind, items in (
            ("binding", sheet_bindings),
            ("unavailable", sheet_unavailable),
        ):
            for item in items:
                if not isinstance(item, dict):
                    return self._reject(
                        f"Each sheet {kind} override must be an object.",
                        "submit_layout_bindings",
                    )
                sheet_name = str(item.get("sheet_name") or "")
                period_id = str(item.get("period_id") or "")
                if sheet_name not in layout_for_sheet:
                    return self._reject(
                        f"Sheet override {sheet_name!r} is outside routed layouts.",
                        "submit_layout_bindings",
                    )
                if period_id not in self.period_ids:
                    return self._reject(
                        f"Unknown selected period_id {period_id!r} in sheet override.",
                        "submit_layout_bindings",
                    )
                pair = (sheet_name, period_id)
                if pair in overrides:
                    return self._reject(
                        f"{sheet_name!r} and {period_id} have multiple overrides.",
                        "submit_layout_bindings",
                    )
                overrides[pair] = (kind, item)

        bindings: list[PeriodBinding] = []
        unavailable: list[UnavailablePeriod] = []
        notes = [str(value) for value in arguments.get("notes") or []]
        for (layout_id, period_id), (kind, item) in outcomes.items():
            layout = layouts[layout_id]
            if kind == "binding":
                column = str(item.get("excel_column") or "").strip().upper()
                candidates = {
                    value["excel_column"]: value
                    for value in layout["candidate_columns"]
                }
                if column not in candidates:
                    return self._reject(
                        f"{layout_id} has no listed numeric candidate column {column!r}.",
                        "submit_layout_bindings",
                    )
                forbidden_hints = set(candidates[column]["forbidden_hints"])
                if forbidden_hints:
                    return self._reject(
                        f"{layout_id} column {column} is explicitly marked as "
                        f"{', '.join(sorted(forbidden_hints))}, so it is not a "
                        "period amount column.",
                        "submit_layout_bindings",
                    )
                scenario_hints = set(candidates[column]["scenario_hints"])
                allowed_hints = _allowed_scenario_hints(period_id)
                if scenario_hints and not scenario_hints & allowed_hints:
                    return self._reject(
                        f"{layout_id} column {column} has explicit scenario hints "
                        f"{sorted(scenario_hints)}, which do not match {period_id}.",
                        "submit_layout_bindings",
                    )
                evidence = [str(value) for value in item.get("evidence") or []]
            else:
                reason = str(item.get("reason") or "").strip()
                if not reason:
                    return self._reject(
                        f"Unavailable outcome for {layout_id}/{period_id} needs a reason.",
                        "submit_layout_bindings",
                    )

            for sheet_name in layout["sheet_names"]:
                override = overrides.get((sheet_name, period_id))
                if override is not None:
                    override_kind, override_item = override
                    if override_kind == "binding":
                        override_column = str(
                            override_item.get("excel_column") or ""
                        ).strip().upper()
                        if not override_column.isalpha():
                            return self._reject(
                                f"Sheet override for {sheet_name!r} requires a valid "
                                "Excel column.",
                                "submit_layout_bindings",
                            )
                        override_forbidden = column_forbidden_markers(
                            self.sheets[sheet_name],
                            _excel_column_number(override_column),
                        )
                        if override_forbidden:
                            return self._reject(
                                f"Sheet override {sheet_name!r} column "
                                f"{override_column} is explicitly marked as "
                                f"{', '.join(sorted(override_forbidden))}, so it is "
                                "not a period amount column.",
                                "submit_layout_bindings",
                            )
                        override_hints = set(
                            _column_scenario_hints(
                                self.sheets[sheet_name],
                                _excel_column_number(override_column),
                            )
                        )
                        if override_hints and not override_hints & _allowed_scenario_hints(
                            period_id
                        ):
                            return self._reject(
                                f"Sheet override {sheet_name!r} column "
                                f"{override_column} has explicit scenario hints "
                                f"{sorted(override_hints)}, which do not match "
                                f"{period_id}.",
                                "submit_layout_bindings",
                            )
                        bindings.append(
                            PeriodBinding(
                                sheet_name=sheet_name,
                                period_id=period_id,
                                excel_column=override_column,
                                evidence=[
                                    str(value)
                                    for value in override_item.get("evidence") or []
                                ],
                            )
                        )
                    else:
                        override_reason = str(
                            override_item.get("reason") or ""
                        ).strip()
                        if not override_reason:
                            return self._reject(
                                f"Unavailable override for {sheet_name!r}/{period_id} "
                                "needs a reason.",
                                "submit_layout_bindings",
                            )
                        unavailable.append(
                            UnavailablePeriod(
                                sheet_name=sheet_name,
                                period_id=period_id,
                                reason=override_reason,
                            )
                        )
                    continue
                if kind == "unavailable":
                    unavailable.append(
                        UnavailablePeriod(
                            sheet_name=sheet_name,
                            period_id=period_id,
                            reason=reason,
                        )
                    )
                    continue
                column_number = _excel_column_number(column)
                counts = self.value_column_counts[sheet_name].get(column_number, {})
                if counts.get("nonzero", 0):
                    bindings.append(
                        PeriodBinding(
                            sheet_name=sheet_name,
                            period_id=period_id,
                            excel_column=column,
                            evidence=evidence,
                        )
                    )
                else:
                    state = (
                        "all zero"
                        if counts.get("numeric", 0)
                        else "blank"
                    )
                    unavailable.append(
                        UnavailablePeriod(
                            sheet_name=sheet_name,
                            period_id=period_id,
                            reason=(
                                f"Layout {layout_id} selected column {column}, but the "
                                f"column is {state} on this sheet."
                            ),
                        )
                    )

        expanded = WorkbookBindings(
            bindings=bindings,
            unavailable=unavailable,
            notes=notes,
        )
        if self.controlling_summary_sheet is not None:
            bound_on_anchor = {
                item.period_id
                for item in expanded.bindings
                if item.sheet_name == self.controlling_summary_sheet
            }
            missing_anchor = [
                period_id
                for period_id in self.period_ids
                if period_id not in bound_on_anchor
            ]
            if missing_anchor:
                return self._reject(
                    f"Discovery confirmed every selected period on controlling summary "
                    f"{self.controlling_summary_sheet!r}, but it has no usable binding "
                    f"for: {', '.join(missing_anchor)}. Re-read that summary header and "
                    "bind its displayed amount columns.",
                    "submit_layout_bindings",
                )
        return self._submit_bindings(
            expanded.model_dump(mode="json"),
            submission_tool="submit_layout_bindings",
        )

    def _submit_bindings(
        self,
        arguments: dict[str, Any],
        *,
        submission_tool: str = "submit_bindings",
    ) -> dict[str, Any]:
        try:
            submission = WorkbookBindings.model_validate(arguments)
        except Exception as exc:  # noqa: BLE001 - the message goes back to the model
            return self._reject(
                f"That did not match the schema: {exc}", submission_tool
            )

        # Before anything about the columns: did the session actually open the
        # sheets it is making claims about? Asking in the prompt does not carry
        # this. Exploration measured it -- 24 of 29 sessions read exactly one
        # sheet after being asked in text for five -- and the corpus measured it
        # again here: the fast model bound columns on 28 tabs it never opened,
        # and got the period block wrong on every one of them.
        unread = self._unopened_claims(submission)
        unanswered = self._unanswered_pairs(submission)
        if unread or unanswered:
            self._refusals["coverage"] = self._refusals.get("coverage", 0) + 1
            if self._refusals["coverage"] <= self.MAX_COVERAGE_REFUSALS:
                return self._coverage_refusal(unread, unanswered)
            # Out of patience. Keep only claims grounded in opened sheets, then
            # fail closed for every unresolved sheet-period pair. Never infer a
            # missing column from another sheet's convention.
            if unread:
                submission = _drop_unopened(submission, self.opened_sheets)
                self.observations.append(
                    f"Dropped bindings for {len(unread)} sheet(s) the session never "
                    f"opened ({', '.join(sorted(unread)[:6])})."
                )
            unanswered = self._unanswered_pairs(submission)
            if unanswered:
                submission = _mark_unresolved_unavailable(submission, unanswered)
                self.observations.append(
                    f"Marked {len(unanswered)} unresolved sheet-period pair(s) "
                    "unavailable; no default column was inferred."
                )

        result = check_bindings(
            submission,
            self.sheets,
            period_ids=self.period_ids,
            financial_sheets=self.financial_sheets,
        )
        if not result.accepted and not self._out_of_patience(submission_tool):
            return self._reject(" ".join(result.rejections), submission_tool)

        if result.rejections:
            # Preserve only mechanically valid, unambiguous outcomes, then fail
            # closed for every pair the salvage removed.
            submission = _drop_rejected_bindings(
                submission,
                self.sheets,
                self.period_ids,
                self.financial_sheets,
            )
            unresolved = self._unanswered_pairs(submission)
            submission = _mark_unresolved_unavailable(submission, unresolved)
            self.observations.extend(
                f"Removed during deterministic salvage: {message}"
                for message in result.rejections
            )
            if unresolved:
                self.observations.append(
                    f"Marked {len(unresolved)} invalid or ambiguous sheet-period "
                    "pair(s) unavailable; no default column was inferred."
                )
        self.observations.extend(result.observations)
        self.submission = submission.model_copy(
            update={"observations": list(self.observations)}
        )
        return {
            "ok": True,
            "accepted": True,
            "structure": self.submission.model_dump(mode="json"),
        }

    def _unanswered_pairs(self, submission: WorkbookBindings) -> set[tuple[str, str]]:
        """Routed sheet-period pairs this submission says nothing about.

        The hole the first version of the read gate left. It asked only that
        claims be backed by reads, which a *small* answer satisfies trivially --
        and one run found that out, growing its submission two bindings at a time
        and being accepted at five sheets of twenty-five. Routing already decided
        these sheets hold P&L content, so silence about one is not an answer;
        `unavailable` is always available where a period genuinely is not there.
        """
        answered = {
            (binding.sheet_name, binding.period_id) for binding in submission.bindings
        } | {(item.sheet_name, item.period_id) for item in submission.unavailable}
        return {
            (name, period_id)
            for name in self.financial_sheets
            for period_id in self.period_ids
            if (name, period_id) not in answered
        }

    def _coverage_refusal(
        self,
        unread: set[str],
        unanswered: set[tuple[str, str]],
    ) -> dict[str, Any]:
        """One message covering the binding and read obligations."""
        parts: list[str] = []
        if unanswered:
            examples = ", ".join(
                f"{sheet} [{period}]"
                for sheet, period in sorted(unanswered)[:12]
            )
            parts.append(
                f"{len(unanswered)} routed sheet-period pair(s) have no binding "
                f"and no unavailable note: {examples}. Routing "
                "decided these hold P&L content, so each one needs an answer for "
                "every chosen period -- a column, or an unavailable note saying "
                "what you saw instead."
            )
        if unread:
            parts.append(
                f"{len(unread)} sheet(s) in the submission have not been opened: "
                f"{', '.join(sorted(unread)[:12])}. Which column holds a period is "
                "a question only that sheet's header rows answer."
            )
        parts.append(
            f"Use read_headers to open up to {MAX_SHEETS_PER_HEADER_READ} sheets "
            "in a single call and compare their header blocks side by side, then "
            "submit once with every sheet answered for. Do not shrink the "
            "submission to satisfy this: a sheet left out is excluded from "
            "mapping for that period rather than assigned another sheet's column."
        )
        message = " ".join(parts)
        self.rejections.append(message)
        return {
            "ok": True,
            "accepted": False,
            "error": message,
            "opened_so_far": sorted(self.opened_sheets),
            "sheets_to_answer_for": self.financial_sheets,
            "instruction": "Read what you are missing, then call submit_bindings again.",
        }

    def _unopened_claims(self, submission: WorkbookBindings) -> set[str]:
        """Sheets this submission makes a claim about without having opened.

        Both a binding and an `unavailable` note are claims: saying a period is
        not on a sheet is as much an assertion about its headers as saying which
        column holds it.
        """
        claimed = {binding.sheet_name for binding in submission.bindings} | {
            item.sheet_name for item in submission.unavailable
        }
        return {
            name
            for name in claimed
            if name in self.sheets and name not in self.opened_sheets
        }

    def _out_of_patience(self, tool: str) -> bool:
        """Has this submission been refused often enough to stop refusing it?"""
        return self._refusals.get(tool, 0) >= self.MAX_REJECTIONS_PER_PHASE

    def _reject(self, message: str, tool: str) -> dict[str, Any]:
        self.rejections.append(message)
        self._refusals[tool] = self._refusals.get(tool, 0) + 1
        remaining = self.MAX_REJECTIONS_PER_PHASE - self._refusals[tool]
        instruction = f"Correct it and call {tool} again."
        if remaining <= 0:
            instruction += (
                " This is the last time this will be refused -- the next "
                "well-formed answer is taken as it stands, so send your best "
                "one rather than stopping."
            )
        return {
            "ok": True,
            "accepted": False,
            "error": message,
            "instruction": instruction,
        }

    def best_effort(self) -> WorkbookBindings | None:
        """Whatever valid submission this session established."""
        if self.submission is not None:
            return self.submission
        return None


def _drop_unopened(
    submission: WorkbookBindings, opened: set[str]
) -> WorkbookBindings:
    """Remove claims about sheets the session never opened.

    Safer than keeping them. A dropped binding leaves the sheet on the workbook's
    most common column, which is right far more often than a column chosen
    without looking; a kept one puts a specific wrong figure in front of the
    mapper, which nothing downstream can detect.
    """
    return submission.model_copy(
        update={
            "bindings": [
                binding
                for binding in submission.bindings
                if binding.sheet_name in opened
            ],
            "unavailable": [
                item for item in submission.unavailable if item.sheet_name in opened
            ],
        }
    )


def _drop_rejected_bindings(
    submission: WorkbookBindings,
    sheets,
    period_ids: list[str],
    financial_sheets: list[str],
) -> WorkbookBindings:
    """Keep the bindings that stand on their own, drop the ones that cannot.

    Used only when patience has run out. A binding naming a sheet that is not
    here, or a column with no numbers, would put a wrong figure in front of the
    mapper -- which is the one outcome worse than no figure -- so those go. The
    rest are kept, because they were never the problem.
    """
    chosen = set(period_ids)
    financial = set(financial_sheets)
    unavailable = []
    unavailable_pairs: set[tuple[str, str]] = set()
    for item in submission.unavailable:
        pair = (item.sheet_name, item.period_id)
        if (
            pair in unavailable_pairs
            or item.period_id not in chosen
            or item.sheet_name not in sheets
            or item.sheet_name not in financial
            or not item.reason.strip()
        ):
            continue
        unavailable.append(item)
        unavailable_pairs.add(pair)

    candidates = []
    for binding in submission.bindings:
        sheet = sheets.get(binding.sheet_name)
        pair = (binding.sheet_name, binding.period_id)
        if (
            sheet is None
            or binding.sheet_name not in financial
            or binding.period_id not in chosen
            or pair in unavailable_pairs
        ):
            continue
        letters = (binding.excel_column or "").strip().upper()
        if not letters.isalpha():
            continue
        column = _excel_column_number(letters)
        if not _sheet_column_holds_numbers(sheet, column):
            unavailable.append(
                UnavailablePeriod(
                    sheet_name=binding.sheet_name,
                    period_id=binding.period_id,
                    reason=(
                        f"Column {letters} was inspected and contains no numeric "
                        "values for the selected period."
                    ),
                )
            )
            unavailable_pairs.add(pair)
            continue
        candidates.append(binding)

    pair_counts: dict[tuple[str, str], int] = {}
    column_periods: dict[tuple[str, str], set[str]] = {}
    for binding in candidates:
        pair = (binding.sheet_name, binding.period_id)
        pair_counts[pair] = pair_counts.get(pair, 0) + 1
        column_periods.setdefault(
            (binding.sheet_name, binding.excel_column.strip().upper()), set()
        ).add(binding.period_id)
    kept = [
        binding
        for binding in candidates
        if pair_counts[(binding.sheet_name, binding.period_id)] == 1
        and len(
            column_periods[
                (binding.sheet_name, binding.excel_column.strip().upper())
            ]
        )
        == 1
    ]
    return submission.model_copy(
        update={"bindings": kept, "unavailable": unavailable}
    )


def _mark_unresolved_unavailable(
    submission: WorkbookBindings,
    unresolved: set[tuple[str, str]],
) -> WorkbookBindings:
    """Fail closed for pairs the model did not establish explicitly."""
    if not unresolved:
        return submission
    additions = [
        UnavailablePeriod(
            sheet_name=sheet_name,
            period_id=period_id,
            reason=(
                "Binding was not established before the model session ended; "
                "no column was inferred from another sheet."
            ),
        )
        for sheet_name, period_id in sorted(unresolved)
    ]
    return submission.model_copy(
        update={"unavailable": [*submission.unavailable, *additions]}
    )


def _excel_column_number(letters: str) -> int:
    number = 0
    for character in letters:
        number = number * 26 + ord(character) - ord("A") + 1
    return number


def _sheet_column_holds_numbers(sheet, column: int) -> bool:
    return any(
        isinstance(cell.raw_value, int | float)
        and not isinstance(cell.raw_value, bool)
        for row in sheet.rows
        for cell in row.cells
        if cell.column == column
    )


def _sheet_column_has_nonzero(
    sheet, column: int, header_end: int | None = None
) -> bool:
    header_end = header_end if header_end is not None else _header_end_row(sheet)
    return any(
        _is_number(cell.raw_value) and float(cell.raw_value) != 0
        for row in sheet.rows
        if row.row_index > header_end
        for cell in row.cells
        if cell.column == column
    )


def _sheet_value_column_holds_numbers(
    sheet, column: int, header_end: int | None = None
) -> bool:
    header_end = header_end if header_end is not None else _header_end_row(sheet)
    return any(
        _is_number(cell.raw_value)
        for row in sheet.rows
        if row.row_index > header_end
        for cell in row.cells
        if cell.column == column
    )


def _labelled_column_counts(sheet, header_end: int) -> dict[int, dict[str, int]]:
    rows = [row for row in sheet.rows if row.row_index > header_end]
    layout = infer_label_layout(rows)
    counts: dict[int, dict[str, int]] = {}
    for row in rows:
        if select_row_label(row, layout).cell is None:
            continue
        for cell in row.cells:
            if not _is_number(cell.raw_value):
                continue
            tally = counts.setdefault(cell.column, {"numeric": 0, "nonzero": 0})
            tally["numeric"] += 1
            if float(cell.raw_value) != 0:
                tally["nonzero"] += 1
    return counts


def _is_number(value: Any) -> bool:
    return isinstance(value, int | float) and not isinstance(value, bool)


def _sheet_header_signature(sheet) -> tuple[Any, ...]:
    header_end = _header_end_row(sheet)
    markers: list[tuple[int, tuple[str, ...]]] = []
    for row in sheet.rows:
        if row.row_index > header_end:
            continue
        for cell in row.cells:
            text = _text(cell)
            if not text:
                continue
            found = _header_markers(text)
            if found:
                markers.append((cell.column, found))
    merge_markers: list[tuple[int, int, tuple[str, ...]]] = []
    cells = {
        (cell.row, cell.column): cell
        for row in sheet.rows
        if row.row_index <= header_end
        for cell in row.cells
    }
    for merged in sheet.merged_ranges:
        if merged.top_left_row > header_end:
            continue
        cell = cells.get((merged.top_left_row, merged.top_left_column))
        text = _text(cell) if cell is not None else None
        found = _header_markers(text) if text else ()
        if found:
            end_column = _merged_end_column(merged.range)
            merge_markers.append((merged.top_left_column, end_column, found))
    if not markers and not merge_markers:
        return ()
    return tuple(sorted(markers)), tuple(sorted(merge_markers))


def _semantic_header_cells(sheet) -> list[dict[str, Any]]:
    header_end = _header_end_row(sheet)
    result: list[dict[str, Any]] = []
    for row in sheet.rows:
        if row.row_index > header_end:
            continue
        for cell in row.cells:
            text = _text(cell)
            markers = _header_markers(text) if text else ()
            if markers:
                result.append(
                    {"at": cell.address, "value": text, "markers": list(markers)}
                )
    return result[:120]


def _column_header_text(sheet, column: int) -> list[str]:
    header_end = _header_end_row(sheet)
    values: list[str] = []
    for row in sheet.rows:
        if row.row_index > header_end:
            continue
        for cell in row.cells:
            if cell.column != column:
                continue
            text = _text(cell)
            if text and text not in values:
                values.append(text)
    for merged in sheet.merged_ranges:
        if merged.top_left_row > header_end:
            continue
        if merged.top_left_column <= column <= _merged_end_column(merged.range):
            value = str(merged.value or "").strip()
            if value and value not in values:
                values.append(value)
    return values[:12]


def _column_scenario_hints(sheet, column: int) -> list[str]:
    return sorted(column_scenario_markers(sheet, column))


def _allowed_scenario_hints(period_id: str) -> set[str]:
    if "_budget" in period_id:
        return {"budget"}
    if "_forecast" in period_id:
        return {"forecast"}
    return {"actual", "prior_year"}


def _merged_end_column(cell_range: str) -> int:
    end = cell_range.split(":", 1)[-1]
    letters = "".join(character for character in end if character.isalpha())
    return _excel_column_number(letters)


def _header_end_row(sheet) -> int:
    """Stop before the first substantive value row, capped for prompt safety."""
    saw_period_header = False
    for row in sheet.rows:
        numeric = [cell for cell in row.cells if _is_number(cell.raw_value)]
        row_markers = {
            marker
            for cell in row.cells
            if (text := _text(cell))
            for marker in _header_markers(text)
        }
        labelled_value_row = bool(numeric) and any(
            (text := _text(cell)) and not _header_markers(text)
            for cell in row.cells
            if not _is_number(cell.raw_value)
        )
        if saw_period_header and (len(numeric) >= 2 or labelled_value_row):
            return max(1, min(LAYOUT_HEADER_ROWS, row.row_index - 1))
        saw_period_header = saw_period_header or bool(
            row_markers
            & {
                "actual",
                "budget",
                "forecast",
                "prior_year",
                "ytd",
                "ptd",
                "ttm",
                "total",
            }
            or any(marker.startswith(("month:", "year:")) for marker in row_markers)
        )
    return LAYOUT_HEADER_ROWS


def _text(cell) -> str | None:
    """One cell as short display text, or None when it holds nothing."""
    value = cell.display_value if cell.display_value is not None else cell.raw_value
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = " ".join(str(value).split())
    return text[:160] or None


__all__ = ["PeriodBindingToolset", "MAX_ROWS_PER_READ"]
