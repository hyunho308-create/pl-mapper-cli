"""Render a normalized workbook in the KHP output format.

Design notes
------------
This is the artifact a client judges the system by, and the second tab is the one
an analyst actually works in, so the constraints are unusually tight.

**The template is authored, not generated.** `data/output_template.xlsx` holds
the COA grid and ~180 hand-built formulas on `KHP Model Accounts`. Rebuilding
those in code would mean maintaining them in two places, where a typo stays
invisible until an analyst hits it. This module opens the template and writes
values into it; it never authors a model formula.

**Row order is load-bearing.** Every model formula is a hard row reference
(`=COA!C178+COA!C180`). The COA tab must therefore carry all 271 accounts in
`coa_v2.csv` order on every run, including unmapped ones, or every figure on the
model tab silently shifts by a row. `_assert_template_matches` refuses to write
rather than let that happen quietly.

**Periods grow rightward into reserved space.** Twenty slots at C-V, metadata
pinned beyond them at W/X/Y. Unused slots are hidden, so the common two-period
run looks exactly like the authored format. The model tab's period column is
translated once per period, which is exact rather than approximate: translating
the authored column C onto D reproduces the authored column D formula for
formula, all 139 of them.

**Say which lines to check.** Accuracy runs ~95%, and the honest move is to mark
the risky lines rather than present every number with equal confidence. Model
Feedback carries validation checks, mapper review items, and accounts nothing
mapped to, per account. Run Notes stays a compact summary of the completed job.
"""

from __future__ import annotations

import csv
import math
import re
import tempfile
import zipfile
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.formula.translate import Translator
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.views import Selection

from hotel_pl_normalizer.atomic import replace_atomically
from hotel_pl_normalizer.mapping import GENERIC_VENUE_SLOTS
from hotel_pl_normalizer.pipeline import NormalizationResult

DATA_DIR = Path(__file__).resolve().parent / "data"
TEMPLATE = DATA_DIR / "output_template.xlsx"
COA_CSV = DATA_DIR / "coa_v2.csv"

# Must match scripts/build_output_template.py, which imports them from here.
HEADER_ROW = 2
FIRST_ACCOUNT_ROW = 3
ID_COL = 2                 # B
FIRST_PERIOD_COL = 3       # C
MAX_PERIODS = 20           # C..V
LABELS_COL = 23            # W
FEEDBACK_COL = 24          # X
VENUE_COL = 25             # Y
MAX_OUTPUT_FILENAME_CHARS = 80

MODEL_SHEET_PART = "xl/worksheets/sheet3.xml"
MODEL_SHEET_RELS = "xl/worksheets/_rels/sheet3.xml.rels"
MODEL_DRAWING_PART = "xl/drawings/drawing1.xml"

SEVERITY_ORDER = {"error": 0, "warning": 1}

SUMMARY_DEPARTMENT_NAMES = {
    "S12.total_departmental_expenses": (
        "Total Departmental Expenses",
        "combined Rooms, F&B, and OOD department",
        True,
    ),
    "S12.total_rooms_revenue": ("Rooms revenue", "Rooms", False),
    "S12.total_rooms_expenses": ("Rooms expenses", "Rooms", True),
    "S12.total_food_and_beverage_revenue": ("F&B revenue", "F&B", False),
    "S12.total_food_and_beverage_expenses": ("F&B expenses", "F&B", True),
    "S12.total_other_operated_departments_revenue": ("OOD revenue", "OOD", False),
    "S12.total_other_operated_departments_expenses": ("OOD expenses", "OOD", True),
    "S12.total_miscellaneous_income": ("miscellaneous income", "Miscellaneous Income", False),
    "S12.total_administrative_and_general_expenses": ("A&G expenses", "A&G", True),
    "S12.total_information_and_telecommunications_systems_expenses": ("IT expenses", "IT", True),
    "S12.total_sales_and_marketing_expenses": ("S&M expenses", "S&M", True),
    "S12.total_property_operation_and_maintenance_expenses": ("POM expenses", "POM", True),
    "S12.total_utilities_expenses": ("Utilities expenses", "Utilities", True),
    "S12.total_management_fees": ("management fees", "Management Fees", True),
    "S12.total_non_operating_income_and_expenses": ("Non-Op total", "Non-Op", False),
}

SUMMARY_MATH_NAMES = {
    "S12.total_revenue": (
        "Summary Total Revenue",
        "Rooms, F&B, OOD, and Miscellaneous Income revenue combined",
        False,
    ),
    "S12.total_departmental_expenses": (
        "Summary Total Departmental Expenses",
        "Rooms, F&B, and OOD expenses combined",
        True,
    ),
    "S12.departmental_profit": (
        "Summary Departmental Profit",
        "Total Revenue less Total Departmental Expenses",
        False,
    ),
    "S12.total_undistributed_expenses": (
        "Summary Total Undistributed Expenses",
        "A&G, IT, S&M, POM, and Utilities expenses combined",
        True,
    ),
    "S12.gop": (
        "Summary GOP",
        "Departmental Profit less Total Undistributed Expenses",
        False,
    ),
    "S12.income_after_management_fees": (
        "Summary Income After Management Fees",
        "GOP less Total Management Fees",
        False,
    ),
    "S12.total_non_operating_income_and_expenses": (
        "Summary Total Non-Operating Income and Expenses",
        "the detailed Non-Operating accounts combined",
        True,
    ),
    "S12.ebitda": (
        "Summary EBITDA",
        "Income After Management Fees less Non-Operating Income and Expenses",
        False,
    ),
}


class OutputTemplateError(RuntimeError):
    """The template and the COA have drifted apart."""


def mapped_output_name(source_name: str) -> str:
    """Return a short, consistent user-facing name for a mapped workbook."""
    suffix = " [MAPPED].xlsx"
    max_stem_chars = MAX_OUTPUT_FILENAME_CHARS - len(suffix)
    stem = Path(source_name).stem.strip()
    if len(stem) > max_stem_chars:
        shortened = stem[:max_stem_chars].rstrip(" ._-")
        last_space = shortened.rfind(" ")
        stem = shortened[:last_space] if last_space >= max_stem_chars // 2 else shortened
    return f"{stem or 'Mapped P&L'}{suffix}"


def _field(item, name, default=None):
    """Read a field whether the caller passed a model object or a plain dict."""
    if isinstance(item, dict):
        return item.get(name, default)
    return getattr(item, name, default)


def _canonical_coa_ids() -> list[str]:
    with COA_CSV.open(encoding="utf-8-sig", newline="") as handle:
        return [row["coa_id"] for row in csv.DictReader(handle)]


def _assert_template_matches(sheet, canonical: list[str]) -> int:
    """Refuse to write if the template's rows no longer match the COA.

    A mismatch is not cosmetic. The model tab addresses accounts by row number,
    so one inserted account shifts every figure below it by a line and the
    workbook still opens, still calculates, and is wrong throughout. Failing here
    turns that into an obvious build error instead of a plausible-looking model.
    """
    last_row = FIRST_ACCOUNT_ROW + len(canonical) - 1
    found = [
        str(sheet.cell(row=row, column=ID_COL).value or "")
        for row in range(FIRST_ACCOUNT_ROW, last_row + 1)
    ]
    if found != canonical:
        for index, (in_template, expected) in enumerate(zip(found, canonical)):
            if in_template != expected:
                raise OutputTemplateError(
                    f"Template row {FIRST_ACCOUNT_ROW + index} holds "
                    f"{in_template!r}, but coa_v2.csv has {expected!r}. Rebuild "
                    f"the template with scripts/build_output_template.py."
                )
        raise OutputTemplateError(
            f"Template holds {len(found)} accounts, coa_v2.csv has "
            f"{len(canonical)}. Rebuild with scripts/build_output_template.py."
        )
    if sheet.cell(row=last_row + 1, column=ID_COL).value not in (None, ""):
        raise OutputTemplateError(
            f"Template has an account at row {last_row + 1}, past the last "
            f"COA row. Rebuild with scripts/build_output_template.py."
        )
    return last_row


def _periods(result: NormalizationResult) -> list[tuple[str, str, dict]]:
    """Ordered period id, label, and mapped values."""
    if not result.period_values:
        return [("selected", result.period_label, result.values)]
    return [
        (
            period_id,
            result.period_labels.get(period_id, result.period_label),
            values,
        )
        for period_id, values in result.period_values.items()
    ]


def _mapped_label_period(
    periods: list[tuple[str, str, dict]],
) -> tuple[str, str, dict]:
    """Choose the single most useful period for source-label explanations."""

    def priority(period: tuple[str, str, dict]) -> tuple[int, int, int]:
        period_id, label, _ = period
        text = f"{period_id} {label}".lower()
        if re.fullmatch(r"20\d{2} (?:actual|forecast|budget)", label.lower()) or any(
            token in text for token in ("full year", "full_year", "fy", "total")
        ):
            range_priority = 5
        elif "ttm" in text or "trailing 12" in text:
            range_priority = 4
        elif "ytd" in text or "year to date" in text:
            range_priority = 3
        elif any(token in text for token in ("mtd", "ptd", "current period")):
            range_priority = 2
        elif re.search(
            r"\b(?:jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*\b",
            text,
        ):
            range_priority = 1
        else:
            range_priority = 0

        years = [int(year) for year in re.findall(r"\b(20\d{2})\b", text)]
        recency = max(years, default=0)
        if "current" in text:
            recency = max(recency, 9_999)
        elif "last year" in text or "prior" in text:
            recency = max(recency, 9_998)

        scenario = 3 if "actual" in text or "last year" in text else 2 if "forecast" in text else 1 if "budget" in text else 0
        return range_priority, recency, scenario

    return max(periods, key=priority)


def _describe_check(check) -> tuple[str, str, str]:
    """Split a `severity|rule|target|detail...` check into readable parts.

    Checks are pipe-delimited strings, not objects. Parsing them properly is what
    puts them in front of a reviewer at all -- the previous reader expected
    dictionaries, matched nothing, and so produced an empty column on every run
    while the interface still reported a flag count.
    """
    text = str(check)
    parts = [part for part in text.split("|") if part != ""]
    if not parts:
        return "", "", text
    severity = parts[0].strip().lower()
    if severity not in SEVERITY_ORDER:
        return "", "", text
    rule = parts[1] if len(parts) > 1 else ""
    target = parts[2] if len(parts) > 2 else ""
    details = {
        key.strip(): value.strip()
        for part in parts[3:]
        if "=" in part
        for key, value in [part.split("=", 1)]
    }
    if rule in {
        "hierarchy_complete",
        "source_detail_incomplete",
        "hierarchy_partial_with_residual",
    } and {
        "parent",
        "children",
    } <= details.keys():
        try:
            parent = float(details["parent"])
            children = float(details["children"])
            difference = children - parent
            direction = "greater than" if difference > 0 else "less than"
            rendered = (
                f"Rollup warning: Child accounts are {abs(difference):,.0f} "
                f"{direction} the parent account."
            )
        except ValueError:
            rendered = "Rollup warning: Child accounts do not reconcile to the parent account."
    elif rule == "source_detail_incomplete":
        rendered = "Rollup warning: Child accounts do not reconcile to the parent account."
    elif rule == "summary_department":
        try:
            variance = float(details["variance"])
            subject, schedule, plural = SUMMARY_DEPARTMENT_NAMES.get(
                target,
                (target.rsplit(".", 1)[-1].replace("_", " "), "department", False),
            )
            amount = f"{abs(variance):,.0f}"
            if variance > 0:
                comparison = "exceed" if plural else "exceeds"
            else:
                comparison = "are below" if plural else "is below"
            rendered = (
                f"Summary {subject} {comparison} the detailed {schedule} "
                f"schedule by {amount}."
            )
        except (KeyError, ValueError):
            rendered = "The Summary total does not match the detailed department schedule."
    elif rule == "source_layer_conflict":
        try:
            variance = float(details["variance"])
            rendered = (
                "Source layer conflict: The selected mapped layer differs from "
                f"an alternate reported subtotal by {abs(variance):,.0f}."
            )
        except (KeyError, ValueError):
            rendered = (
                "The selected mapped layer differs from an alternate reported "
                "source subtotal."
            )
    elif rule == "source_discrepancy":
        try:
            variance = float(details["variance"])
            subject, schedule, plural = SUMMARY_DEPARTMENT_NAMES.get(
                target,
                (target.rsplit(".", 1)[-1].replace("_", " "), "department", False),
            )
            comparison = (
                ("exceed" if plural else "exceeds")
                if variance > 0
                else ("are below" if plural else "is below")
            )
            rendered = (
                f"Source discrepancy: Summary {subject} {comparison} the "
                f"independently reported {schedule} schedule by "
                f"{abs(variance):,.0f}."
            )
        except (KeyError, ValueError):
            rendered = "The independently reported Summary and department values differ."
    elif rule == "source_presentation_exception":
        try:
            variance = float(details["variance"])
            subject, schedule, plural = SUMMARY_DEPARTMENT_NAMES.get(
                target,
                (target.rsplit(".", 1)[-1].replace("_", " "), "department", False),
            )
            comparison = (
                ("exceed" if plural else "exceeds")
                if variance > 0
                else ("are below" if plural else "is below")
            )
            rendered = (
                f"Combined source presentation: Summary {subject} {comparison} "
                f"the separately reported {schedule} schedule by "
                f"{abs(variance):,.0f}."
            )
        except (KeyError, ValueError):
            rendered = (
                "The operator Summary combines Other Operated Departments and "
                "Miscellaneous Income that detail reports separately."
            )
    elif rule == "summary_math":
        try:
            variance = float(details["variance"])
            subject, equation, plural = SUMMARY_MATH_NAMES.get(
                target,
                (
                    f"Summary {target.rsplit('.', 1)[-1].replace('_', ' ').title()}",
                    "its required equation",
                    False,
                ),
            )
            verb = "are" if plural else "is"
            direction = "greater than" if variance > 0 else "less than"
            rendered = (
                f"{subject} {verb} {abs(variance):,.0f} {direction} {equation}."
            )
        except (KeyError, ValueError):
            rendered = "The Summary total does not satisfy its required equation."
    elif rule == "unresolved_negative_residual":
        try:
            rendered = (
                "A negative residual of "
                f"{abs(float(details['plug'])):,.0f} was not forced into the account; "
                "review source offsets, signs, or overlapping children."
            )
        except (KeyError, ValueError):
            rendered = "A material negative residual remains unresolved."
    elif rule == "small_source_reconciliation_difference":
        try:
            variance = abs(float(details.get("variance", "")))
            rendered = (
                f"The reported total differs from the related accounts by "
                f"{_money(variance)}. The reported values were retained."
            )
        except ValueError:
            rendered = "The reported total differs slightly from the related accounts; the reported values were retained."
    elif rule == "coverage_unspecified":
        rendered = "The source does not clearly show how this parent divides among its child accounts."
    elif rule == "hierarchy_partial_with_residual":
        rendered = "Some child detail is missing even though an all-other account is available."
    elif rule == "large_residual_plug":
        try:
            plug = abs(float(details.get("plug", "")))
            ratio = float(details.get("ratio", ""))
            rendered = (
                f"The remaining {_money(plug)} difference "
                f"({ratio:.1%} of the parent) was assigned to the all-other account."
            )
        except ValueError:
            rendered = "A material remaining difference was assigned to the all-other account."
    elif rule == "coverage_inconsistent":
        rendered = "The selected child detail conflicts with the stated parent coverage."
    elif rule == "parent_no_value_with_children":
        rendered = "Child accounts were mapped, but the related parent account is blank."
    elif rule == "non_operating_sign":
        rendered = "The non-operating income sign appears inconsistent with the source presentation."
    elif rule in {"source_row_repeated", "source_row_double_count"}:
        rendered = "A source line may have been used more than once."
    elif rule == "source_row_included_and_excluded":
        rendered = "The same source line was both included and excluded."
    elif rule == "ood_misc_summary_mode_unknown":
        rendered = "The source does not clearly separate Other Operated Departments from Miscellaneous Income."
    else:
        readable_rule = rule.replace("_", " ").strip().capitalize()
        rendered = readable_rule or str(check)
        if not rendered.endswith("."):
            rendered += "."
    if rendered.startswith("Rollup warning:"):
        pass
    elif severity == "error":
        rendered = f"Needs review: {rendered}"
    elif severity == "warning":
        rendered = f"Note: {rendered}"
    return severity, target, rendered


def _clean_feedback_message(message: str, coa: dict[str, dict]) -> str:
    """Remove implementation details from model-written reviewer messages."""
    text = str(message).strip()
    for coa_id in sorted(coa, key=len, reverse=True):
        account_name = str(coa.get(coa_id, {}).get("account_name") or "").strip()
        replacement = account_name or coa_id.split(".", 1)[-1].replace("_", " ")
        text = text.replace(coa_id, replacement)
    text = re.sub(
        r"\s*\([^)]*(?:\brow\s+\d+\b|\b[\w&.-]+!\d+\b)[^)]*\)",
        "",
        text,
        flags=re.IGNORECASE,
    )
    text = re.sub(r"\b[\w&.-]+!\d+\b", "", text)
    text = re.sub(r"\brow\s+\d+\b", "", text, flags=re.IGNORECASE)
    text = text.replace("no_value", "left blank")
    text = re.sub(
        r"\b[a-z][a-z0-9]*(?:_[a-z0-9]+)+\b",
        lambda match: match.group(0).replace("_f_b", " F&B").replace("_", " "),
        text,
    )
    text = re.sub(r"\s+([,.;:])", r"\1", text)
    text = re.sub(r"\s{2,}", " ", text)
    return text.strip(" ,;")


def _round_feedback_numbers(message: str) -> str:
    """Round reviewer figures and omit currency symbols without changing values."""

    def replace(match: re.Match) -> str:
        number = float(match.group("number").replace(",", ""))
        return f"{match.group('sign')}{number:,.0f}{match.group('percent')}"

    rounded = re.sub(
        r"(?<![\w.])(?P<sign>-?)(?P<currency>\$?)"
        r"(?P<number>\d[\d,]*\.\d+)(?P<percent>%?)(?!\w)",
        replace,
        str(message),
    )
    return rounded.replace("$", "")


def _feedback(
    result: NormalizationResult,
    known_ids: set[str],
    periods: list[tuple[str, str, dict]],
) -> tuple[dict[str, list[tuple[int, str]]], list[str]]:
    """Per-account feedback, plus everything that belongs to no single account."""
    by_account: dict[str, list[tuple[int, str]]] = {}
    orphans: list[str] = []
    multi = len(periods) > 1
    labels = {period_id: label for period_id, label, _ in periods}

    grouped_checks: dict[tuple[str, str, str], list[str]] = {}
    checks_by_period = result.checks_by_period or {
        periods[0][0]: list(result.checks or [])
    }
    for period_id, period_checks in checks_by_period.items():
        for check in period_checks:
            severity, target, rendered = _describe_check(check)
            grouped_checks.setdefault((severity, target, rendered), []).append(
                labels[period_id]
            )
    for (severity, target, rendered), affected_labels in grouped_checks.items():
        prefix = ""
        if multi and len(affected_labels) < len(periods):
            prefix = f"{', '.join(affected_labels)} — "
        line = f"{prefix}{rendered}"
        if target in known_ids:
            by_account.setdefault(target, []).append(
                (SEVERITY_ORDER.get(severity, 2), line)
            )
        else:
            orphans.append(f"{line} [{target}]" if target else line)

    for item in result.review_items or []:
        message = _clean_feedback_message(
            str(_field(item, "message", "") or ""), result.coa
        )
        if not message:
            continue
        kind = str(_field(item, "kind", "") or "")
        prefix = "Ambiguity" if kind == "ambiguity" else "Unusual presentation"
        rendered = f"{prefix}: {message}" if kind else message
        targets = [
            coa_id
            for coa_id in (_field(item, "coa_ids", []) or [])
            if coa_id in known_ids
        ]
        if targets:
            # Keep every affected ID in the run log, but show the human one note
            # only once. Prefer the Summary account, where reviewers encounter
            # cross-department presentation decisions first.
            primary = next(
                (coa_id for coa_id in targets if coa_id.startswith("S12.")),
                targets[0],
            )
            by_account.setdefault(primary, []).append((2, rendered))
        else:
            orphans.append(rendered)

    for issue in result.execution_issues or []:
        orphans.append(str(issue))
    for period_id, issues in (result.execution_issues_by_period or {}).items():
        prefix = f"{labels.get(period_id, period_id)} — " if multi else ""
        orphans.extend(f"{prefix}{issue}" for issue in issues or [])

    return by_account, orphans


def _validation_findings(
    result: NormalizationResult,
    periods: list[tuple[str, str, dict]],
) -> list[tuple[str, str, str, list[str]]]:
    """Group the final validator findings consistently across selected periods."""
    labels = {period_id: label for period_id, label, _ in periods}
    grouped: dict[tuple[str, str, str], list[str]] = {}
    checks_by_period = result.checks_by_period or {
        periods[0][0]: list(result.checks or [])
    }
    for period_id, checks in checks_by_period.items():
        for check in checks:
            severity, target, rendered = _describe_check(check)
            if severity:
                grouped.setdefault((severity, target, rendered), []).append(
                    labels[period_id]
                )
    return [(*finding, period_labels) for finding, period_labels in grouped.items()]


def _plural(count: int, singular: str) -> str:
    if count == 0:
        suffix = "es" if singular.endswith(("s", "x", "ch", "sh")) else "s"
        return f"No {singular}{suffix}"
    if count == 1:
        return f"{count} {singular}"
    suffix = "es" if singular.endswith(("s", "x", "ch", "sh")) else "s"
    return f"{count} {singular}{suffix}"


def _money(value) -> str:
    if value is None:
        return "blank"
    number = float(value)
    rendered = f"${abs(number):,.0f}" if math.isclose(number, round(number), abs_tol=0.005) else f"${abs(number):,.2f}"
    return f"({rendered})" if number < 0 else rendered


def _source_value(value, coa_id: str) -> str:
    if value is None:
        return "blank"
    lowered_id = coa_id.lower()
    if "occupancy" in lowered_id:
        return f"{float(value):.1%}"
    if lowered_id.endswith(".adr"):
        return _money(value)
    number = float(value)
    rendered = f"{abs(number):,.0f}"
    return f"({rendered})" if number < 0 else rendered


def _row_amount(
    row: dict,
    period: tuple[str, str, dict],
    coa_id: str,
) -> str:
    selected = row.get("selected_values") or row.get("values") or {}
    period_id, _, _ = period
    value = selected.get(period_id)
    if value is None and not selected:
        value = row.get("selected_value", row.get("value"))
    return _source_value(value, coa_id)


def _mapped_from(
    decision,
    evidence_by_key: dict[str, dict],
    period: tuple[str, str, dict],
    residual_plug: float | None = None,
) -> str:
    """Render cited rows, amounts, and simple arithmetic in analyst language."""
    coa_id = str(_field(decision, "coa_id", "") or "")

    def describe(key, prefix: str = "") -> str:
        row_key = str(key)
        row = evidence_by_key.get(row_key, {})
        label = str(row.get("label") or row_key).replace("\n", " ").strip()
        sheet_name = row_key.rsplit("!", 1)[0] if "!" in row_key else "Source"
        amount = _row_amount(row, period, coa_id)
        return f"{prefix}{sheet_name} - {label}: {amount}"

    operation = str(_field(decision, "operation", "direct") or "direct")
    operation = operation.split(".")[-1]
    headings = {
        "sum": "Added together:",
        "adjusted_subtotal": "Calculated from the subtotal less exclusions:",
        "negate": "Sign reversed:",
        "ratio": "Calculated as the first line divided by the second:",
        "product": "Calculated by multiplying:",
        "scale": "Calculated from the cited line using the stated scale:",
        "coa_rollup": "Derived from mapped department accounts.",
    }
    source_rows = list(dict.fromkeys(_field(decision, "source_rows", []) or []))
    excluded_rows = list(dict.fromkeys(_field(decision, "excluded_rows", []) or []))
    if operation in {"direct", "sum", "adjusted_subtotal"}:
        lines = []
        if len(source_rows) == 1:
            lines.append(describe(source_rows[0]))
        elif source_rows:
            lines.append("Add:")
            lines.extend(describe(key) for key in source_rows)
        if excluded_rows:
            lines.append("Less:")
            lines.extend(describe(key) for key in excluded_rows)
    else:
        lines = [headings[operation]] if operation in headings else []
        lines.extend(describe(key) for key in source_rows)
        if excluded_rows:
            lines.append("Less:")
            lines.extend(describe(key) for key in excluded_rows)
    if residual_plug is not None:
        lines.append(f"Residual plug: {_source_value(residual_plug, coa_id)}")
    return "\n".join(dict.fromkeys(lines))[:32_000]


def _inferred_venue_name(decision, evidence_by_key: dict[str, dict]) -> str:
    explicit = str(_field(decision, "venue_name", "") or "").strip()
    if explicit:
        return explicit
    rationale = str(_field(decision, "rationale", "") or "")
    named = re.search(r"\bnamed\s+['\"]([^'\"]+)['\"]", rationale, re.IGNORECASE)
    if named:
        return named.group(1).strip()
    source_rows = _field(decision, "source_rows", []) or []
    if len(source_rows) == 1:
        label = str(
            evidence_by_key.get(str(source_rows[0]), {}).get("label") or ""
        ).strip()
        if label and not re.search(
            r"\b(total|food revenue|beverage revenue)\b", label, re.IGNORECASE
        ):
            return label
    return ""


def _venue_names(by_account: dict, evidence_by_key: dict[str, dict]) -> dict[str, str]:
    """Operator venue names, keyed to the food row of each generic slot.

    The model tab labels its venue rows from these cells, so a slot with no
    operator name still gets a placeholder -- a blank label there reads as a
    broken model rather than as an unused venue.
    """
    names: dict[str, str] = {}
    for number, slot in enumerate(GENERIC_VENUE_SLOTS, start=1):
        name = ""
        for coa_id in slot:
            candidate = _inferred_venue_name(
                by_account.get(coa_id), evidence_by_key
            )
            if candidate:
                name = candidate
                break
        names[slot[0]] = name or f"Venue {number}"
    return names


def _run_note_mismatch_counts(result, periods) -> dict[str, int]:
    """Count the three requested final mismatch classes above 10."""
    period_checks = list(
        (
            result.checks_by_period
            or {periods[0][0]: list(result.checks or [])}
        ).items()
    )
    values_by_period = {period_id: values for period_id, _, values in periods}
    children_by_parent: dict[str, list[str]] = {}
    for coa_id, metadata in result.coa.items():
        parent = metadata.get("parent_coa_id")
        if parent:
            children_by_parent.setdefault(str(parent), []).append(str(coa_id))
    counts = {
        "summary_math": set(),
        "summary_department": set(),
        "hierarchy": set(),
    }

    for period_id, checks in period_checks:
        for check in checks or []:
            parts = str(check).split("|")
            if len(parts) < 3:
                continue
            rule, target = parts[1], parts[2]
            details = {
                key.strip(): value.strip()
                for part in parts[3:]
                if "=" in part
                for key, value in [part.split("=", 1)]
            }
            if rule in {"summary_math", "summary_department"}:
                try:
                    variance = float(details["variance"])
                except (KeyError, ValueError):
                    continue
                if abs(variance) > 10:
                    counts[rule].add((period_id, target))
            elif rule in {
                "hierarchy_complete",
                "source_detail_incomplete",
                "hierarchy_partial_with_residual",
            }:
                try:
                    if {"parent", "children"} <= details.keys():
                        variance = float(details["parent"]) - float(details["children"])
                    else:
                        variance = float(details["variance"])
                except (KeyError, ValueError):
                    values = values_by_period.get(period_id, {})
                    child_ids = children_by_parent.get(target, [])
                    if target not in values or not child_ids:
                        continue
                    parent_value = float(values.get(target) or 0.0)
                    child_value = sum(
                        float(values.get(child) or 0.0)
                        for child in child_ids
                    )
                    variance = parent_value - child_value
                if abs(variance) > 10:
                    counts["hierarchy"].add((period_id, target))

    return {name: len(items) for name, items in counts.items()}


def _run_note_mismatches(result, periods) -> list[str]:
    counts = _run_note_mismatch_counts(result, periods)
    return [
        _plural(counts["summary_math"], "summary math error"),
        _plural(counts["summary_department"], "summary-to-department error"),
        _plural(counts["hierarchy"], "material rollup warning"),
    ]


def _write_run_notes(book, result, orphans, periods) -> None:
    """Write the compact run summary; detailed findings live on the COA tab."""
    if "Run Notes" not in book.sheetnames:
        raise OutputTemplateError("The output template is missing the Run Notes tab.")
    sheet = book["Run Notes"]
    for row in range(4, max(sheet.max_row, 20) + 1):
        sheet.cell(row=row, column=3).value = None
        if row > 8:
            sheet.cell(row=row, column=2).value = None

    findings = _validation_findings(result, periods)
    errors = [finding for finding in findings if finding[0] == "error"]
    warnings = [finding for finding in findings if finding[0] == "warning"]
    human_notes = len(result.review_items or [])
    mismatch_counts = _run_note_mismatch_counts(result, periods)
    status_by_outcome = {
        "clean": "Completed",
        "source_exception": "Completed — source exception",
        "coverage_gap": "Completed — coverage gap",
        "scope_exception": "Rejected — scope decision required",
        "rejected": "Rejected",
    }
    if result.stopped_reason:
        status = "Stopped"
    elif result.accepted and result.outcome == "rejected":
        # Compatibility for callers that predate the explicit outcome field.
        status = None
    else:
        status = status_by_outcome.get(result.outcome)
    if status is None:
        if errors:
            status = "Completed with errors"
        elif warnings or human_notes or any(mismatch_counts.values()):
            status = "Completed with warnings"
        else:
            status = "Completed"

    details = {
        4: result.source_name,
        5: ", ".join(label for _, label, _ in periods),
        6: status,
        7: int(result.mapped_account_count),
        8: (
            "Every value is calculated by code from cited source rows; "
            "model-generated numbers are not used."
        ),
    }
    labels = {
        4: "Source file",
        5: "Periods",
        6: "Status",
        7: "Accounts mapped",
        8: "Calculation policy",
    }
    for row_number, detail in details.items():
        label_cell = sheet.cell(row=row_number, column=2, value=labels[row_number])
        label_cell.alignment = Alignment(horizontal="left", vertical="bottom")
        detail_cell = sheet.cell(row=row_number, column=3, value=detail)
        detail_cell.alignment = copy(detail_cell.alignment)
        detail_cell.alignment = Alignment(
            horizontal="left",
            vertical="bottom",
            wrap_text=row_number == 8,
        )
        if row_number == 7:
            detail_cell.number_format = "0"
        if row_number == 8:
            sheet.row_dimensions[row_number].height = 14.5

    note_lines = _run_note_mismatches(result, periods)
    label_cell = sheet.cell(row=9, column=2, value="Notes")
    label_cell._style = copy(sheet.cell(row=8, column=2)._style)
    label_cell.alignment = Alignment(horizontal="left", vertical="bottom")
    note_cell = sheet.cell(row=9, column=3, value="\n".join(note_lines) or None)
    note_cell._style = copy(sheet.cell(row=8, column=3)._style)
    note_font = copy(note_cell.font)
    note_font.sz = 11
    note_cell.font = note_font
    note_cell.alignment = Alignment(
        horizontal="left", vertical="bottom", wrap_text=True
    )
    estimated_lines = sum(max(1, math.ceil(len(line) / 115)) for line in note_lines)
    sheet.row_dimensions[9].height = min(409, max(14.5, 15 * estimated_lines))


def _fill_model_periods(sheet, count: int) -> None:
    """Translate the authored period column rightward, once per period.

    Exact rather than approximate: translating the authored column C onto D
    reproduces the workbook's own authored column D for all 139 of its formulas,
    which is what makes doing the same for columns E onward safe.
    """
    origin_letter = get_column_letter(FIRST_PERIOD_COL)
    for offset in range(1, count):
        target_col = FIRST_PERIOD_COL + offset
        target_letter = get_column_letter(target_col)
        for row in range(1, sheet.max_row + 1):
            source = sheet.cell(row=row, column=FIRST_PERIOD_COL)
            target = sheet.cell(row=row, column=target_col)
            target._style = copy(source._style)
            value = source.value
            target.value = (
                Translator(value, origin=f"{origin_letter}{row}").translate_formula(
                    f"{target_letter}{row}"
                )
                if isinstance(value, str) and value.startswith("=")
                else value
            )
    for offset in range(MAX_PERIODS):
        letter = get_column_letter(FIRST_PERIOD_COL + offset)
        sheet.column_dimensions[letter].hidden = offset >= count


def _unhide_existing_columns_after(sheet, column: int) -> None:
    """Unhide authored columns outside the reserved output area without bloating it."""
    for dimension in sheet.column_dimensions.values():
        if (dimension.min or 0) > column:
            dimension.hidden = False


def _autofit_coa_rows(sheet, last_row: int) -> None:
    """Estimate Excel row heights for the two variable wrapped-text columns."""
    for row in range(FIRST_ACCOUNT_ROW, last_row + 1):
        line_counts = [1]
        for column, width in ((LABELS_COL, 75), (FEEDBACK_COL, 55)):
            cell = sheet.cell(row=row, column=column)
            cell.alignment = copy(cell.alignment)
            cell.alignment = Alignment(
                horizontal=cell.alignment.horizontal,
                vertical="bottom",
                wrap_text=True,
                indent=cell.alignment.indent,
            )
            text = str(cell.value or "")
            if text:
                line_counts.append(
                    sum(
                        max(1, math.ceil(len(part) / width))
                        for part in text.splitlines()
                    )
                )
        sheet.row_dimensions[row].height = min(409, max(18, 15 * max(line_counts)))


def write_normalized_workbook(result: NormalizationResult, path: Path) -> Path:
    """Write the standardized output and return the path written."""
    if not TEMPLATE.is_file():
        raise OutputTemplateError(
            f"Missing {TEMPLATE}. Build it with scripts/build_output_template.py."
        )

    canonical = _canonical_coa_ids()
    book = openpyxl.load_workbook(TEMPLATE)
    sheet = book["COA"]
    last_row = _assert_template_matches(sheet, canonical)

    periods = _periods(result)
    if len(periods) > MAX_PERIODS:
        raise OutputTemplateError(
            f"{len(periods)} periods selected, but the template reserves "
            f"{MAX_PERIODS} columns."
        )

    by_account = {str(_field(d, "coa_id", "")): d for d in result.decisions}
    evidence_by_key = {
        str(row.get("row_key")): row for row in (result.evidence or [])
    }
    known_ids = set(canonical)
    feedback, orphans = _feedback(result, known_ids, periods)
    venues = _venue_names(by_account, evidence_by_key)
    mapped_label_period = _mapped_label_period(periods)
    sheet.cell(
        row=HEADER_ROW,
        column=LABELS_COL,
        value=f"Mapped Labels - {mapped_label_period[1]}",
    )

    for offset, (_, label, values) in enumerate(periods):
        column = FIRST_PERIOD_COL + offset
        sheet.cell(row=HEADER_ROW, column=column, value=label)
        for index, coa_id in enumerate(canonical):
            value = values.get(coa_id)
            # Blank, not zero, when a period has no figure: a written zero is a
            # claim that the account was mapped and came to nothing.
            sheet.cell(
                row=FIRST_ACCOUNT_ROW + index,
                column=column,
                value=None if value is None else round(float(value), 2),
            )
    for offset in range(MAX_PERIODS):
        letter = get_column_letter(FIRST_PERIOD_COL + offset)
        sheet.column_dimensions[letter].hidden = offset >= len(periods)
    sheet.column_dimensions[get_column_letter(ID_COL)].hidden = False
    for column in (LABELS_COL, FEEDBACK_COL, VENUE_COL):
        sheet.column_dimensions[get_column_letter(column)].hidden = False
    _unhide_existing_columns_after(sheet, VENUE_COL)

    for index, coa_id in enumerate(canonical):
        row = FIRST_ACCOUNT_ROW + index
        decision = by_account.get(coa_id)
        residual_plug = (result.residual_plugs_by_period or {}).get(
            mapped_label_period[0], {}
        ).get(coa_id)
        sheet.cell(
            row=row,
            column=LABELS_COL,
            value=(
                _mapped_from(
                    decision,
                    evidence_by_key,
                    mapped_label_period,
                    residual_plug=residual_plug,
                )
                if decision
                else None
            ),
        )

        notes = list(feedback.get(coa_id, []))
        if decision is None:
            notes.append((3, "No source found — nothing mapped to this account."))
        if notes:
            ordered = [
                _round_feedback_numbers(text)
                for _, text in sorted(notes, key=lambda item: item[0])
            ]
            sheet.cell(
                row=row,
                column=FEEDBACK_COL,
                value="\n".join(dict.fromkeys(ordered))[:2_000],
            )

        if coa_id in venues:
            sheet.cell(row=row, column=VENUE_COL, value=venues[coa_id])

    _autofit_coa_rows(sheet, last_row)
    _write_run_notes(book, result, orphans, periods)
    model_sheet = book["KHP Model Accounts"]
    _fill_model_periods(model_sheet, len(periods))
    _unhide_existing_columns_after(model_sheet, FIRST_PERIOD_COL + MAX_PERIODS - 1)
    _reset_sheet_views(book)

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    book.save(path)
    _restore_template_textbox(path)
    return path


def _reset_sheet_views(book: openpyxl.Workbook) -> None:
    """Open every worksheet at A1 without disturbing its frozen rows/columns."""
    book.active = 0
    for sheet in book.worksheets:
        view = sheet.sheet_view
        view.topLeftCell = "A1"

        pane = view.pane
        selection_pane = None
        if pane is not None and pane.state in {"frozen", "frozenSplit"}:
            frozen_columns = int(float(pane.xSplit or 0))
            frozen_rows = int(float(pane.ySplit or 0))
            if frozen_columns or frozen_rows:
                pane.topLeftCell = (
                    f"{get_column_letter(frozen_columns + 1)}{frozen_rows + 1}"
                )
                pane.activePane = "topLeft"
                selection_pane = "topLeft"

        view.selection = [
            Selection(pane=selection_pane, activeCell="A1", sqref="A1")
        ]


def _restore_template_textbox(path: Path) -> None:
    """Restore the authored KHP textbox that openpyxl otherwise drops on save."""
    with zipfile.ZipFile(TEMPLATE) as source:
        required = {MODEL_DRAWING_PART, MODEL_SHEET_RELS}
        if not required <= set(source.namelist()):
            return
        drawing_info = source.getinfo(MODEL_DRAWING_PART)
        drawing = source.read(MODEL_DRAWING_PART)
        rels_info = source.getinfo(MODEL_SHEET_RELS)
        rels = source.read(MODEL_SHEET_RELS)

    relationship_ids = re.findall(
        rb'<Relationship\b(?=[^>]*\bType="[^"]*/drawing")'
        rb'(?=[^>]*\bTarget="[^"]*drawings/drawing1\.xml")'
        rb'[^>]*\bId="([^"]+)"[^>]*/?>',
        rels,
    )
    if len(relationship_ids) != 1:
        raise OutputTemplateError(
            "The output template must define exactly one relationship for "
            f"{MODEL_DRAWING_PART}; found {len(relationship_ids)}."
        )
    drawing_id = relationship_ids[0]
    drawing_tag = (
        b'<drawing xmlns:r="http://schemas.openxmlformats.org/'
        b'officeDocument/2006/relationships" r:id="'
        + drawing_id
        + b'"/>'
    )
    drawing_override = (
        b'<Override PartName="/xl/drawings/drawing1.xml" '
        b'ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>'
    )
    with zipfile.ZipFile(path) as original:
        archive = {
            item.filename: (item, original.read(item.filename))
            for item in original.infolist()
        }
    archive.pop(MODEL_SHEET_RELS, None)
    archive.pop(MODEL_DRAWING_PART, None)
    sheet_info, sheet_xml = archive[MODEL_SHEET_PART]
    drawing_reference = re.compile(
        rb"<(?:[A-Za-z_][\w.-]*:)?drawing\b[^>]*/>"
    )
    if drawing_reference.search(sheet_xml):
        sheet_xml = drawing_reference.sub(drawing_tag, sheet_xml, count=1)
    else:
        sheet_xml = sheet_xml.replace(
            b"</worksheet>", drawing_tag + b"</worksheet>"
        )
    archive[MODEL_SHEET_PART] = (sheet_info, sheet_xml)
    content_info, content_xml = archive["[Content_Types].xml"]
    if b"/xl/drawings/drawing1.xml" not in content_xml:
        content_xml = content_xml.replace(
            b"</Types>", drawing_override + b"</Types>"
        )
        archive["[Content_Types].xml"] = (content_info, content_xml)

    handle = tempfile.NamedTemporaryFile(
        dir=path.parent, suffix=".xlsx", delete=False
    )
    temporary = Path(handle.name)
    handle.close()
    try:
        with zipfile.ZipFile(temporary, "w") as output:
            for item, data in archive.values():
                output.writestr(item, data)
            output.writestr(rels_info, rels)
            output.writestr(drawing_info, drawing)
        replace_atomically(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)
