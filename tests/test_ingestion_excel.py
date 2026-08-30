from __future__ import annotations

import shutil
import sys
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path
from types import SimpleNamespace

import openpyxl
import pytest

from hotel_pl_normalizer.models.workbook import FileType
from hotel_pl_normalizer.structure.ingestion import read_excel_workbook
from hotel_pl_normalizer.structure.ingestion.openpyxl_compat import (
    load_openpyxl_workbook,
)


def _add_defined_names(path: Path, names: list[tuple[str, str]]) -> None:
    staged = path.with_suffix(".staged.xlsx")
    namespace = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(staged, "w") as target:
        for entry in source.infolist():
            payload = source.read(entry.filename)
            if entry.filename == "xl/workbook.xml":
                root = ET.fromstring(payload)
                container = root.find(f"{{{namespace}}}definedNames")
                if container is None:
                    container = ET.SubElement(root, f"{{{namespace}}}definedNames")
                for name, value in names:
                    item = ET.SubElement(container, f"{{{namespace}}}definedName")
                    item.set("name", name)
                    if name in {"_xlnm.Print_Titles", "_xlnm.Print_Area"}:
                        item.set("localSheetId", "0")
                    item.text = value
                payload = ET.tostring(root, encoding="utf-8", xml_declaration=True)
            target.writestr(entry, payload)
    staged.replace(path)


def test_read_xlsx_workbook_returns_used_range_workbook_record(tmp_path: Path):
    path = tmp_path / "sample.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "P&L"
    ws["A1"] = "Hotel P&L"
    ws["A1"].font = openpyxl.styles.Font(bold=True)
    ws.merge_cells("A1:C1")
    ws["A3"] = "Rooms"
    ws["C3"] = 100
    wb.save(path)

    record = read_excel_workbook(path, source_id="upload_1")

    assert record.source.source_id == "upload_1"
    assert record.source.original_filename == "sample.xlsx"
    assert record.source.local_path == str(path)
    assert record.source.file_hash
    assert record.workbook_metadata.sheet_count == 1

    sheet = record.sheets[0]
    assert sheet.sheet_name == "P&L"
    assert sheet.max_row == 3
    assert sheet.max_column == 3
    assert len(sheet.rows) == 3
    assert sheet.rows[1].cells[0].raw_value is None
    assert sheet.rows[0].cells[0].style.bold is True
    # Merge references are read directly from worksheet XML, so period headers
    # remain available without loading the workbook's cell objects into memory.
    assert len(sheet.merged_ranges) == 1
    assert sheet.merged_ranges[0].range == "A1:C1"
    assert sheet.merged_ranges[0].value == "Hotel P&L"
    assert sheet.rows[0].cells[1].is_merged is True
    assert sheet.rows[0].cells[1].merged_parent == "A1"
    # The merged label still reads from its top-left cell, which is where the
    # value actually lives -- that has not changed.
    assert sheet.rows[0].cells[0].raw_value == "Hotel P&L"
    assert sheet.rows[2].cells[2].raw_value == 100


def test_read_excel_workbook_supports_xlsx_directly(tmp_path: Path):
    path = tmp_path / "sample.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Direct"
    ws["A1"] = "Revenue"
    wb.save(path)

    record = read_excel_workbook(path, source_id="upload_xlsx")

    assert record.source.source_id == "upload_xlsx"
    assert record.source.file_type == FileType.XLSX
    assert record.source.original_filename == "sample.xlsx"
    assert record.sheets[0].rows[0].cells[0].raw_value == "Revenue"


def test_invalid_print_metadata_is_removed_from_temporary_copy(tmp_path: Path):
    path = tmp_path / "invalid_print_titles.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active["A1"] = "Revenue"
    workbook.save(path)
    _add_defined_names(
        path,
        [
            ("_xlnm.Print_Titles", "#N/A"),
            ("_xlnm.Print_Area", "#REF!"),
            ("financial_model_name", "#REF!"),
        ],
    )

    original_bytes = path.read_bytes()
    record = read_excel_workbook(path)

    assert record.sheets[0].rows[0].cells[0].raw_value == "Revenue"
    assert path.read_bytes() == original_bytes
    assert len(record.ingestion_warnings) == 1
    warning = record.ingestion_warnings[0].message
    assert "_xlnm.Print_Titles=#N/A" in warning
    assert "_xlnm.Print_Area=#REF!" in warning
    assert "financial_model_name" not in warning
    assert "source file was not changed" in warning


def test_valid_workbook_uses_original_path_without_repairs(tmp_path: Path):
    path = tmp_path / "valid.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active["A1"] = "Revenue"
    workbook.save(path)

    loaded = load_openpyxl_workbook(path, read_only=True, data_only=True)
    try:
        assert loaded.read_path == path
        assert loaded.repairs == ()
    finally:
        loaded.close()


def test_non_print_corruption_is_not_repaired(tmp_path: Path):
    path = tmp_path / "corrupt.xlsx"
    workbook = openpyxl.Workbook()
    workbook.save(path)
    path.write_bytes(b"not an OOXML package")

    with pytest.raises(zipfile.BadZipFile):
        load_openpyxl_workbook(path, read_only=True, data_only=True)


def test_workbook_id_is_derived_from_content_not_random(tmp_path: Path):
    """The workbook id must be reproducible for identical content.

    It is embedded in every packet_id and therefore in every prompt, so a random
    id makes each run's prompts unique and silently defeats the response cache --
    which both inflates cost and makes any two runs incomparable.
    """

    def build(path: Path, label: str) -> Path:
        wb = openpyxl.Workbook()
        wb.active["A1"] = label
        wb.save(path)
        return path

    first = build(tmp_path / "a.xlsx", "Revenue")
    different = build(tmp_path / "c.xlsx", "Expenses")
    # A byte-for-byte copy under a different filename must yield the same id.
    # Copied rather than re-saved because openpyxl stamps archive entry times.
    same_content = tmp_path / "b.xlsx"
    shutil.copyfile(first, same_content)

    id_first = read_excel_workbook(first).workbook_id
    assert id_first == read_excel_workbook(first).workbook_id, "unstable across reads"
    assert id_first == read_excel_workbook(same_content).workbook_id
    assert id_first != read_excel_workbook(different).workbook_id


def test_workbook_id_ignores_excel_archive_metadata(tmp_path: Path):
    """Re-saving unchanged parsed content must not invalidate every prompt."""
    first = tmp_path / "first.xlsx"
    second = tmp_path / "second.xlsx"
    wb = openpyxl.Workbook()
    wb.active["A1"] = "Revenue"
    wb.active["B1"] = 100
    wb.save(first)

    reloaded = openpyxl.load_workbook(first)
    reloaded.save(second)

    first_record = read_excel_workbook(first)
    second_record = read_excel_workbook(second)
    assert first_record.workbook_id == second_record.workbook_id

    changed = openpyxl.load_workbook(second)
    changed.active["B1"] = 101
    changed.save(second)
    assert first_record.workbook_id != read_excel_workbook(second).workbook_id


def test_workbook_id_changes_when_interpretive_formatting_changes(tmp_path: Path):
    plain = tmp_path / "plain.xlsx"
    bold = tmp_path / "bold.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active["A1"] = "Rooms"
    workbook.save(plain)
    workbook.active["A1"].font = openpyxl.styles.Font(bold=True)
    workbook.save(bold)

    assert read_excel_workbook(plain).workbook_id != read_excel_workbook(bold).workbook_id


def test_read_excel_workbook_supports_xlsm_directly(tmp_path: Path):
    path = tmp_path / "macro_enabled.xlsm"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "P&L"
    ws["A1"] = "Rooms Revenue"
    wb.save(path)

    record = read_excel_workbook(path, source_id="upload_xlsm")

    assert record.source.source_id == "upload_xlsm"
    assert record.source.file_type == FileType.XLSM
    assert record.source.original_filename == "macro_enabled.xlsm"
    assert record.sheets[0].rows[0].cells[0].raw_value == "Rooms Revenue"


def test_read_excel_workbook_supports_xls_natively_with_xlrd(tmp_path: Path, monkeypatch):
    source_path = tmp_path / "legacy.xls"
    source_path.write_bytes(b"legacy workbook bytes")

    class FakeCell:
        def __init__(self, ctype: int, value):
            self.ctype = ctype
            self.value = value
            self.xf_index = 0

    class FakeSheet:
        name = "Legacy"
        nrows = 1
        ncols = 2
        merged_cells = []
        rowinfo_map = {}

        def cell(self, rowx: int, colx: int):
            values = {
                (0, 0): FakeCell(1, "Rooms"),
                (0, 1): FakeCell(2, 123.0),
            }
            return values.get((rowx, colx), FakeCell(0, ""))

    class FakeBook:
        nsheets = 1
        datemode = 0
        sheet_visibility = [0]
        xf_list = []
        format_map = {}
        font_list = []

        def sheet_by_index(self, idx: int):
            assert idx == 0
            return FakeSheet()

    fake_xlrd = SimpleNamespace(
        XL_CELL_EMPTY=0,
        XL_CELL_TEXT=1,
        XL_CELL_NUMBER=2,
        XL_CELL_DATE=3,
        XL_CELL_BOOLEAN=4,
        XL_CELL_ERROR=5,
        XL_CELL_BLANK=6,
        error_text_from_code={},
        open_workbook=lambda path, formatting_info: FakeBook(),
        xldate_as_datetime=lambda value, datemode: value,
    )
    monkeypatch.setitem(sys.modules, "xlrd", fake_xlrd)

    record = read_excel_workbook(source_path, source_id="upload_xls")

    assert record.source.source_id == "upload_xls"
    assert record.source.file_type == FileType.XLS
    assert record.source.original_filename == "legacy.xls"
    assert record.source.local_path == str(source_path)
    assert record.source.file_hash
    assert record.sheets[0].sheet_name == "Legacy"
    assert record.sheets[0].rows[0].cells[0].raw_value == "Rooms"
    assert record.sheets[0].rows[0].cells[1].raw_value == 123
    assert record.ingestion_warnings == []


def test_cells_sharing_a_style_share_one_style_object(tmp_path: Path):
    """Styles are resolved once per distinct style, not once per cell.

    `cell.font` and friends resolve against the workbook's style tables on every
    access, which is six lookups per cell across hundreds of thousands of cells.
    Reusing the instance is what makes that cheap, so the sharing is asserted
    directly rather than inferred from a timing.
    """
    path = tmp_path / "styles.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    for address in ("A1", "A2"):
        ws[address] = "Rooms"
        ws[address].font = openpyxl.styles.Font(bold=True)
    ws["B1"] = 1
    ws["B2"] = 2
    wb.save(path)

    record = read_excel_workbook(path, source_id="styles")
    rows = record.sheets[0].rows
    first_bold, second_bold = rows[0].cells[0], rows[1].cells[0]
    plain = rows[0].cells[1]

    assert first_bold.style.bold is True
    assert second_bold.style.bold is True
    assert plain.style.bold is False
    assert first_bold.style is second_bold.style
    assert first_bold.style is not plain.style


def test_cell_style_cannot_be_mutated(tmp_path: Path):
    """One style object backs many cells, so mutating it would corrupt all of them."""
    import dataclasses

    path = tmp_path / "frozen.xlsx"
    wb = openpyxl.Workbook()
    wb.active["A1"] = "Rooms"
    wb.save(path)

    record = read_excel_workbook(path, source_id="frozen")
    style = record.sheets[0].rows[0].cells[0].style

    with pytest.raises(dataclasses.FrozenInstanceError):
        style.bold = True
