"""The output workbook is what a client and an analyst both judge us by.

The tests that matter most here are the structural ones: the COA tab must stay
row-for-row aligned with coa_v2.csv, because every formula on the model tab is a
hard row reference and a one-row shift produces a workbook that opens,
calculates, and is wrong throughout.
"""

import csv
import re
import zipfile
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from hotel_pl_normalizer.output import (
    COA_CSV,
    FEEDBACK_COL,
    FIRST_ACCOUNT_ROW,
    FIRST_PERIOD_COL,
    HEADER_ROW,
    ID_COL,
    LABELS_COL,
    MAX_PERIODS,
    TEMPLATE,
    VENUE_COL,
    OutputTemplateError,
    _mapped_from,
    mapped_output_name,
    write_normalized_workbook,
)
from hotel_pl_normalizer.pipeline import NormalizationResult


def canonical_ids():
    with COA_CSV.open(encoding="utf-8-sig", newline="") as handle:
        return [row["coa_id"] for row in csv.DictReader(handle)]


def build_result(**overrides):
    defaults = {
        "workbook_id": "wb_test",
        "source_name": "Hotel P&L.xlsx",
        "period_label": "YTD Actual",
        "values": {},
        "coa": {},
        "accepted": True,
    }
    defaults.update(overrides)
    return NormalizationResult(**defaults)


def test_derived_summary_mapped_label_explains_coa_rollup():
    decision = SimpleNamespace(
        coa_id="S12.total_revenue",
        operation="coa_rollup",
        source_rows=[],
        excluded_rows=[],
    )

    assert _mapped_from(
        decision,
        {},
        ("selected", "FY 2025", {"S12.total_revenue": 100.0}),
    ) == "Derived from mapped department accounts."


def test_deterministic_ffe_and_noi_labels_need_no_model_decisions(tmp_path):
    ids = canonical_ids()
    result = build_result(
        coa={i: {} for i in ids},
        values={"S12.ffe_reserve": 40.0, "S12.noi": 110.0},
    )

    sheet = load_workbook(
        write_normalized_workbook(result, tmp_path / "o.xlsx")
    )["COA"]

    assert sheet.cell(
        row=FIRST_ACCOUNT_ROW + ids.index("S12.ffe_reserve"),
        column=LABELS_COL,
    ).value == "Calculated: 4% of Total Revenue."
    assert sheet.cell(
        row=FIRST_ACCOUNT_ROW + ids.index("S12.noi"),
        column=LABELS_COL,
    ).value == "Calculated: EBITDA less FF&E Reserve."


def test_coa_tab_carries_every_account_in_canonical_order(tmp_path):
    """The guarantee the whole model tab rests on."""
    ids = canonical_ids()
    result = build_result(values={ids[0]: 100.0}, coa={i: {} for i in ids})

    sheet = load_workbook(write_normalized_workbook(result, tmp_path / "o.xlsx"))["COA"]

    written = [
        sheet.cell(row=FIRST_ACCOUNT_ROW + n, column=ID_COL).value
        for n in range(len(ids))
    ]
    assert written == ids
    # Nothing may sit immediately after the block; the model reads to that row.
    assert sheet.cell(row=FIRST_ACCOUNT_ROW + len(ids), column=ID_COL).value is None


def test_unmapped_accounts_are_still_written_as_rows(tmp_path):
    """A skipped row would shift every account below it on the model tab."""
    ids = canonical_ids()
    result = build_result(values={ids[5]: 42.0}, coa={i: {} for i in ids})

    sheet = load_workbook(write_normalized_workbook(result, tmp_path / "o.xlsx"))["COA"]

    assert sheet.cell(row=FIRST_ACCOUNT_ROW, column=ID_COL).value == ids[0]
    assert sheet.cell(row=FIRST_ACCOUNT_ROW + 5, column=FIRST_PERIOD_COL).value == 42


def test_one_column_per_period_and_the_rest_stay_hidden(tmp_path):
    ids = canonical_ids()
    result = build_result(
        coa={i: {} for i in ids},
        period_labels={"current": "2025 YTD", "prior": "2024 YTD"},
        period_values={
            "current": {ids[0]: 10.0, ids[1]: None},
            "prior": {ids[0]: 20.0, ids[1]: 30.0},
        },
    )

    book = load_workbook(write_normalized_workbook(result, tmp_path / "o.xlsx"))
    sheet = book["COA"]

    assert sheet.cell(row=HEADER_ROW, column=FIRST_PERIOD_COL).value == "2025 YTD"
    assert sheet.cell(row=HEADER_ROW, column=FIRST_PERIOD_COL + 1).value == "2024 YTD"
    assert sheet.cell(row=FIRST_ACCOUNT_ROW, column=FIRST_PERIOD_COL).value == 10
    assert sheet.cell(row=FIRST_ACCOUNT_ROW, column=FIRST_PERIOD_COL + 1).value == 20
    # A period with no figure stays blank rather than claiming a mapped zero.
    assert sheet.cell(row=FIRST_ACCOUNT_ROW + 1, column=FIRST_PERIOD_COL).value is None
    assert sheet.column_dimensions["C"].hidden is False
    assert sheet.column_dimensions["D"].hidden is False
    assert sheet.column_dimensions["E"].hidden is True
    model = book["KHP Model Accounts"]
    assert model.column_dimensions["C"].hidden is False
    assert model.column_dimensions["D"].hidden is False
    assert model.column_dimensions["E"].hidden is True
    assert model.column_dimensions["V"].hidden is True
    assert model.column_dimensions["W"].hidden is False
    assert sheet.column_dimensions["Z"].hidden is False


def test_output_preserves_the_period_labels_selected_by_the_user(tmp_path):
    ids = canonical_ids()
    result = build_result(
        coa={i: {} for i in ids},
        period_labels={"current": "YTD Current", "prior": "YTD Last Year"},
        period_values={"current": {}, "prior": {}},
    )

    sheet = load_workbook(write_normalized_workbook(result, tmp_path / "o.xlsx"))["COA"]

    assert sheet.cell(row=HEADER_ROW, column=FIRST_PERIOD_COL).value == "YTD Current"
    assert sheet.cell(row=HEADER_ROW, column=FIRST_PERIOD_COL + 1).value == "YTD Last Year"
    assert sheet.cell(row=HEADER_ROW, column=LABELS_COL).value == (
        "Mapped Labels - YTD Current"
    )


def test_metadata_columns_do_not_move_when_periods_grow(tmp_path):
    """The model tab reads venue names by fixed cell; W/X/Y must never shift."""
    ids = canonical_ids()
    many = {f"p{n}": {ids[0]: float(n)} for n in range(8)}
    result = build_result(
        coa={i: {} for i in ids},
        period_labels={f"p{n}": f"Period {n}" for n in range(8)},
        period_values=many,
    )

    sheet = load_workbook(write_normalized_workbook(result, tmp_path / "o.xlsx"))["COA"]

    assert sheet.cell(row=HEADER_ROW, column=LABELS_COL).value == (
        "Mapped Labels - Period 0"
    )
    assert sheet.cell(row=HEADER_ROW, column=FEEDBACK_COL).value == "MODEL FEEDBACK"
    assert sheet.cell(row=HEADER_ROW, column=VENUE_COL).value == "VENUE NAME"


def test_more_periods_than_reserved_slots_is_refused(tmp_path):
    ids = canonical_ids()
    too_many = MAX_PERIODS + 1
    result = build_result(
        coa={i: {} for i in ids},
        period_labels={f"p{n}": f"Period {n}" for n in range(too_many)},
        period_values={f"p{n}": {} for n in range(too_many)},
    )

    with pytest.raises(OutputTemplateError, match="reserves"):
        write_normalized_workbook(result, tmp_path / "o.xlsx")


def test_model_tab_formulas_are_translated_per_period(tmp_path):
    ids = canonical_ids()
    result = build_result(
        coa={i: {} for i in ids},
        period_labels={"a": "A", "b": "B", "c": "C"},
        period_values={"a": {}, "b": {}, "c": {}},
    )

    book = load_workbook(write_normalized_workbook(result, tmp_path / "o.xlsx"))
    model = book["KHP Model Accounts"]

    # EBITDA, authored at C238 as =C225-C236 in the expanded model.
    assert model["C238"].value == "=C225-C236"
    assert model["D238"].value == "=D225-D236"
    assert model["E238"].value == "=E225-E236"
    assert model["C20"].value == "=COA!C3"
    assert model["D20"].value == "=COA!D3"
    assert model["C233"].value == "=COA!C30-'KHP Model Accounts'!C234"
    assert model["D233"].value == "=COA!D30-'KHP Model Accounts'!D234"
    assert model["C234"].value == "=COA!C270"
    assert model["D234"].value == "=COA!D270"
    # The header pull-through follows too, so each column names its own period.
    assert model["E17"].value == "=COA!E2"
    assert model.column_dimensions["F"].hidden is True


def test_model_tab_venue_labels_point_at_the_venue_column(tmp_path):
    """Venue labels are read by cell, so they must survive the column move."""
    ids = canonical_ids()
    result = build_result(coa={i: {} for i in ids})

    book = load_workbook(write_normalized_workbook(result, tmp_path / "o.xlsx"))
    assert book["KHP Model Accounts"]["B40"].value == "=COA!Y86"


def test_model_tab_preserves_authored_label_indentation(tmp_path):
    ids = canonical_ids()
    result = build_result(coa={i: {} for i in ids})

    model = load_workbook(
        write_normalized_workbook(result, tmp_path / "o.xlsx")
    )["KHP Model Accounts"]

    assert model["B36"].alignment.indent == 2
    assert model["B38"].alignment.indent == 5
    assert model["B66"].alignment.indent == 8
    assert sum(
        bool(model.cell(row=row, column=2).alignment.indent)
        for row in range(2, model.max_row + 1)
    ) == 172


def test_mapped_labels_show_sources_and_exclusions(tmp_path):
    ids = canonical_ids()
    target = "S2.venue1_food_revenue"
    result = build_result(
        coa={i: {} for i in ids},
        values={target: 80.0},
        decisions=[
            SimpleNamespace(
                coa_id=target,
                operation="adjusted_subtotal",
                source_rows=["Rooms!14"],
                excluded_rows=["Rooms!12"],
                venue_name="La Loba",
            )
        ],
        evidence=[
            {
                "row_key": "Rooms!12",
                "label": "Contract Labor",
                "selected_value": 20,
            },
            {
                "row_key": "Rooms!14",
                "label": "Total Salaries",
                "selected_value": 100,
            },
        ],
    )

    sheet = load_workbook(write_normalized_workbook(result, tmp_path / "o.xlsx"))["COA"]
    row = FIRST_ACCOUNT_ROW + ids.index(target)

    mapped = sheet.cell(row=row, column=LABELS_COL).value
    assert mapped.splitlines() == [
        "Rooms - Total Salaries: 100",
        "Less:",
        "Rooms - Contract Labor: 20",
    ]
    assert "Current YTD" not in mapped
    assert "row 14" not in mapped
    assert sheet.cell(row=row, column=VENUE_COL).value == "La Loba"


def test_direct_mapping_with_multiple_rows_still_displays_add_math(tmp_path):
    ids = canonical_ids()
    target = "S12.total_rooms_revenue"
    result = build_result(
        coa={i: {} for i in ids},
        decisions=[
            SimpleNamespace(
                coa_id=target,
                operation="direct",
                source_rows=["Summary!10", "Summary!11"],
                excluded_rows=[],
                venue_name=None,
            )
        ],
        evidence=[
            {"row_key": "Summary!10", "label": "Rooms", "selected_value": 100},
            {"row_key": "Summary!11", "label": "Resort Fees", "selected_value": 20},
        ],
    )

    sheet = load_workbook(write_normalized_workbook(result, tmp_path / "o.xlsx"))["COA"]
    row = FIRST_ACCOUNT_ROW + ids.index(target)
    assert sheet.cell(row=row, column=LABELS_COL).value.splitlines() == [
        "Add:",
        "Summary - Rooms: 100",
        "Summary - Resort Fees: 20",
    ]


@pytest.mark.parametrize(
    ("target", "value", "expected"),
    [
        ("S12.occupancy", 0.7094, "Summary - Source label: 70.9%"),
        ("S12.adr", 281.534, "Summary - Source label: $281.53"),
        ("S12.revpar", 199.584, "Summary - Source label: 200"),
        ("S12.total_revenue", -1234.56, "Summary - Source label: (1,235)"),
    ],
)
def test_mapped_label_amount_formatting(tmp_path, target, value, expected):
    ids = canonical_ids()
    result = build_result(
        coa={i: {} for i in ids},
        values={target: value},
        decisions=[
            SimpleNamespace(
                coa_id=target,
                operation="direct",
                source_rows=["Summary!10"],
                excluded_rows=[],
                venue_name=None,
            )
        ],
        evidence=[
            {
                "row_key": "Summary!10",
                "label": "Source label",
                "selected_value": value,
            }
        ],
    )

    sheet = load_workbook(write_normalized_workbook(result, tmp_path / "o.xlsx"))["COA"]
    row = FIRST_ACCOUNT_ROW + ids.index(target)
    assert expected in sheet.cell(row=row, column=LABELS_COL).value


def test_unnamed_venue_slots_still_get_a_label(tmp_path):
    """A blank venue label reads as a broken model, not an unused venue."""
    ids = canonical_ids()
    result = build_result(coa={i: {} for i in ids})

    sheet = load_workbook(write_normalized_workbook(result, tmp_path / "o.xlsx"))["COA"]
    row = FIRST_ACCOUNT_ROW + ids.index("S2.venue1_food_revenue")

    assert sheet.cell(row=row, column=VENUE_COL).value == "Venue 1"


def test_validation_checks_reach_the_feedback_column(tmp_path):
    """The old reader parsed checks as dicts and silently produced nothing."""
    ids = canonical_ids()
    target = "S12.total_revenue"
    result = build_result(
        coa={i: {} for i in ids},
        checks=[f"error|hierarchy_complete|{target}|parent=100.00|children=90.00"],
        accepted=False,
    )

    sheet = load_workbook(write_normalized_workbook(result, tmp_path / "o.xlsx"))["COA"]
    note = sheet.cell(row=FIRST_ACCOUNT_ROW + ids.index(target), column=FEEDBACK_COL).value

    assert "Rollup warning: Child accounts are 10 less than the parent account." in note


def test_review_items_attach_to_their_accounts(tmp_path):
    ids = canonical_ids()
    target = ids[3]
    result = build_result(
        coa={i: {} for i in ids},
        review_items=[
            {
                "kind": "unusual_convention",
                "message": "Contract labor sits inside the salary subtotal.",
                "coa_ids": [target],
                "source_rows": ["Rooms!14"],
            }
        ],
    )

    book = load_workbook(write_normalized_workbook(result, tmp_path / "o.xlsx"))
    sheet = book["COA"]
    note = sheet.cell(row=FIRST_ACCOUNT_ROW + ids.index(target), column=FEEDBACK_COL).value

    assert note == "Mapping treatment: Contract labor sits inside the salary subtotal."
    assert "Contract labor sits inside the salary subtotal." not in book["Run Notes"]["C9"].value


def test_review_item_is_displayed_once_on_summary_account(tmp_path):
    ids = canonical_ids()
    summary = "S12.total_property_operation_and_maintenance_expenses"
    detail = "S8.total_property_operation_and_maintenance_expenses"
    message = "A source-only adjustment was retained in Summary POM."
    result = build_result(
        coa={i: {} for i in ids},
        review_items=[
            {
                "kind": "unusual_convention",
                "message": message,
                "coa_ids": [detail, summary],
                "source_rows": ["Summary!40"],
            }
        ],
    )

    book = load_workbook(write_normalized_workbook(result, tmp_path / "o.xlsx"))
    sheet = book["COA"]
    summary_note = sheet.cell(
        row=FIRST_ACCOUNT_ROW + ids.index(summary), column=FEEDBACK_COL
    ).value
    detail_note = sheet.cell(
        row=FIRST_ACCOUNT_ROW + ids.index(detail), column=FEEDBACK_COL
    ).value

    assert message not in book["Run Notes"]["C9"].value
    assert message in (summary_note or "")
    assert message not in (detail_note or "")


def test_review_items_hide_internal_ids_but_keep_readable_source_rows(tmp_path):
    ids = canonical_ids()
    target = "S12.total_sales_and_marketing_expenses"
    result = build_result(
        coa={
            i: {"account_name": i.split(".", 1)[-1].replace("_", " ")}
            for i in ids
        },
        review_items=[
            {
                "kind": "unusual_convention",
                "message": (
                    "Franchise Fees (Rooms!40) were moved to "
                    "S12.total_sales_and_marketing_expenses; the source was no_value."
                ),
                "coa_ids": [target],
                "source_rows": ["Rooms!40"],
            }
        ],
    )
    book = load_workbook(write_normalized_workbook(result, tmp_path / "o.xlsx"))
    sheet = book["COA"]
    note = sheet.cell(row=FIRST_ACCOUNT_ROW + ids.index(target), column=FEEDBACK_COL).value
    run_note = book["Run Notes"]["C9"].value

    assert "Rooms row 40" in note
    assert "S12." not in note
    assert "no_value" not in note
    assert "left blank" in note
    assert "Franchise Fees" not in run_note


def test_per_period_checks_name_their_period(tmp_path):
    ids = canonical_ids()
    target = "S12.total_revenue"
    result = build_result(
        coa={i: {} for i in ids},
        period_labels={"cur": "2025 YTD", "pri": "2024 YTD"},
        period_values={"cur": {}, "pri": {}},
        checks_by_period={
            "pri": [
                f"warning|source_detail_incomplete|{target}|partial|"
                "parent=100.00|children=80.00|variance=20.00"
            ]
        },
    )

    sheet = load_workbook(write_normalized_workbook(result, tmp_path / "o.xlsx"))["COA"]
    note = sheet.cell(row=FIRST_ACCOUNT_ROW + ids.index(target), column=FEEDBACK_COL).value

    assert note.startswith("2024 YTD — ")
    assert "Rollup warning: Child accounts are 20 less than the parent account." in note


def test_repeated_period_rollups_are_combined_into_one_feedback_line(tmp_path):
    ids = canonical_ids()
    target = "S1.total_rooms_expenses"
    result = build_result(
        coa={i: {} for i in ids},
        period_labels={"cur": "2025 Actual", "pri": "2024 Actual"},
        period_values={"cur": {}, "pri": {}},
        checks_by_period={
            "cur": [
                f"warning|source_detail_incomplete|{target}|"
                "parent=100.00|children=80.00|variance=20.00"
            ],
            "pri": [
                f"warning|source_detail_incomplete|{target}|"
                "parent=90.00|children=75.00|variance=15.00"
            ],
        },
    )

    sheet = load_workbook(
        write_normalized_workbook(result, tmp_path / "o.xlsx")
    )["COA"]
    note = sheet.cell(
        row=FIRST_ACCOUNT_ROW + ids.index(target), column=FEEDBACK_COL
    ).value

    assert note == (
        "Coverage gap: child accounts versus parent — "
        "2025 Actual: 20 below; 2024 Actual: 15 below."
    )


def test_missing_decisions_are_summarized_in_run_notes_not_every_account(tmp_path):
    ids = canonical_ids()
    target = ids[0]
    result = build_result(
        coa={i: {} for i in ids},
        decisions=[
            SimpleNamespace(
                coa_id=target,
                operation="no_value",
                source_rows=[],
                excluded_rows=[],
                venue_name=None,
            )
        ],
        accepted=False,
    )

    book = load_workbook(write_normalized_workbook(result, tmp_path / "o.xlsx"))
    feedback = [
        book["COA"].cell(row=FIRST_ACCOUNT_ROW + index, column=FEEDBACK_COL).value
        for index in range(len(ids))
    ]

    assert all("No source found" not in str(value or "") for value in feedback)
    assert "COA accounts have no submitted mapping decision" in book["Run Notes"]["C9"].value


def test_residual_plug_is_shown_in_mapped_labels(tmp_path):
    ids = canonical_ids()
    target = "S4.other_misc_income"
    result = build_result(
        coa={i: {} for i in ids},
        values={target: 8.0},
        decisions=[
            SimpleNamespace(
                coa_id=target,
                operation="no_value",
                source_rows=[],
                excluded_rows=[],
                venue_name=None,
                rationale="",
            )
        ],
        residual_plugs_by_period={"selected": {target: 8.0}},
    )

    sheet = load_workbook(write_normalized_workbook(result, tmp_path / "o.xlsx"))["COA"]
    mapped = sheet.cell(
        row=FIRST_ACCOUNT_ROW + ids.index(target), column=LABELS_COL
    ).value

    assert "Residual plug: 8" in mapped


def test_large_residual_plug_warning_is_human_readable(tmp_path):
    ids = canonical_ids()
    parent = "S4.total_miscellaneous_income"
    result = build_result(
        coa={i: {} for i in ids},
        period_labels={"selected": "Current YTD"},
        period_values={"selected": {}},
        checks_by_period={
            "selected": [
                f"warning|large_residual_plug|{parent}|"
                "residual_id=S4.other_misc_income|plug=120.00|ratio=0.0750"
            ]
        },
    )

    sheet = load_workbook(write_normalized_workbook(result, tmp_path / "o.xlsx"))["COA"]
    note = sheet.cell(
        row=FIRST_ACCOUNT_ROW + ids.index(parent), column=FEEDBACK_COL
    ).value

    assert "120" in note
    assert "8%" in note
    assert "all-other account" in note


@pytest.mark.parametrize(
    ("actual", "expected", "message"),
    [
        (
            1_872_364.54,
            1_441_814.20,
            "Needs review: Summary S&M expenses exceed the detailed S&M schedule by 430,550.",
        ),
        (
            5_143_848.14,
            5_150_329.75,
            "Needs review: Summary A&G expenses are below the detailed A&G schedule by 6,482.",
        ),
        (
            11_822_120.13,
            11_786_544.37,
            "Needs review: Summary Total Departmental Expenses exceed the detailed combined Rooms, F&B, and OOD department schedule by 35,576.",
        ),
    ],
)
def test_summary_department_feedback_shows_direction_and_variance(
    tmp_path, actual, expected, message
):
    ids = canonical_ids()
    target = (
        "S12.total_sales_and_marketing_expenses"
        if "S&M" in message
        else (
            "S12.total_departmental_expenses"
            if "Total Departmental" in message
            else "S12.total_administrative_and_general_expenses"
        )
    )
    result = build_result(
        coa={i: {} for i in ids},
        checks=[
            f"error|summary_department|{target}|actual={actual:.2f}|"
            f"expected={expected:.2f}|variance={actual - expected:.2f}"
        ],
    )

    sheet = load_workbook(write_normalized_workbook(result, tmp_path / "o.xlsx"))["COA"]
    note = sheet.cell(
        row=FIRST_ACCOUNT_ROW + ids.index(target), column=FEEDBACK_COL
    ).value

    assert note.splitlines()[0] == message


def test_summary_math_feedback_names_equation_and_variance(tmp_path):
    ids = canonical_ids()
    target = "S12.total_departmental_expenses"
    result = build_result(
        coa={i: {} for i in ids},
        checks=[
            f"error|summary_math|{target}|actual=11822120.13|"
            "expected=11786544.37|variance=35575.76|"
            "equation=S12.total_rooms_expenses + "
            "S12.total_food_and_beverage_expenses + "
            "S12.total_other_operated_departments_expenses"
        ],
    )

    sheet = load_workbook(write_normalized_workbook(result, tmp_path / "o.xlsx"))["COA"]
    note = sheet.cell(
        row=FIRST_ACCOUNT_ROW + ids.index(target), column=FEEDBACK_COL
    ).value

    assert note.splitlines()[0] == (
        "Needs review: Summary Total Departmental Expenses are 35,576 greater "
        "than Rooms, F&B, and OOD expenses combined."
    )


def test_all_model_feedback_numbers_are_rounded_to_whole_numbers(tmp_path):
    ids = canonical_ids()
    target = "S12.total_revenue"
    result = build_result(
        coa={i: {} for i in ids},
        checks=[
            f"warning|small_source_reconciliation_difference|{target}|"
            "variance=1652.03"
        ],
        review_items=[
            {
                "kind": "unusual_convention",
                "message": "Adjustment of -713,193.41 represents 7.5% of the source.",
                "coa_ids": [target],
                "source_rows": [],
            }
        ],
    )

    book = load_workbook(write_normalized_workbook(result, tmp_path / "o.xlsx"))
    sheet = book["COA"]
    note = sheet.cell(
        row=FIRST_ACCOUNT_ROW + ids.index(target), column=FEEDBACK_COL
    ).value

    assert "1,652" in note
    assert "$" not in note
    assert "-713,193" in note
    assert "8%" in note
    assert ".03" not in note
    assert ".41" not in note
    assert "7.5%" not in note


def test_run_notes_lists_final_rollup_mismatches_over_ten(tmp_path):
    ids = canonical_ids()
    hierarchy_target = "S2.total_food_and_beverage_expenses"
    result = build_result(
        coa={
            coa_id: {
                "account_name": (
                    "Total F&B Outlet Expense"
                    if coa_id == hierarchy_target
                    else coa_id
                )
            }
            for coa_id in ids
        },
        checks=[
            "error|summary_department|S12.total_rooms_expenses|"
            "actual=120.00|expected=100.00|variance=20.00",
            "error|summary_math|S12.total_departmental_expenses|"
            "actual=140.00|expected=100.00|variance=40.00",
            f"error|hierarchy_complete|{hierarchy_target}|"
            "parent=150.00|children=125.00|variance=25.00",
            "error|hierarchy_complete|S1.total_rooms_revenue|"
            "parent=100.00|children=90.00|variance=10.00",
        ],
    )

    notes = load_workbook(
        write_normalized_workbook(result, tmp_path / "o.xlsx")
    )["Run Notes"]

    assert notes["B9"].value == "Notes"
    assert notes["C9"].value.splitlines() == [
        "1 summary math error",
        "1 summary-to-department error",
        "1 material rollup warning",
    ]
    assert notes["C9"].font.sz == 11
    assert notes["C9"].alignment.wrap_text is True
    assert notes.row_dimensions[9].height > 14.5


def test_run_notes_only_surfaces_rollup_mismatches(tmp_path):
    """Unrelated validation and execution detail stays in COA feedback."""
    ids = canonical_ids()
    result = build_result(
        coa={i: {} for i in ids},
        checks=["error|source_row_double_count|Rooms!14|used by two accounts"],
        execution_issues=["Sheet 'Budget' could not be read."],
    )

    book = load_workbook(write_normalized_workbook(result, tmp_path / "o.xlsx"))
    assert book.sheetnames[0] == "Run Notes"
    notes = book["Run Notes"]
    assert notes["B9"].value == "Notes"
    assert notes["C9"].value.splitlines() == [
        "No summary math errors",
        "No summary-to-department errors",
        "No material rollup warnings",
    ]
    assert all(
        cell.value in (None, "")
        for row in notes.iter_rows(min_row=10)
        for cell in row
    )


def test_run_notes_uses_the_approved_template_format(tmp_path):
    ids = canonical_ids()
    book = load_workbook(
        write_normalized_workbook(
            build_result(coa={i: {} for i in ids}), tmp_path / "o.xlsx"
        )
    )
    notes = book["Run Notes"]

    assert notes["B2"].value == "RUN NOTES"
    assert notes["B4"].value == "Source file"
    assert notes["B5"].value == "Periods"
    assert notes["B6"].value == "Status"
    assert notes["C6"].value == "Completed"
    assert notes["B7"].value == "Accounts mapped"
    assert notes["C7"].value == 0
    assert notes["C7"].number_format == "0"
    assert notes["C7"].alignment.horizontal == "left"
    assert notes["B8"].value == "Calculation policy"
    assert notes.row_dimensions[8].height == 30.0
    assert notes["B9"].value == "Notes"
    assert notes["C9"].value.splitlines() == [
        "No summary math errors",
        "No summary-to-department errors",
        "No material rollup warnings",
    ]
    assert "Accounts populated" not in {
        cell.value for row in notes.iter_rows() for cell in row
    }
    assert notes["B2"].fill.fgColor.rgb == "FF39617A"
    assert notes["C2"].fill.fgColor.rgb == "FF39617A"
    assert notes["C4"].font.name == "Calibri"
    assert notes["C4"].font.sz == 11
    assert all(
        notes.cell(row=row, column=column).alignment.vertical == "bottom"
        for row in range(4, 10)
        for column in (2, 3)
    )
    assert notes.column_dimensions["B"].width == 28.0
    assert notes.column_dimensions["C"].width == 95.0
    assert book["COA"].column_dimensions["D"].hidden is True
    assert book["COA"].column_dimensions["V"].hidden is True
    assert book["COA"].column_dimensions["Z"].hidden is False
    assert book["KHP Model Accounts"].column_dimensions["D"].hidden is True
    assert book["KHP Model Accounts"].column_dimensions["V"].hidden is True
    assert book["KHP Model Accounts"].column_dimensions["W"].hidden is False


def test_mapped_output_name_uses_source_stem():
    assert (
        mapped_output_name("Mission Inn Riverside - 2024 Operating Statements.xlsx")
        == "Mission Inn Riverside - 2024 Operating Statements [MAPPED].xlsx"
    )


def test_mapped_output_name_shortens_long_source_stem():
    source = (
        "SFOJD KABUKI--2026-FORECAST-12-Mos-PL Statement_Jan-May Actuals "
        "and Jun-Dec Forecast_06.23.2026.xlsx"
    )

    name = mapped_output_name(source)

    assert name == (
        "SFOJD KABUKI--2026-FORECAST-12-Mos-PL Statement_Jan-May Actuals [MAPPED].xlsx"
    )
    assert len(name) <= 80


def test_run_notes_summarizes_final_validation_and_review_counts(tmp_path):
    ids = canonical_ids()
    target = "S12.total_revenue"
    result = build_result(
        coa={
            coa_id: {"account_name": "Total Revenue" if coa_id == target else coa_id}
            for coa_id in ids
        },
        values={target: 100.0},
        checks=[
            f"error|hierarchy_complete|{target}|parent=100.00|children=90.00",
            f"warning|source_detail_incomplete|{target}|partial",
        ],
        review_items=[
            {
                "kind": "ambiguity",
                "message": "One source convention needs human confirmation.",
                "coa_ids": [target],
                "source_rows": [],
            }
        ],
        accepted=False,
    )
    notes = load_workbook(
        write_normalized_workbook(result, tmp_path / "o.xlsx")
    )["Run Notes"]

    assert notes["C6"].value == "Rejected"
    assert notes["C7"].value == 1
    assert notes["B9"].value == "Notes"
    assert notes["C9"].value.splitlines() == [
        "No summary math errors",
        "No summary-to-department errors",
        "No material rollup warnings",
        "Mapping incomplete: 269 COA accounts have no submitted mapping decision.",
    ]


@pytest.mark.parametrize(
    "overrides",
    [
        {"checks": ["warning|coverage_unspecified|S12.total_revenue|review"]},
        {
            "review_items": [
                {
                    "kind": "ambiguity",
                    "message": "One source convention needs confirmation.",
                    "coa_ids": ["S12.total_revenue"],
                    "source_rows": [],
                }
            ]
        },
    ],
)
def test_run_notes_status_uses_warning_level_without_errors(tmp_path, overrides):
    ids = canonical_ids()
    result = build_result(coa={i: {} for i in ids}, **overrides)

    notes = load_workbook(
        write_normalized_workbook(result, tmp_path / "o.xlsx")
    )["Run Notes"]

    assert notes["C6"].value == "Completed with warnings"


def test_run_notes_status_treats_rollup_mismatch_as_warning(tmp_path):
    ids = canonical_ids()
    target = "S2.total_food_and_beverage_expenses"
    result = build_result(
        coa={i: {} for i in ids},
        checks=[
            f"warning|source_detail_incomplete|{target}|"
            "parent=150.00|children=125.00|variance=25.00"
        ],
    )

    notes = load_workbook(
        write_normalized_workbook(result, tmp_path / "o.xlsx")
    )["Run Notes"]

    assert notes["C6"].value == "Completed with warnings"
    assert "1 material rollup warning" in notes["C9"].value


def test_run_notes_status_preserves_stopped_state(tmp_path):
    ids = canonical_ids()
    result = build_result(
        coa={i: {} for i in ids},
        checks=["error|summary_math|S12.total_revenue|variance=100.00"],
        stopped_reason="Stopped by user",
    )

    notes = load_workbook(
        write_normalized_workbook(result, tmp_path / "o.xlsx")
    )["Run Notes"]

    assert notes["C6"].value == "Stopped"


def test_authored_khp_textbox_survives_output_save(tmp_path):
    ids = canonical_ids()
    path = write_normalized_workbook(
        build_result(coa={i: {} for i in ids}), tmp_path / "o.xlsx"
    )

    with zipfile.ZipFile(path) as archive:
        names = set(archive.namelist())
        assert "xl/drawings/drawing1.xml" in names
        assert "xl/worksheets/_rels/sheet3.xml.rels" in names
        sheet_xml = archive.read("xl/worksheets/sheet3.xml")
        rels_xml = archive.read("xl/worksheets/_rels/sheet3.xml.rels")
        sheet_drawing_id = re.search(
            rb'drawing\b[^>]*r:id="([^"]+)"', sheet_xml
        )
        relationship_id = re.search(
            rb'<Relationship\b(?=[^>]*\bType="[^"]*/drawing")'
            rb'[^>]*\bId="([^"]+)"',
            rels_xml,
        )
        assert sheet_drawing_id is not None
        assert relationship_id is not None
        assert sheet_drawing_id.group(1) == relationship_id.group(1)
        assert b"TextBox 1" in archive.read("xl/drawings/drawing1.xml")


def test_output_freeze_panes(tmp_path):
    ids = canonical_ids()
    book = load_workbook(
        write_normalized_workbook(
            build_result(coa={i: {} for i in ids}), tmp_path / "o.xlsx"
        )
    )
    template = load_workbook(TEMPLATE)
    assert book.active.title == "Run Notes"

    for sheet in book.worksheets:
        assert sheet.sheet_view.topLeftCell == "A1"
        assert len(sheet.sheet_view.selection) == 1
        selection = sheet.sheet_view.selection[0]
        assert selection.activeCell == "A1"
        assert selection.sqref == "A1"

    for name in ("COA", "KHP Model Accounts"):
        pane = book[name].sheet_view.pane
        template_pane = template[name].sheet_view.pane
        assert pane.xSplit == template_pane.xSplit
        assert pane.ySplit == template_pane.ySplit
        assert pane.topLeftCell == (
            f"{get_column_letter(int(pane.xSplit) + 1)}{int(pane.ySplit) + 1}"
        )
        assert pane.activePane == "topLeft"
        assert book[name].sheet_view.selection[0].pane == "topLeft"


@pytest.mark.parametrize(
    ("labels", "expected_amount", "expected_period"),
    [
        (
            {"fy24": "2024 Actual", "fy25f": "2025 Forecast", "ytd": "Current YTD"},
            200,
            "2025 Forecast",
        ),
        (
            {"actual": "2025 Actual", "forecast": "2025 Forecast", "budget": "2025 Budget"},
            100,
            "2025 Actual",
        ),
        (
            {"mtd": "Current MTD", "ytd": "Current YTD", "month": "Jan 2025 Actual"},
            200,
            "Current YTD",
        ),
    ],
)
def test_mapped_labels_choose_one_highest_priority_period(
    tmp_path, labels, expected_amount, expected_period
):
    ids = canonical_ids()
    target = ids[0]
    period_values = {
        period_id: {target: float(index + 1)}
        for index, period_id in enumerate(labels)
    }
    evidence_values = {
        period_id: float((index + 1) * 100)
        for index, period_id in enumerate(labels)
    }
    result = build_result(
        coa={i: {} for i in ids},
        period_labels=labels,
        period_values=period_values,
        decisions=[
            SimpleNamespace(
                coa_id=target,
                operation="direct",
                source_rows=["Summary!10"],
                excluded_rows=[],
                venue_name=None,
            )
        ],
        evidence=[
            {
                "row_key": "Summary!10",
                "label": "Rooms Available",
                "selected_values": evidence_values,
            }
        ],
    )
    sheet = load_workbook(write_normalized_workbook(result, tmp_path / "o.xlsx"))["COA"]
    mapped = sheet.cell(row=FIRST_ACCOUNT_ROW, column=LABELS_COL).value

    assert mapped == f"Summary - Rooms Available: {expected_amount:,}"
    assert all(label not in mapped for label in labels.values())
    assert sheet.cell(row=HEADER_ROW, column=LABELS_COL).value == (
        f"Mapped Labels - {expected_period}"
    )


def test_wrapped_output_sets_row_height_and_keeps_metadata_visible(tmp_path):
    ids = canonical_ids()
    target = ids[0]
    result = build_result(
        coa={i: {} for i in ids},
        decisions=[
            SimpleNamespace(
                coa_id=target,
                operation="sum",
                source_rows=["Rooms!10", "Rooms!11", "Rooms!12"],
                excluded_rows=[],
                venue_name=None,
            )
        ],
        evidence=[
            {"row_key": f"Rooms!{row}", "label": "A long mapped operator account label", "selected_value": row}
            for row in (10, 11, 12)
        ],
        checks=[f"warning|source_detail_incomplete|{target}|partial"],
    )
    sheet = load_workbook(write_normalized_workbook(result, tmp_path / "o.xlsx"))["COA"]
    row = FIRST_ACCOUNT_ROW
    assert sheet.row_dimensions[row].height > 18
    assert sheet.cell(row=row, column=LABELS_COL).alignment.wrap_text is True
    assert sheet.cell(row=row, column=LABELS_COL).alignment.vertical == "bottom"
    assert sheet.cell(row=row, column=FEEDBACK_COL).alignment.vertical == "bottom"
    assert all(
        sheet.column_dimensions[get_column_letter(column)].hidden is False
        for column in (ID_COL, LABELS_COL, FEEDBACK_COL, VENUE_COL)
    )


def test_conditional_formatting_spans_the_full_coa_table(tmp_path):
    ids = canonical_ids()
    result = build_result(coa={i: {} for i in ids})
    sheet = load_workbook(write_normalized_workbook(result, tmp_path / "o.xlsx"))["COA"]

    assert [str(block.sqref) for block in sheet.conditional_formatting] == [
        f"B{FIRST_ACCOUNT_ROW}:Y{FIRST_ACCOUNT_ROW + len(ids) - 1}"
    ]


def test_drifted_template_is_refused_rather_than_written(tmp_path):
    """A silent row shift is the one failure that produces a plausible model."""
    ids = canonical_ids()
    result = build_result(coa={i: {} for i in ids})

    import hotel_pl_normalizer.output as output

    with pytest.raises(OutputTemplateError, match="Rebuild"):
        original = output._canonical_coa_ids
        try:
            output._canonical_coa_ids = lambda: ["S99.invented", *ids]
            write_normalized_workbook(result, tmp_path / "o.xlsx")
        finally:
            output._canonical_coa_ids = original
